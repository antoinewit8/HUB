"""
6____Tractionnaires_KM.py
──────────────────────────────────────────────────────────────────
Outil TX-FLEX : Analyse Tractionnaires — KM PTV + CA + Rentabilité
──────────────────────────────────────────────────────────────────
Entrée : Export tractionnaires (.xlsx)
  Colonnes : Tractionnaire, Chauffeur, Véhicule, Remorque, Dossier,
             Référence, Type transport, CMR, Date chargement,
             CP chargement, Localité chargement, Pays chargement,
             Date déchargement, CP déchargement, Localité déchargement,
             Pays déchargement, Statut facturation, Ventes totales,
             Département vente, Client

Sorties :
  • Tableau détail dossiers avec KM PTV estimés
  • Résumé par tractionnaire : dossiers, KM, CA, rentabilité
  • Export Excel
──────────────────────────────────────────────────────────────────
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import os
import io
import re
import datetime as _dt
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════════════════════════════
#  CONFIG PTV
# ══════════════════════════════════════════════════════════════════

PTV_API_KEY  = os.environ.get("PTV_API_KEY", "METS_TA_CLE_ICI")
PTV_BASE_URL = "https://api.myptv.com/routing/v1"
GEOCODE_URL  = "https://api.myptv.com/geocoding/v1"
HEADERS      = {"apiKey": PTV_API_KEY}
MAX_RETRIES  = 3
RETRY_DELAY  = 2
VEHICLE      = "EUR_TRAILER_TRUCK"

PAYS_MAP = {
    "F": "France", "B": "Belgium", "D": "Germany", "L": "Luxembourg",
    "NL": "Netherlands", "E": "Spain", "I": "Italy", "CH": "Switzerland",
    "GB": "United Kingdom", "A": "Austria", "P": "Portugal",
    "FR": "France", "BE": "Belgium", "DE": "Germany", "LU": "Luxembourg",
    "IT": "Italy", "ES": "Spain", "AT": "Austria", "PT": "Portugal",
}

PAYS_TO_ISO2 = {
    "F": "FR", "B": "BE", "D": "DE", "L": "LU", "I": "IT",
    "E": "ES", "A": "AT", "P": "PT", "CH": "CH", "GB": "GB",
    "NL": "NL", "FR": "FR", "BE": "BE", "DE": "DE", "LU": "LU",
    "IT": "IT", "ES": "ES", "AT": "AT", "PT": "PT",
}

MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

# ══════════════════════════════════════════════════════════════════
#  UTILITAIRES
# ══════════════════════════════════════════════════════════════════

def _norm_col(s: str) -> str:
    s = str(s).strip().lower()
    for src, dst in [("é","e"),("è","e"),("ê","e"),("à","a"),("â","a"),
                     ("ô","o"),("û","u"),("î","i"),("ù","u"),("ç","c")]:
        s = s.replace(src, dst)
    return re.sub(r"[^a-z0-9]", "", s)


def _find_col(df: pd.DataFrame, candidates: list) -> str | None:
    cols_lower = {_norm_col(c): c for c in df.columns}
    for cand in candidates:
        key = _norm_col(cand)
        if key in cols_lower:
            return cols_lower[key]
    return None


def _clean(v) -> str:
    v = str(v or "").strip()
    return "" if v.lower() in ("nan", "none") else v


def _to_float(s) -> float:
    try:
        return float(str(s).replace(",", ".").replace("\u00a0", "")
                     .replace(" ", "").replace("€", "").strip())
    except Exception:
        return 0.0


# ══════════════════════════════════════════════════════════════════
#  GEOCODAGE
# ══════════════════════════════════════════════════════════════════

def _ptv_by_text(query: str) -> tuple | None:
    if not query:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-text",
                params={"searchText": query},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt); continue
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def _ptv_by_postal_code(cp: str, iso2: str) -> tuple | None:
    if not cp or not iso2:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-postal-code",
                params={"postalCode": cp, "countryCode": iso2},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt); continue
            if resp.status_code in (400, 404):
                return None
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def geocode_with_fallback(ville: str, cp: str, pays: str) -> tuple | None:
    pays_full = PAYS_MAP.get(pays.upper(), pays) if pays else ""
    iso2      = PAYS_TO_ISO2.get(pays.upper(), pays.upper() if len(pays) == 2 else "")

    if ville and cp and pays_full:
        r = _ptv_by_text(f"{ville}, {cp}, {pays_full}")
        if r: return r
    if cp and iso2:
        r = _ptv_by_postal_code(cp, iso2)
        if r: return r
    if ville and pays_full:
        r = _ptv_by_text(f"{ville}, {pays_full}")
        if r: return r
    return None


# ══════════════════════════════════════════════════════════════════
#  CALCUL ROUTE PTV
# ══════════════════════════════════════════════════════════════════

def calculate_route(coords_list: list) -> dict | None:
    if len(coords_list) < 2:
        return None
    query_params = [("profile", VEHICLE), ("results", "POLYLINE")]
    for lat, lon in coords_list:
        query_params.append(("waypoints", f"{lat},{lon}"))
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{PTV_BASE_URL}/routes",
                headers=HEADERS, params=query_params, timeout=30,
            )
            if resp.status_code != 200:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY); continue
                return None
            data = resp.json()
            return {"km": round(data.get("distance", 0) / 1000, 1)}
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


# ══════════════════════════════════════════════════════════════════
#  PARSING
# ══════════════════════════════════════════════════════════════════

# Noms de colonnes exacts de l'export tractionnaires
TRACT_COL_CANDIDATES = {
    "tractionnaire":   ["Tractionnaire"],
    "chauffeur":       ["Chauffeur"],
    "vehicule":        ["Véhicule", "Vehicule"],
    "remorque":        ["Remorque"],
    "dossier":         ["Dossier", "N° Dossier"],
    "reference":       ["Référence", "Reference"],
    "type_transport":  ["Type de transport", "Type transport"],
    "cmr":             ["CMR"],
    "date_charg":      ["Date chargement", "Date Chargement"],
    "cp_charg":        ["C.P. chargement", "CP chargement"],
    "localite_charg":  ["Localité chargement", "Localite chargement"],
    "pays_charg":      ["Pays chargement"],
    "date_decharg":    ["Date déchargement", "Date dechargement"],
    "cp_decharg":      ["C.P. déchargement", "CP dechargement"],
    "localite_decharg":["Localité déchargement", "Localite dechargement"],
    "pays_decharg":    ["Pays déchargement", "Pays dechargement"],
    "statut":          ["Statut facturation", "Statut"],
    "ventes_totales":  ["Ventes totales", "Total ventes", "CA"],
    "dept_vente":      ["Département vente", "Dept vente"],
    "client":          ["Client"],
}


def parse_tractionnaires(file) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cols_lower = {_norm_col(c): c for c in df.columns}
    col_map = {}
    for role, candidates in TRACT_COL_CANDIDATES.items():
        found = None
        for cand in candidates:
            key = _norm_col(cand)
            if key in cols_lower:
                found = cols_lower[key]
                break
        col_map[role] = found

    # Avertissement colonnes critiques manquantes
    critiques = ["tractionnaire", "dossier", "localite_charg", "localite_decharg"]
    manquantes = [r for r in critiques if col_map.get(r) is None]
    if manquantes:
        st.warning(f"⚠️ Colonnes non détectées : {manquantes} — Colonnes dispo : {list(df.columns)}")

    rename = {v: k for k, v in col_map.items() if v}
    df = df.rename(columns=rename)
    for col in TRACT_COL_CANDIDATES:
        if col not in df.columns:
            df[col] = ""

    # Nettoyage
    df["dossier"] = df["dossier"].str.strip()
    df = df[df["dossier"].notna() & (df["dossier"] != "") & (df["dossier"] != "nan")]
    df = df[df["dossier"].str.match(r"^\d+", na=False)]

    df["ventes_totales"] = df["ventes_totales"].apply(_to_float)

    # Nettoyage tractionnaire/chauffeur
    for col in ["tractionnaire", "chauffeur", "vehicule", "remorque",
                "localite_charg", "cp_charg", "pays_charg",
                "localite_decharg", "cp_decharg", "pays_decharg"]:
        df[col] = df[col].apply(_clean)

    # Date chargement pour la période
    df["_date_dt"] = pd.to_datetime(
        df["date_charg"].str.strip().str[:10], format="%Y-%m-%d", errors="coerce"
    )

    return df


# ══════════════════════════════════════════════════════════════════
#  CALCUL KM PTV
# ══════════════════════════════════════════════════════════════════

def compute_km(df: pd.DataFrame, progress_cb=None) -> pd.DataFrame:
    """
    Géocode tous les points uniques (ville, cp, pays) et calcule
    les km PTV pour chaque dossier (chargement → déchargement).
    """
    # Collecter points uniques
    points = set()
    for _, row in df.iterrows():
        if row["localite_charg"] or row["cp_charg"]:
            points.add((_clean(row["localite_charg"]), _clean(row["cp_charg"]), _clean(row["pays_charg"])))
        if row["localite_decharg"] or row["cp_decharg"]:
            points.add((_clean(row["localite_decharg"]), _clean(row["cp_decharg"]), _clean(row["pays_decharg"])))

    # Géocodage
    geo_cache = {}
    total = len(points)
    for i, (ville, cp, pays) in enumerate(points):
        if progress_cb:
            progress_cb(f"🌍 Géocodage {i+1}/{total} : {ville} {cp}...")
        coords = geocode_with_fallback(ville, cp, pays)
        geo_cache[(ville, cp, pays)] = coords
        if coords is None:
            st.warning(f"⚠️ Géocodage échoué : {ville}, {cp}, {pays}")

    # Calcul route par dossier
    km_results = []
    total_dos = len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        if progress_cb:
            progress_cb(f"📍 Calcul KM dossier {row['dossier']} ({i+1}/{total_dos})...")

        c_ch = geo_cache.get((_clean(row["localite_charg"]),  _clean(row["cp_charg"]),  _clean(row["pays_charg"])))
        c_de = geo_cache.get((_clean(row["localite_decharg"]), _clean(row["cp_decharg"]), _clean(row["pays_decharg"])))

        if c_ch and c_de:
            res = calculate_route([c_ch, c_de])
            km_results.append(res["km"] if res else None)
        else:
            km_results.append(None)

    df = df.copy()
    df["km_ptv"] = km_results
    return df


# ══════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════

def export_excel(df_detail: pd.DataFrame, df_resume: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # Feuille détail
            col_rename_detail = {
                "tractionnaire":    "Tractionnaire",
                "chauffeur":        "Chauffeur",
                "vehicule":         "Véhicule",
                "remorque":         "Remorque",
                "dossier":          "N° Dossier",
                "client":           "Client",
                "statut":           "Statut facturation",
                "localite_charg":   "Localité chargement",
                "localite_decharg":  "Localité déchargement",
                "ventes_totales":   "Ventes totales (€)",
                "km_ptv":           "KM PTV estimés",
                "rentabilite":      "Rentabilité €/km",
            }
            df_d = df_detail.copy()
            df_d["rentabilite"] = (
                df_d["ventes_totales"] / df_d["km_ptv"].replace(0, np.nan)
            ).round(2)
            cols = [c for c in col_rename_detail if c in df_d.columns]
            df_d = df_d[cols].rename(columns=col_rename_detail).fillna("")
            df_d.to_excel(writer, sheet_name="Détail dossiers", index=False)
            _style_sheet(writer.sheets["Détail dossiers"], len(df_d))

            # Feuille résumé
            df_resume.to_excel(writer, sheet_name="Résumé tractionnaires", index=False)
            _style_sheet(writer.sheets["Résumé tractionnaires"], len(df_resume))

    except Exception as e:
        st.error(f"❌ Erreur génération Excel : {e}")
        return b""

    return output.getvalue()


def _style_sheet(ws, nb_rows: int):
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
    for cell in ws[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, nb_rows + 2):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


# ══════════════════════════════════════════════════════════════════
#  INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Tractionnaires KM + CA", page_icon="🚛", layout="wide")

st.title("🚛 Analyse Tractionnaires — KM PTV + CA")
st.caption("Calcule les km estimés PTV et analyse le CA par tractionnaire.")

if not PTV_API_KEY or PTV_API_KEY == "METS_TA_CLE_ICI":
    st.error("⚠️ Clé PTV_API_KEY non configurée.")

st.divider()

file_tract = st.file_uploader("📋 Export tractionnaires (.xlsx)", type=["xlsx"])

if file_tract:
    with st.spinner("📂 Lecture du fichier..."):
        try:
            df = parse_tractionnaires(file_tract)
        except Exception as e:
            st.error(f"❌ Erreur lecture : {e}")
            st.stop()

    # ── Période détectée ──────────────────────────────────────
    dates_valides = df["_date_dt"].dropna()
    if not dates_valides.empty:
        d_min, d_max = dates_valides.min(), dates_valides.max()
        if d_min.month == d_max.month and d_min.year == d_max.year:
            periode_label = f"{MOIS_FR[d_min.month]} {d_min.year}"
        else:
            periode_label = f"{d_min.strftime('%d/%m/%Y')} → {d_max.strftime('%d/%m/%Y')}"
    else:
        periode_label = "Période inconnue"

    # ── KPIs globaux ──────────────────────────────────────────
    st.markdown(f"### 📊 Aperçu — {periode_label}")
    ca_total = df["ventes_totales"].sum()
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("📁 Dossiers",       len(df))
    k2.metric("🏢 Tractionnaires", df["tractionnaire"].nunique())
    k3.metric("🚛 Véhicules",      df["vehicule"].nunique())
    k4.metric("💶 CA Total",       f"{ca_total:,.0f} €")
    k5.metric("📈 CA moy/dossier", f"{ca_total / len(df):,.0f} €" if len(df) else "—")

    st.divider()

    # ── Filtres ───────────────────────────────────────────────
    st.markdown("### 📋 Tableau détail")

    tract_dispo = sorted([t for t in df["tractionnaire"].unique() if t and t != "nan"])
    f1, f2 = st.columns(2)
    with f1:
        filtre_tract = st.multiselect("🏢 Tractionnaire :", options=tract_dispo, default=[],
                                       placeholder="Tous les tractionnaires")
    with f2:
        statuts_dispo = sorted([s for s in df["statut"].unique() if s and s != "nan"])
        filtre_statut = st.multiselect("📄 Statut facturation :", options=statuts_dispo,
                                        default=[], placeholder="Tous")

    df_display = df.copy()
    if filtre_tract:
        df_display = df_display[df_display["tractionnaire"].isin(filtre_tract)]
    if filtre_statut:
        df_display = df_display[df_display["statut"].isin(filtre_statut)]

    # KPIs filtre
    if filtre_tract or filtre_statut:
        _ca_f = df_display["ventes_totales"].sum()
        _nd_f = len(df_display)
        fk1, fk2, fk3, fk4 = st.columns(4)
        fk1.metric("📁 Dossiers sélectionnés", _nd_f)
        fk2.metric("💶 CA sélection",           f"{_ca_f:,.0f} €")
        fk3.metric("📈 CA moy/dossier",          f"{_ca_f/_nd_f:,.0f} €" if _nd_f else "—")
        fk4.metric("% du CA total",              f"{_ca_f/ca_total*100:.1f}%" if ca_total else "—")

    # Tableau
    cols_show = {
        "dossier": "N° Dossier", "tractionnaire": "Tractionnaire",
        "chauffeur": "Chauffeur", "vehicule": "Véhicule", "remorque": "Remorque",
        "client": "Client", "statut": "Statut",
        "localite_charg": "Chargement", "localite_decharg": "Déchargement",
        "ventes_totales": "CA (€)",
    }
    df_table = df_display[[c for c in cols_show if c in df_display.columns]].rename(columns=cols_show)
    st.dataframe(df_table, use_container_width=True, height=380)

    st.divider()

    # ── Résumé par tractionnaire (sans KM) ───────────────────
    st.markdown("### 🏢 Résumé par tractionnaire")
    df_resume_base = df_display.groupby("tractionnaire", as_index=False).agg(
        Dossiers      = ("dossier",        "count"),
        Véhicules     = ("vehicule",        pd.Series.nunique),
        CA_Total      = ("ventes_totales",  "sum"),
    ).round(1)
    df_resume_base["CA moy/dossier"] = (df_resume_base["CA_Total"] / df_resume_base["Dossiers"]).round(0)
    df_resume_base = df_resume_base.rename(columns={"CA_Total": "CA Total (€)"})
    df_resume_base = df_resume_base.sort_values("CA Total (€)", ascending=False)
    st.dataframe(df_resume_base, use_container_width=True)

    st.divider()

    # ── Section calcul PTV ───────────────────────────────────
    st.markdown("### 🗺️ Calcul KM via PTV")

    tract_ptv = st.multiselect(
        "🏢 Tractionnaires à calculer :",
        options=tract_dispo, default=[],
        placeholder="Sélectionner des tractionnaires..."
    )

    df_ptv_scope = df[df["tractionnaire"].isin(tract_ptv)] if tract_ptv else pd.DataFrame()

    if tract_ptv:
        st.info(f"ℹ️ {len(df_ptv_scope)} dossiers à calculer pour {len(tract_ptv)} tractionnaire(s).")

    btn_ptv = st.button("🚀 Lancer le calcul PTV", disabled=(not tract_ptv), type="primary")

    if btn_ptv and tract_ptv:
        progress_bar = st.progress(0)
        status_text  = st.empty()

        def _progress(msg):
            status_text.text(msg)

        try:
            df_ptv_result = compute_km(df_ptv_scope, progress_cb=_progress)
            st.session_state["df_ptv_result"] = df_ptv_result
            status_text.success("✅ Calcul PTV terminé !")
            progress_bar.progress(100)
        except Exception as e:
            st.error(f"❌ Erreur calcul PTV : {e}")

    # ── Résultats PTV ─────────────────────────────────────────
    if "df_ptv_result" in st.session_state:
        df_r = st.session_state["df_ptv_result"]

        km_sum  = df_r["km_ptv"].sum()
        ca_sum  = df_r["ventes_totales"].sum()
        rent    = ca_sum / km_sum if km_sum > 0 else 0
        dos_ok  = df_r["km_ptv"].notna().sum()

        st.divider()
        st.markdown("### 📈 Résultats KM")
        rk1, rk2, rk3, rk4 = st.columns(4)
        rk1.metric("📁 Dossiers calculés",  f"{dos_ok} / {len(df_r)}")
        rk2.metric("📏 KM PTV Total",        f"{km_sum:,.0f} km")
        rk3.metric("💶 CA Total",             f"{ca_sum:,.0f} €")
        rk4.metric("📈 Rentabilité",          f"{rent:.2f} €/km")

        tab1, tab2 = st.tabs(["📋 Détail dossiers", "🏢 Résumé tractionnaires"])

        with tab1:
            df_detail_show = df_r.copy()
            df_detail_show["rentabilite"] = (
                df_detail_show["ventes_totales"] / df_detail_show["km_ptv"].replace(0, np.nan)
            ).round(2)
            cols_det = {
                "dossier": "N° Dossier", "tractionnaire": "Tractionnaire",
                "chauffeur": "Chauffeur", "client": "Client",
                "localite_charg": "Chargement", "localite_decharg": "Déchargement",
                "ventes_totales": "CA (€)", "km_ptv": "KM PTV", "rentabilite": "€/km",
            }
            df_det_tab = df_detail_show[[c for c in cols_det if c in df_detail_show.columns]].rename(columns=cols_det)
            st.dataframe(df_det_tab, use_container_width=True, height=400)

        with tab2:
            df_res_ptv = df_r.groupby("tractionnaire", as_index=False).agg(
                Dossiers    = ("dossier",        "count"),
                KM_PTV      = ("km_ptv",          "sum"),
                CA_Total    = ("ventes_totales",  "sum"),
            ).round(1)
            df_res_ptv["CA moy/dossier"]  = (df_res_ptv["CA_Total"] / df_res_ptv["Dossiers"]).round(0)
            df_res_ptv["KM moy/dossier"]  = (df_res_ptv["KM_PTV"]  / df_res_ptv["Dossiers"]).round(0)
            df_res_ptv["Rentabilité €/km"]= (df_res_ptv["CA_Total"] / df_res_ptv["KM_PTV"].replace(0, np.nan)).round(2)
            df_res_ptv = df_res_ptv.rename(columns={
                "tractionnaire": "Tractionnaire", "KM_PTV": "KM PTV Total",
                "CA_Total": "CA Total (€)"
            }).sort_values("CA Total (€)", ascending=False)
            st.dataframe(df_res_ptv, use_container_width=True)

        # Export
        st.divider()
        df_resume_export = df_res_ptv.copy()
        excel_bytes = export_excel(df_r, df_resume_export)
        if excel_bytes:
            st.download_button(
                label="📥 Télécharger le rapport Excel",
                data=excel_bytes,
                file_name="Rapport_Tractionnaires_KM.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

else:
    st.markdown("""
    #### Comment utiliser cet outil

    1. **Chargez l'export tractionnaires** (.xlsx)
    2. Consultez l'**aperçu global** et le résumé CA par tractionnaire
    3. Filtrez par tractionnaire ou statut de facturation
    4. Sélectionnez les tractionnaires et **lancez le calcul PTV** pour estimer les km
    5. **Téléchargez le rapport Excel**

    > ⚙️ La clé PTV doit être configurée dans `.env` (`PTV_API_KEY`).
    """)
