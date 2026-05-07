"""
6____Renta_Benne.py
──────────────────────────────────────────────────────────────────
Outil Rentabilité Benne
──────────────────────────────────────────────────────────────────
Entrées :
  • Fichier Benne  (.xlsx) — colonnes :
      Dossier | Date charg. | C.P. charg. (ex: "F 91270") |
      Localité charg. | Date Décharg. | C.P. Déharg. | Localité Décharg.

  • Fichier CA     (.xlsx) — colonnes :
      N° Dossier | Total des ventes | Client facturation | Etat vente

Sorties :
  • Tableau consolidé avec CA jointé sur N° Dossier
  • KPIs globaux : CA, km chargés, km à vide, rentabilité €/km
  • Résumé semaine + mois
  • Export Excel
──────────────────────────────────────────────────────────────────
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import os
import io
import re
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════════════════════════════
#  CONFIG PTV
# ══════════════════════════════════════════════════════════════════

PTV_API_KEY  = os.environ.get("PTV_API_KEY", "METS_TA_CLE_ICI")
PTV_BASE_URL = "https://api.myptv.com/routing/v1"
GEOCODE_URL  = "https://api.myptv.com/geocoding/v1"
HEADERS      = {"apiKey": PTV_API_KEY}
MAX_RETRIES  = 3
RETRY_DELAY  = 2
VEHICLE      = "EUR_TRAILER_TRUCK"

PAYS_CODE_TO_ISO2 = {
    "F": "FR", "B": "BE", "D": "DE", "L": "LU", "NL": "NL",
    "E": "ES", "I": "IT", "CH": "CH", "GB": "GB", "A": "AT",
    "P": "PT", "FR": "FR", "BE": "BE", "DE": "DE", "LU": "LU",
}

PAYS_CODE_TO_FULL = {
    "F": "France", "B": "Belgium", "D": "Germany", "L": "Luxembourg",
    "NL": "Netherlands", "E": "Spain", "I": "Italy", "CH": "Switzerland",
    "GB": "United Kingdom", "A": "Austria", "P": "Portugal",
    "FR": "France", "BE": "Belgium", "DE": "Germany", "LU": "Luxembourg",
}


# ══════════════════════════════════════════════════════════════════
#  PARSING CP FORMAT "F 91270" / "B 9000"
# ══════════════════════════════════════════════════════════════════

def parse_cp_pays(val: str):
    """
    Parse le format 'F 91270' ou 'B 9000' → (code_pays, cp_num).
    Retourne ("", "") si impossible.
    """
    val = str(val or "").strip()
    if not val or val.lower() == "nan":
        return "", ""
    m = re.match(r"^([A-Za-z]{1,3})\s+(.+)$", val)
    if m:
        return m.group(1).upper(), m.group(2).strip()
    return "", val


# ══════════════════════════════════════════════════════════════════
#  GEOCODAGE PTV
# ══════════════════════════════════════════════════════════════════

def _ptv_by_postal_code(cp, iso2):
    if not cp or not iso2:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-postal-code",
                params={"postalCode": cp, "countryCode": iso2},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code not in (200,):
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def _ptv_by_text(query):
    if not query:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-text",
                params={"searchText": query},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


@st.cache_data(show_spinner=False)
def geocode_stop(ville: str, cp: str, code_pays: str):
    """Géocode un arrêt. Priorité : CP+ISO → ville+pays → ville seule."""
    iso2      = PAYS_CODE_TO_ISO2.get(code_pays.upper(), "")
    pays_full = PAYS_CODE_TO_FULL.get(code_pays.upper(), "")

    if cp and iso2:
        r = _ptv_by_postal_code(cp, iso2)
        if r:
            return r
    if ville and cp and pays_full:
        r = _ptv_by_text(f"{ville}, {cp}, {pays_full}")
        if r:
            return r
    if ville and pays_full:
        r = _ptv_by_text(f"{ville}, {pays_full}")
        if r:
            return r
    if ville:
        return _ptv_by_text(ville)
    return None


def calculate_route_km(coord_a, coord_b):
    """Distance PTV entre deux points GPS. Retourne km (float) ou None."""
    if not coord_a or not coord_b:
        return None
    query_params = [
        ("profile",   VEHICLE),
        ("waypoints", f"{coord_a[0]},{coord_a[1]}"),
        ("waypoints", f"{coord_b[0]},{coord_b[1]}"),
    ]
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{PTV_BASE_URL}/routes",
                headers=HEADERS, params=query_params, timeout=30,
            )
            if resp.status_code != 200:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return None
            return round(resp.json().get("distance", 0) / 1000, 1)
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


# ══════════════════════════════════════════════════════════════════
#  PARSING FICHIER BENNE
# ══════════════════════════════════════════════════════════════════

def _norm(s):
    s = str(s).lower()
    for a, b in [("é","e"),("è","e"),("ê","e"),("à","a"),("â","a"),
                 ("ô","o"),("û","u"),("î","i"),("ù","u"),("ç","c")]:
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]", "", s)


BENNE_COLS = {
    "dossier":       ["Dossier", "N° Dossier", "N°Dossier"],
    "date_charg":    ["Date charg.", "Date charg", "Date chargement"],
    "cp_charg":      ["C.P. charg.", "CP charg.", "CP chargement", "C.P. charg"],
    "ville_charg":   ["Localité charg.", "Localite charg.", "Ville charg.", "Localité chargement"],
    "date_decharg":  ["Date Décharg.", "Date Decharg.", "Date décharg.", "Date déchargement"],
    "cp_decharg":    ["C.P. Déharg.", "C.P. Décharg.", "CP Décharg.", "CP decharg.", "C.P. decharg"],
    "ville_decharg": ["Localité Décharg.", "Localite Decharg.", "Ville décharg.", "Localité déchargement"],
}


def parse_benne(file):
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cols_lower = {_norm(c): c for c in df.columns}

    col_map = {}
    for role, candidates in BENNE_COLS.items():
        found = None
        for cand in candidates:
            if _norm(cand) in cols_lower:
                found = cols_lower[_norm(cand)]
                break
        if not found:
            for col in df.columns:
                for cand in sorted(candidates, key=len, reverse=True):
                    if _norm(cand) in _norm(col):
                        found = col
                        break
                if found:
                    break
        col_map[role] = found

    manquantes = [r for r in ["dossier", "cp_charg", "ville_charg", "cp_decharg", "ville_decharg"]
                  if col_map.get(r) is None]
    if manquantes:
        st.warning(f"⚠️ Colonnes non détectées : **{manquantes}** — Disponibles : `{list(df.columns)}`")

    df = df.rename(columns={v: k for k, v in col_map.items() if v})

    for col in BENNE_COLS.keys():
        if col not in df.columns:
            df[col] = ""

    def _clean(v):
        v = str(v or "").strip()
        return "" if v.lower() in ("nan", "none") else v

    for col in df.columns:
        df[col] = df[col].apply(_clean)

    df = df[df["dossier"] != ""].reset_index(drop=True)

    import datetime as _dt

    def _parse_date(s):
        """
        Parse les dates benne : ISO '2026-04-01 00:00:00', FR '20/04/2026'.
        Dates < 2000 = cellule vide/corrompue (numéro série Excel) -> NaT.
        """
        s = str(s or "").strip()
        if not s or s.lower() == "nan":
            return pd.NaT
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
            try:
                d = _dt.datetime.strptime(s, fmt)
                if d.year < 2000:
                    return pd.NaT
                return pd.Timestamp(d)
            except Exception:
                pass
        return pd.NaT

    df["date_charg_dt"]   = df["date_charg"].apply(_parse_date)
    df["date_decharg_dt"] = df["date_decharg"].apply(_parse_date)

    # Parse "F 91270" → (pays, cp_num)
    df[["pays_charg",  "cp_charg_num"]]  = df["cp_charg"].apply(
        lambda v: pd.Series(parse_cp_pays(v))
    )
    df[["pays_decharg", "cp_decharg_num"]] = df["cp_decharg"].apply(
        lambda v: pd.Series(parse_cp_pays(v))
    )

    return df


# ══════════════════════════════════════════════════════════════════
#  PARSING FICHIER CA
# ══════════════════════════════════════════════════════════════════

def parse_ca(file):
    """
    Retourne un dict {dossier_str: {total_ventes, client, etat_vente}}.
    Si plusieurs lignes pour un même dossier → somme des Total des ventes.
    """
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    cols_lower = {_norm(c): c for c in df.columns}

    def _find(candidates):
        for cand in candidates:
            if _norm(cand) in cols_lower:
                return cols_lower[_norm(cand)]
        return None

    col_dossier = _find(["N° Dossier", "N°Dossier", "Dossier"])
    col_total   = _find(["Total des ventes", "Total ventes", "Total des vente"])
    col_client  = _find(["Client facturation", "Client Facturation", "Client"])
    col_etat    = _find(["Etat vente", "État vente", "Etat"])

    if not col_dossier or not col_total:
        st.error(f"❌ Colonnes CA manquantes — disponibles : `{list(df.columns)}`")
        return {}

    def to_float(s):
        try:
            return float(str(s).replace(",", ".").replace("\xa0", "")
                         .replace(" ", "").replace("€", "").strip())
        except Exception:
            return 0.0

    ca_dict = {}
    for _, row in df.iterrows():
        dos = str(row[col_dossier] or "").strip()
        if not dos or dos.lower() == "nan":
            continue
        total = to_float(row[col_total])
        if dos in ca_dict:
            ca_dict[dos]["total_ventes"] += total
        else:
            ca_dict[dos] = {
                "total_ventes": total,
                "client":     str(row[col_client] or "").strip() if col_client else "",
                "etat_vente": str(row[col_etat]   or "").strip() if col_etat   else "",
            }
    return ca_dict


# ══════════════════════════════════════════════════════════════════
#  JOINTURE BENNE + CA
# ══════════════════════════════════════════════════════════════════

def join_ca(df_benne, ca_dict):
    df = df_benne.copy()
    df["total_ventes"] = df["dossier"].map(lambda d: ca_dict.get(d, {}).get("total_ventes", 0.0))
    df["client"]       = df["dossier"].map(lambda d: ca_dict.get(d, {}).get("client",       ""))
    df["etat_vente"]   = df["dossier"].map(lambda d: ca_dict.get(d, {}).get("etat_vente",   ""))
    df["_ca_trouve"]   = df["dossier"].map(lambda d: d in ca_dict)
    return df


# ══════════════════════════════════════════════════════════════════
#  CALCUL KM VIA PTV
# ══════════════════════════════════════════════════════════════════

def compute_km_benne(df, progress_cb=None):
    df = df.copy().sort_values("date_charg_dt", na_position="last").reset_index(drop=True)

    # Lignes sans date valide : km chargé calculé quand même, mais EXCLUES des km à vide
    # (une rotation sans date ne peut pas être placée chronologiquement → trajet à vide incalculable)
    df_valid   = df[df["date_charg_dt"].notna()].copy().reset_index(drop=True)
    df_no_date = df[df["date_charg_dt"].isna()].copy()

    if not df_no_date.empty:
        dossiers_nd = df_no_date["dossier"].tolist()
        st.warning(
            f"⚠️ **{len(df_no_date)} rotation(s) sans date valide** — km à vide non calculé pour : "
            f"`{dossiers_nd}`. Vérifiez la cellule date dans le fichier source."
        )

    # 1. Géocodage points uniques (sur TOUTES les rotations y compris sans date)
    points = {}
    for _, row in df.iterrows():
        for prefix in ("charg", "decharg"):
            key = (
                str(row.get(f"ville_{prefix}",   "") or "").strip(),
                str(row.get(f"cp_{prefix}_num",  "") or "").strip(),
                str(row.get(f"pays_{prefix}",    "") or "").upper(),
            )
            if key not in points:
                points[key] = None

    total_geo = len(points)
    for i, key in enumerate(list(points.keys())):
        ville, cp, pays = key
        label = f"{ville} {cp}".strip() or "(inconnu)"
        if progress_cb:
            progress_cb(f"🌍 Géocodage {i+1}/{total_geo} : {label}…", i / max(total_geo, 1))
        coords = geocode_stop(ville, cp, pays)
        points[key] = coords
        if coords is None and (ville or cp):
            st.warning(f"⚠️ Géocodage échoué : **{label}** ({pays})")

    # 2. Km chargés
    km_charges = []
    for i, row in df.iterrows():
        key_ch = (str(row.get("ville_charg",    "") or "").strip(),
                  str(row.get("cp_charg_num",   "") or "").strip(),
                  str(row.get("pays_charg",     "") or "").upper())
        key_de = (str(row.get("ville_decharg",  "") or "").strip(),
                  str(row.get("cp_decharg_num", "") or "").strip(),
                  str(row.get("pays_decharg",   "") or "").upper())
        if progress_cb:
            pct = (total_geo + i) / (total_geo + len(df) * 2)
            progress_cb(
                f"📍 Km chargé {i+1}/{len(df)} : "
                f"{row.get('ville_charg','?')} → {row.get('ville_decharg','?')}…", pct)
        km_charges.append(calculate_route_km(points.get(key_ch), points.get(key_de)))

    df["km_charge"] = km_charges

    # 3. Km à vide — uniquement sur les rotations avec date valide (ordre chronologique garanti)
    # On travaille sur df_valid ; on réinjectera les km_vide dans df complet après
    km_vides_valid = [None] * len(df_valid)
    vide_detail    = []

    for i in range(len(df_valid) - 1):
        key_de = (str(df_valid.at[i,   "ville_decharg"]  or "").strip(),
                  str(df_valid.at[i,   "cp_decharg_num"] or "").strip(),
                  str(df_valid.at[i,   "pays_decharg"]   or "").upper())
        key_ch_next = (str(df_valid.at[i+1, "ville_charg"]    or "").strip(),
                       str(df_valid.at[i+1, "cp_charg_num"]   or "").strip(),
                       str(df_valid.at[i+1, "pays_charg"]     or "").upper())
        if progress_cb:
            pct = (total_geo + len(df) + i) / (total_geo + len(df) * 2)
            progress_cb(
                f"⚡ Km à vide {i+1}/{len(df_valid)-1} : "
                f"{df_valid.at[i, 'ville_decharg']} → {df_valid.at[i+1, 'ville_charg']}…", pct)
        km_v = calculate_route_km(points.get(key_de), points.get(key_ch_next))
        km_vides_valid[i] = km_v
        vide_detail.append({
            "rotation_depart":  i + 1,
            "rotation_arrivee": i + 2,
            "dossier_depart":   df_valid.at[i,   "dossier"],
            "dossier_arrivee":  df_valid.at[i+1, "dossier"],
            "ville_depart":     df_valid.at[i,   "ville_decharg"],
            "cp_depart":        df_valid.at[i,   "cp_decharg"],
            "ville_arrivee":    df_valid.at[i+1, "ville_charg"],
            "cp_arrivee":       df_valid.at[i+1, "cp_charg"],
            "date_depart":      df_valid.at[i,   "date_decharg_dt"],
            "date_arrivee":     df_valid.at[i+1, "date_charg_dt"],
            "km_vide":          km_v,
        })

    df_valid["km_vide"] = km_vides_valid

    # Réassembler : rotations sans date → km_vide = None
    df = pd.concat([df_valid, df_no_date], ignore_index=True)

    df["km_complet"] = df["km_charge"].fillna(0) + df["km_vide"].fillna(0)
    df["renta_km"]   = (df["total_ventes"] / df["km_complet"].replace(0, np.nan)).round(3)

    return df, pd.DataFrame(vide_detail)


# ══════════════════════════════════════════════════════════════════
#  RÉSUMÉ PAR PÉRIODE
# ══════════════════════════════════════════════════════════════════

def make_resume_periode(df, group_by="semaine"):
    df = df[df["date_charg_dt"].notna()].copy()
    if group_by == "semaine":
        df["_periode"] = df["date_charg_dt"].dt.to_period("W").astype(str)
        label_col = "Semaine"
    else:
        df["_periode"] = df["date_charg_dt"].dt.to_period("M").astype(str)
        label_col = "Mois"

    agg = df.groupby("_periode", as_index=False).agg(
        Rotations  = ("dossier",      "count"),
        CA_Total   = ("total_ventes", "sum"),
        KM_Charges = ("km_charge",    "sum"),
        KM_Vide    = ("km_vide",      "sum"),
    ).round(1)
    agg["KM_Complet"]      = agg["KM_Charges"] + agg["KM_Vide"]
    agg["Pct_Vide"]        = (agg["KM_Vide"] / agg["KM_Complet"].replace(0, np.nan) * 100).round(1)
    agg["Renta_km"]        = (agg["CA_Total"] / agg["KM_Complet"].replace(0, np.nan)).round(3)
    agg["CA_moy_rotation"] = (agg["CA_Total"] / agg["Rotations"].replace(0, np.nan)).round(0)

    return agg.rename(columns={
        "_periode": label_col, "Rotations": "Nb Rotations",
        "CA_Total": "CA (€)", "KM_Charges": "KM Chargés", "KM_Vide": "KM À Vide",
        "KM_Complet": "KM Complet", "Pct_Vide": "% À Vide",
        "Renta_km": "Renta €/km", "CA_moy_rotation": "CA moy/rotation (€)",
    })


# ══════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════

def export_excel_benne(df_result, df_vide, df_periode):
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            col_rename = {
                "dossier":       "N° Dossier",
                "date_charg":    "Date charg.",
                "ville_charg":   "Chargement",
                "cp_charg":      "CP charg.",
                "date_decharg":  "Date décharg.",
                "ville_decharg": "Déchargement",
                "cp_decharg":    "CP décharg.",
                "client":        "Client",
                "etat_vente":    "État vente",
                "total_ventes":  "Total Ventes (€)",
                "km_charge":     "KM Chargés",
                "km_vide":       "KM À Vide",
                "km_complet":    "KM Complet",
                "renta_km":      "Renta €/km",
            }
            cols = [c for c in col_rename if c in df_result.columns]
            df_result[cols].rename(columns=col_rename).fillna("").to_excel(
                writer, sheet_name="Rotations", index=False)
            _style_sheet(writer.sheets["Rotations"], len(df_result))

            if not df_periode.empty:
                df_periode.to_excel(writer, sheet_name="Résumé Période", index=False)
                _style_sheet(writer.sheets["Résumé Période"], len(df_periode))

            if not df_vide.empty:
                vide_rename = {
                    "rotation_depart": "Rotation départ", "rotation_arrivee": "Rotation arrivée",
                    "dossier_depart": "Dossier départ", "dossier_arrivee": "Dossier arrivée",
                    "ville_depart": "Ville départ", "cp_depart": "CP départ",
                    "ville_arrivee": "Ville arrivée", "cp_arrivee": "CP arrivée",
                    "date_depart": "Date départ", "date_arrivee": "Date arrivée",
                    "km_vide": "KM À Vide",
                }
                cols_v = [c for c in vide_rename if c in df_vide.columns]
                df_vide[cols_v].rename(columns=vide_rename).fillna("").to_excel(
                    writer, sheet_name="KM À Vide Détail", index=False)
                _style_sheet(writer.sheets["KM À Vide Détail"], len(df_vide))

    except Exception as e:
        st.error(f"❌ Erreur génération Excel : {e}")
        return b""
    return output.getvalue()


def _style_sheet(ws, nb_rows):
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, nb_rows + 2):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ══════════════════════════════════════════════════════════════════
#  INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Rentabilité Benne", page_icon="🪣", layout="wide")
st.title("🪣 Rentabilité Benne")
st.caption("Rotations benne : CA (Total Ventes), km chargés, km à vide, rentabilité €/km.")

if not PTV_API_KEY or PTV_API_KEY == "METS_TA_CLE_ICI":
    st.error("⚠️ Clé PTV_API_KEY non configurée. Le calcul de distances ne fonctionnera pas.")

st.divider()

# ── Upload ──
col_up1, col_up2 = st.columns(2)
with col_up1:
    st.markdown("#### 🪣 Fichier Benne")
    st.caption("Colonnes : Dossier · Date charg. · C.P. charg. · Localité charg. · "
               "Date Décharg. · C.P. Déharg. · Localité Décharg.")
    file_benne = st.file_uploader("Export benne (.xlsx)", type=["xlsx"], key="benne")
with col_up2:
    st.markdown("#### 💶 Fichier CA")
    st.caption("Colonnes utilisées : N° Dossier · Total des ventes · "
               "Client facturation · Etat vente")
    file_ca = st.file_uploader("Export CA (.xlsx)", type=["xlsx"], key="ca")

st.divider()

# ── Traitement principal ──
if file_benne and file_ca:

    with st.spinner("📂 Lecture des fichiers…"):
        try:
            df_benne = parse_benne(file_benne)
            ca_dict  = parse_ca(file_ca)
        except Exception as e:
            st.error(f"❌ Erreur lecture : {e}")
            st.stop()

    df_joint = join_ca(df_benne, ca_dict)

    nb_matches = int(df_joint["_ca_trouve"].sum())
    nb_total   = len(df_joint)
    if nb_matches < nb_total:
        manquants = df_joint[~df_joint["_ca_trouve"]]["dossier"].tolist()
        st.warning(
            f"⚠️ **{nb_total - nb_matches} dossier(s)** non trouvés dans le fichier CA "
            f"(Total Ventes = 0 €) : `{manquants}`"
        )

    # Période
    dates_val = df_joint["date_charg_dt"].dropna()
    MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
               "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    if not dates_val.empty:
        d_min, d_max = dates_val.min(), dates_val.max()
        if d_min.month == d_max.month and d_min.year == d_max.year:
            periode_label = f"{MOIS_FR[d_min.month]} {d_min.year}"
        else:
            periode_label = f"{d_min.strftime('%d/%m/%Y')} → {d_max.strftime('%d/%m/%Y')}"
    else:
        periode_label = "Période inconnue"

    # KPIs aperçu
    st.markdown(f"### 📊 Aperçu — {periode_label}")
    ca_total = df_joint["total_ventes"].sum()
    nb_rot   = len(df_joint)
    ca_moy   = ca_total / nb_rot if nb_rot > 0 else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("🔄 Rotations",        nb_rot)
    k2.metric("✅ Matchés CA",        f"{nb_matches}/{nb_total}")
    k3.metric("💶 Total Ventes",     f"{ca_total:,.2f} €")
    k4.metric("📈 CA moy/rotation",  f"{ca_moy:,.0f} €")
    k5.metric("🗓️ Période",           periode_label)

    st.divider()

    # Tableau consolidé + filtres
    st.markdown("### 📋 Tableau des rotations")

    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        villes_ch = sorted(df_joint["ville_charg"].replace("", np.nan).dropna().unique())
        filtre_ch = st.multiselect("📍 Lieu de chargement :", options=villes_ch,
                                   default=[], placeholder="Tous")
    with fc2:
        villes_de = sorted(df_joint["ville_decharg"].replace("", np.nan).dropna().unique())
        filtre_de = st.multiselect("📍 Lieu de déchargement :", options=villes_de,
                                   default=[], placeholder="Tous")
    with fc3:
        clients = sorted(df_joint["client"].replace("", np.nan).dropna().unique())
        filtre_cl = st.multiselect("🏢 Client :", options=clients,
                                   default=[], placeholder="Tous les clients")

    df_display = df_joint.copy()
    if filtre_ch:
        df_display = df_display[df_display["ville_charg"].isin(filtre_ch)]
    if filtre_de:
        df_display = df_display[df_display["ville_decharg"].isin(filtre_de)]
    if filtre_cl:
        df_display = df_display[df_display["client"].isin(filtre_cl)]

    cols_show = ["dossier", "date_charg", "ville_charg", "cp_charg",
                 "ville_decharg", "cp_decharg", "client", "etat_vente", "total_ventes"]
    st.dataframe(
        df_display[[c for c in cols_show if c in df_display.columns]].rename(columns={
            "dossier":       "N° Dossier",
            "date_charg":    "Date charg.",
            "ville_charg":   "Chargement",
            "cp_charg":      "CP charg.",
            "ville_decharg": "Déchargement",
            "cp_decharg":    "CP décharg.",
            "client":        "Client",
            "etat_vente":    "État vente",
            "total_ventes":  "Total Ventes (€)",
        }),
        use_container_width=True,
        height=380,
    )

    st.divider()

    # Calcul PTV
    st.markdown("### 🗺️ Calcul KM via PTV")
    st.info(
        f"ℹ️ Géocodage des arrêts et calcul km PTV pour les **{nb_rot} rotations**. "
        f"Les km à vide sont calculés de chaque déchargement vers le chargement "
        f"suivant (ordre chronologique)."
    )

    btn_ptv = st.button("🚀 Lancer le calcul PTV", type="primary")

    if btn_ptv:
        for key in ["df_result_benne", "df_vide_benne"]:
            st.session_state.pop(key, None)

        progress_bar = st.progress(0)
        status_text  = st.empty()

        def _progress(msg, pct=0.0):
            status_text.text(msg)
            progress_bar.progress(min(float(pct), 1.0))

        try:
            df_result, df_vide = compute_km_benne(df_joint, progress_cb=_progress)
            progress_bar.progress(1.0)
            status_text.success("✅ Calcul PTV terminé !")
            st.session_state["df_result_benne"] = df_result
            st.session_state["df_vide_benne"]   = df_vide
        except Exception as e:
            st.error(f"❌ Erreur durant le calcul PTV : {e}")

    # Résultats PTV
    if "df_result_benne" in st.session_state:
        df_result = st.session_state["df_result_benne"]
        df_vide   = st.session_state.get("df_vide_benne", pd.DataFrame())

        st.divider()
        st.markdown("### 📈 Résultats KM")

        km_ch_sum   = df_result["km_charge"].fillna(0).sum()
        km_vide_sum = df_result["km_vide"].fillna(0).sum()
        km_complet  = km_ch_sum + km_vide_sum
        pct_vide    = (km_vide_sum / km_complet * 100) if km_complet > 0 else 0
        ca_tot      = df_result["total_ventes"].sum()
        rent_global = ca_tot / km_complet if km_complet > 0 else 0

        kp1, kp2, kp3, kp4, kp5, kp6 = st.columns(6)
        kp1.metric("📏 KM Chargés",    f"{km_ch_sum:,.0f} km")
        kp2.metric("⚡ KM À Vide",      f"{km_vide_sum:,.0f} km")
        kp3.metric("🔄 KM Complet",    f"{km_complet:,.0f} km")
        kp4.metric("% À Vide",          f"{pct_vide:.1f}%")
        kp5.metric("💶 Total Ventes",   f"{ca_tot:,.2f} €")
        kp6.metric("📈 Rentabilité",    f"{rent_global:.3f} €/km")

        tab1, tab2, tab3, tab4 = st.tabs([
            "📋 Détail rotations",
            "📅 Résumé semaine",
            "📆 Résumé mois",
            "⚡ KM à vide détail",
        ])

        with tab1:
            cols_res = ["dossier", "date_charg", "ville_charg", "cp_charg",
                        "ville_decharg", "cp_decharg", "client",
                        "total_ventes", "km_charge", "km_vide", "km_complet", "renta_km"]
            st.dataframe(
                df_result[[c for c in cols_res if c in df_result.columns]].rename(columns={
                    "dossier":       "N° Dossier",
                    "date_charg":    "Date charg.",
                    "ville_charg":   "Chargement",
                    "cp_charg":      "CP charg.",
                    "ville_decharg": "Déchargement",
                    "cp_decharg":    "CP décharg.",
                    "client":        "Client",
                    "total_ventes":  "Total Ventes (€)",
                    "km_charge":     "KM Chargés",
                    "km_vide":       "KM À Vide",
                    "km_complet":    "KM Complet",
                    "renta_km":      "Renta €/km",
                }),
                use_container_width=True,
                height=400,
            )

        with tab2:
            st.dataframe(make_resume_periode(df_result, "semaine"), use_container_width=True)

        with tab3:
            st.dataframe(make_resume_periode(df_result, "mois"), use_container_width=True)

        with tab4:
            if not df_vide.empty:
                st.dataframe(
                    df_vide.rename(columns={
                        "rotation_depart": "Rotation départ", "rotation_arrivee": "Rotation arrivée",
                        "dossier_depart": "Dossier départ", "dossier_arrivee": "Dossier arrivée",
                        "ville_depart": "Ville départ", "cp_depart": "CP départ",
                        "ville_arrivee": "Ville arrivée", "cp_arrivee": "CP arrivée",
                        "km_vide": "KM À Vide",
                    }),
                    use_container_width=True,
                )
            else:
                st.info("Aucun trajet à vide détecté.")

        st.divider()
        df_mois_exp = make_resume_periode(df_result, "mois")
        excel_bytes = export_excel_benne(df_result, df_vide, df_mois_exp)
        st.download_button(
            label="📥 Télécharger le rapport Excel",
            data=excel_bytes,
            file_name="Rapport_Renta_Benne.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
        if st.button("🔄 Nouveau calcul"):
            for key in ["df_result_benne", "df_vide_benne"]:
                st.session_state.pop(key, None)
            st.rerun()

elif file_benne and not file_ca:
    st.info("📂 Fichier benne chargé — en attente du fichier CA…")
elif file_ca and not file_benne:
    st.info("📂 Fichier CA chargé — en attente du fichier benne…")
else:
    st.markdown("""
    #### Comment utiliser cet outil

    1. **Chargez le fichier Benne** — export avec les colonnes :
       `Dossier` · `Date charg.` · `C.P. charg.` · `Localité charg.` ·
       `Date Décharg.` · `C.P. Déharg.` · `Localité Décharg.`

    2. **Chargez le fichier CA** — export complet (seuls `N° Dossier` et
       `Total des ventes` sont utilisés pour la jointure)

    3. La jointure se fait automatiquement sur le **N° Dossier** →
       chaque rotation récupère son **Total des ventes** depuis le fichier CA

    4. Consultez le tableau consolidé et utilisez les filtres

    5. Cliquez sur **Lancer le calcul PTV** pour obtenir :
       - **KM chargés** par rotation (chargement → déchargement)
       - **KM à vide** entre chaque déchargement et le chargement suivant
       - **Rentabilité €/km** par rotation et par période (semaine / mois)

    6. **Téléchargez le rapport Excel** (rotations + résumé mensuel + détail km à vide)

    ---
    > Le format `"F 91270"` / `"B 9000"` des codes postaux est géré automatiquement.
    > Clé PTV requise dans `.env` (`PTV_API_KEY`).
    """)
