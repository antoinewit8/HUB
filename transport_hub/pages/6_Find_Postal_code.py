"""
Page Streamlit — Enrichissement des codes postaux manquants
Supporte deux formats de fichier :
  1. Format standard transport_hub (ville/cp/pays en colonnes séparées)
  2. Format "grille" : from / From Country / to / To Country (sans CP)
"""

import streamlit as st
import os, sys, io, time, re, unicodedata, requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

st.set_page_config(page_title="Enrichissement CP", page_icon="📮", layout="wide")

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #0d1b2a; }
[data-testid="stSidebar"]           { background: #0d1b2a; }
h1, h2, h3, .stMarkdown p          { color: #e8eaf6; }
.stat-card {
    background: #1a2a3a; border: 1px solid #2F5496;
    border-radius: 10px; padding: 18px 24px; text-align: center; margin: 6px;
}
.stat-card .val { font-size: 2rem; font-weight: 700; color: #4fc3f7; }
.stat-card .lbl { font-size: 0.85rem; color: #90a4ae; margin-top: 4px; }
.stButton > button {
    background: #2F5496; color: white; border: none;
    border-radius: 8px; padding: 10px 28px; font-weight: 600;
}
.stButton > button:hover { background: #1a3a6e; }
</style>
""", unsafe_allow_html=True)

st.title("📮 Enrichissement des Codes Postaux")
st.markdown("Détecte les villes sans code postal et les complète via l'API PTV — supporte les formats standards et les grilles tarifaires.")
st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
PTV_API_KEY = os.environ.get("PTV_API_KEY", "")
GEOCODE_URL = "https://api.myptv.com/geocoding/v1"

PAYS_MAP = {
    "FR": "France", "BE": "Belgium", "DE": "Germany", "NL": "Netherlands",
    "LU": "Luxembourg", "IT": "Italy", "ES": "Spain", "PT": "Portugal",
    "GB": "United Kingdom", "CH": "Switzerland", "AT": "Austria",
    "PL": "Poland", "CZ": "Czech Republic", "HU": "Hungary",
    "RO": "Romania", "BG": "Bulgaria", "SK": "Slovakia", "SI": "Slovenia",
    "HR": "Croatia", "DK": "Denmark", "SE": "Sweden", "NO": "Norway",
    "FI": "Finland", "IE": "Ireland", "GR": "Greece",
}
PAYS_TO_ISO = {v.lower(): k for k, v in PAYS_MAP.items()}
PAYS_TO_ISO.update({
    "france": "FR", "belgique": "BE", "allemagne": "DE", "pays-bas": "NL",
    "luxembourg": "LU", "italie": "IT", "espagne": "ES", "portugal": "PT",
    "suisse": "CH", "autriche": "AT", "pologne": "PL", "tcheque": "CZ",
    "hongrie": "HU", "roumanie": "RO", "bulgarie": "BG", "croatie": "HR",
    "danemark": "DK", "suede": "SE", "norvege": "NO", "finlande": "FI",
    "irlande": "IE", "grece": "GR",
})
CP_LENGTHS = {
    "FR": 5, "BE": 4, "DE": 5, "IT": 5, "ES": 5, "PL": 5, "CZ": 5,
    "HR": 5, "SK": 5, "GR": 5, "SE": 5, "FI": 5, "NL": 4, "AT": 4,
    "CH": 4, "HU": 4, "DK": 4, "SI": 4, "NO": 4, "LU": 4, "PT": 7, "RO": 6,
}
MAX_RETRIES = 3
RETRY_DELAY = 1.5

# Mots-clés format standard
ORIG_PAYS_KW  = ["pays", "country", "country of origin", "orig cntry", "pays depart", "pays origine"]
ORIG_CP_KW    = ["code postal", "cp", "postal code", "postal code origin", "orig reg", "cp depart"]
ORIG_VILLE_KW = ["ville", "city", "city of origin", "orig zone txt", "localite", "origin", "depart", "origine", "ville depart"]
DEST_PAYS_KW  = ["pays", "country", "country of destination", "dest cntry", "pays destination", "dest pays"]
DEST_CP_KW    = ["departement", "code postal", "cp", "postal code", "postal code destination", "dest reg", "cp destination", "cp dech"]
DEST_VILLE_KW = ["ville2", "ville", "city", "city of destination", "dest zone txt", "destination", "ville destination", "ville dest"]
ORIGIN_GROUP_KW = ["depart", "origin", "origine", "chargement", "loading"]
DEST_GROUP_KW   = ["destination", "dest", "arrivee", "livraison", "delivery", "unloading"]

# Mots-clés format grille (from/to)
GRILLE_FROM_KW         = ["from", "from city", "ville depart", "origine", "origin city"]
GRILLE_FROM_COUNTRY_KW = ["from country", "pays depart", "orig country", "country of origin", "orig cntry"]
GRILLE_FROM_CP_KW      = ["from cp", "from postal", "cp depart", "postal code origin", "orig cp"]
GRILLE_TO_KW           = ["to", "to city", "ville dest", "ville destination", "destination city"]
GRILLE_TO_COUNTRY_KW   = ["to country", "pays dest", "pays destination", "dest country", "country of destination", "dest cntry"]
GRILLE_TO_CP_KW        = ["to cp", "to postal", "cp dest", "cp destination", "postal code destination", "dest cp"]


# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────────────────────
def normalize(val):
    if val is None:
        return ""
    val = str(val).strip().lower()
    val = unicodedata.normalize("NFD", val)
    val = "".join(c for c in val if unicodedata.category(c) != "Mn")
    return re.sub(r'\s+', ' ', val).strip()

def pays_to_iso(pays_str: str) -> str:
    if not pays_str:
        return "FR"
    p = pays_str.strip().upper()
    if p in PAYS_MAP:
        return p
    return PAYS_TO_ISO.get(normalize(pays_str), "FR")

def pad_cp(cp: str, iso: str) -> str:
    if not cp:
        return ""
    target = CP_LENGTHS.get(iso, 5)
    return cp.zfill(target) if target else cp


# ─────────────────────────────────────────────────────────────────────────────
# DÉTECTION FORMAT
# ─────────────────────────────────────────────────────────────────────────────
def detect_format(ws) -> str:
    all_grille = set(GRILLE_FROM_KW + GRILLE_FROM_COUNTRY_KW + GRILLE_TO_KW + GRILLE_TO_COUNTRY_KW)
    for r in range(1, min(10, ws.max_row + 1)):
        score = sum(1 for c in range(1, ws.max_column + 1)
                    if normalize(ws.cell(r, c).value) in all_grille)
        if score >= 2:
            return 'grille'
    return 'standard'


def detect_grille_columns(ws):
    all_grille = set(
        GRILLE_FROM_KW + GRILLE_FROM_COUNTRY_KW + GRILLE_FROM_CP_KW +
        GRILLE_TO_KW   + GRILLE_TO_COUNTRY_KW   + GRILLE_TO_CP_KW
    )
    best_row, best_score = None, 0
    for r in range(1, min(15, ws.max_row + 1)):
        score = sum(1 for c in range(1, ws.max_column + 1)
                    if normalize(ws.cell(r, c).value) in all_grille)
        if score > best_score:
            best_score, best_row = score, r
    if best_score < 2 or best_row is None:
        return None, None

    headers = {c: normalize(ws.cell(best_row, c).value)
               for c in range(1, ws.max_column + 1)
               if ws.cell(best_row, c).value}

    mapping = {}
    for col, val in headers.items():
        if val in GRILLE_FROM_KW         and 'from'         not in mapping: mapping['from']         = col
        if val in GRILLE_FROM_COUNTRY_KW and 'from_country' not in mapping: mapping['from_country'] = col
        if val in GRILLE_FROM_CP_KW      and 'from_cp'      not in mapping: mapping['from_cp']      = col
        if val in GRILLE_TO_KW           and 'to'           not in mapping: mapping['to']           = col
        if val in GRILLE_TO_COUNTRY_KW   and 'to_country'   not in mapping: mapping['to_country']   = col
        if val in GRILLE_TO_CP_KW        and 'to_cp'        not in mapping: mapping['to_cp']        = col

    if 'from' not in mapping or 'to' not in mapping:
        return None, None
    return mapping, best_row


# ─────────────────────────────────────────────────────────────────────────────
# DÉTECTION FORMAT STANDARD
# ─────────────────────────────────────────────────────────────────────────────
def detect_header_row_std(ws, max_scan=15):
    all_kw = set(ORIG_PAYS_KW + ORIG_CP_KW + ORIG_VILLE_KW + DEST_PAYS_KW + DEST_CP_KW + DEST_VILLE_KW)
    best_row, best_score = None, 0
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        score = sum(1 for c in range(1, ws.max_column + 1) if normalize(ws.cell(r, c).value) in all_kw)
        if score > best_score:
            best_score, best_row = score, r
    return best_row if best_score >= 2 else None

def detect_groups_std(ws, header_row):
    if header_row <= 1:
        return {}
    group_row = header_row - 1
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= group_row <= mr.max_row:
            val = ws.cell(mr.min_row, mr.min_col).value
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[c] = val
    groups, current = {}, None
    for col in range(1, ws.max_column + 1):
        val = normalize(merge_map.get(col, ws.cell(group_row, col).value))
        if any(k in val for k in ORIGIN_GROUP_KW):   current = "origin"
        elif any(k in val for k in DEST_GROUP_KW):    current = "dest"
        elif val:                                      current = None
        if current:
            groups[col] = current
    return groups

def map_columns_std(ws):
    header_row = detect_header_row_std(ws)
    if not header_row:
        return None, None, None
    groups  = detect_groups_std(ws, header_row)
    headers = {c: normalize(ws.cell(header_row, c).value)
               for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value}
    ROLES = {
        "orig_pays":  (ORIG_PAYS_KW,  "origin"),
        "orig_cp":    (ORIG_CP_KW,    "origin"),
        "orig_ville": (ORIG_VILLE_KW, "origin"),
        "dest_pays":  (DEST_PAYS_KW,  "dest"),
        "dest_cp":    (DEST_CP_KW,    "dest"),
        "dest_ville": (DEST_VILLE_KW, "dest"),
    }
    mapping, used = {}, set()
    for role, (kws, side) in ROLES.items():
        for col, hval in headers.items():
            if col in used: continue
            if groups.get(col) == side and hval in kws:
                mapping[role] = col; used.add(col); break
        if role not in mapping:
            for col, hval in headers.items():
                if col in used: continue
                if hval in kws:
                    mapping[role] = col; used.add(col); break
    return mapping, header_row, header_row + 1


# ─────────────────────────────────────────────────────────────────────────────
# GÉOCODAGE PTV
# ─────────────────────────────────────────────────────────────────────────────
def fetch_cp_from_ptv(ville: str, iso: str):
    if not PTV_API_KEY:
        return None
    hdrs   = {"apiKey": PTV_API_KEY}
    params = {"searchText": ville, "countryFilter": iso}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(f"{GEOCODE_URL}/locations/by-text",
                             headers=hdrs, params=params, timeout=10)
            if r.status_code == 429:
                time.sleep(RETRY_DELAY * attempt); continue
            if r.status_code in (400, 404):
                return None
            r.raise_for_status()
            locs = r.json().get("locations", [])
            if not locs:
                return None
            vn = normalize(ville)
            for loc in locs:
                addr = loc.get("address", {})
                cn   = normalize(addr.get("city", ""))
                cp   = addr.get("postalCode", "")
                st_  = addr.get("street", "")
                if not cp: continue
                if (vn in cn or cn in vn) and vn not in normalize(st_):
                    return pad_cp(cp, iso)
            for loc in locs:
                cp = loc.get("address", {}).get("postalCode", "")
                if cp:
                    return pad_cp(cp, iso)
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# ANALYSE
# ─────────────────────────────────────────────────────────────────────────────
def analyser_grille(wb, enrichir_origine):
    resultats = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if detect_format(ws) != 'grille':
            continue
        mapping, header_row = detect_grille_columns(ws)
        if not mapping:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            from_v = str(ws.cell(row, mapping['from']).value or "").strip()
            to_v   = str(ws.cell(row, mapping['to']).value   or "").strip()
            if not from_v and not to_v:
                continue
            from_c = str(ws.cell(row, mapping.get('from_country', 0)).value or "").strip() if mapping.get('from_country') else "FR"
            to_c   = str(ws.cell(row, mapping.get('to_country',   0)).value or "").strip() if mapping.get('to_country')   else "FR"
            from_cp = str(ws.cell(row, mapping.get('from_cp', 0)).value or "").strip() if mapping.get('from_cp') else ""
            to_cp   = str(ws.cell(row, mapping.get('to_cp',   0)).value or "").strip() if mapping.get('to_cp')   else ""

            if enrichir_origine and from_v and not from_cp:
                resultats.append({'sheet': sheet_name, 'row': row, 'role': 'from',
                                   'ville': from_v, 'iso': pays_to_iso(from_c),
                                   'cp_col': mapping.get('from_cp'), 'ville_col': mapping['from'],
                                   'mapping': mapping, 'header_row': header_row})
            if to_v and not to_cp:
                resultats.append({'sheet': sheet_name, 'row': row, 'role': 'to',
                                   'ville': to_v, 'iso': pays_to_iso(to_c),
                                   'cp_col': mapping.get('to_cp'), 'ville_col': mapping['to'],
                                   'mapping': mapping, 'header_row': header_row})
    return resultats

def analyser_standard(wb, enrichir_origine):
    resultats = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if detect_format(ws) == 'grille':
            continue
        mapping, header_row, data_start = map_columns_std(ws)
        if not mapping or 'dest_ville' not in mapping:
            continue
        for row in range(data_start, ws.max_row + 1):
            vd = str(ws.cell(row, mapping['dest_ville']).value or "").strip()
            cd = str(ws.cell(row, mapping.get('dest_cp', 0)).value or "").strip() if mapping.get('dest_cp') else ""
            pd_ = str(ws.cell(row, mapping.get('dest_pays', 0)).value or "").strip() if mapping.get('dest_pays') else "France"
            if vd and not cd:
                resultats.append({'sheet': sheet_name, 'row': row, 'role': 'to',
                                   'ville': vd, 'iso': pays_to_iso(pd_),
                                   'cp_col': mapping.get('dest_cp'), 'ville_col': mapping['dest_ville'],
                                   'mapping': mapping, 'header_row': header_row})
            if enrichir_origine and 'orig_ville' in mapping:
                vo = str(ws.cell(row, mapping['orig_ville']).value or "").strip()
                co = str(ws.cell(row, mapping.get('orig_cp', 0)).value or "").strip() if mapping.get('orig_cp') else ""
                po = str(ws.cell(row, mapping.get('orig_pays', 0)).value or "").strip() if mapping.get('orig_pays') else "France"
                if vo and not co:
                    resultats.append({'sheet': sheet_name, 'row': row, 'role': 'from',
                                       'ville': vo, 'iso': pays_to_iso(po),
                                       'cp_col': mapping.get('orig_cp'), 'ville_col': mapping['orig_ville'],
                                       'mapping': mapping, 'header_row': header_row})
    return resultats


# ─────────────────────────────────────────────────────────────────────────────
# ÉCRITURE
# ─────────────────────────────────────────────────────────────────────────────
def ecrire_cp(wb2, lignes, cache_cp, creer_colonne):
    cp_font   = Font(color="1F6B2E", bold=True)
    cp_fill   = PatternFill("solid", fgColor="E8F5E9")
    miss_font = Font(color="B71C1C")
    miss_fill = PatternFill("solid", fgColor="FFEBEE")
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center")

    created_cols = {}  # (sheet, role) → col_index
    ecrites = manquees = 0

    for r_info in lignes:
        ws2       = wb2[r_info['sheet']]
        row       = r_info['row']
        ville     = r_info['ville']
        iso       = r_info['iso']
        cp_col    = r_info.get('cp_col')
        ville_col = r_info['ville_col']
        role      = r_info['role']
        hdr_row   = r_info['header_row']
        key       = (normalize(ville), iso)
        cp_found  = cache_cp.get(key, "")

        if not cp_col and creer_colonne:
            ck = (r_info['sheet'], role)
            if ck in created_cols:
                cp_col = created_cols[ck]
            else:
                insert_col = ville_col + 1
                ws2.insert_cols(insert_col)
                lbl = "CP Depart" if role == 'from' else "CP Destination"
                hc = ws2.cell(hdr_row, insert_col)
                hc.value = lbl; hc.font = hdr_font
                hc.fill = hdr_fill; hc.alignment = hdr_align
                cp_col = insert_col
                created_cols[ck] = cp_col
                for other in lignes:
                    if other['sheet'] == r_info['sheet'] and other['role'] == role and not other.get('cp_col'):
                        other['cp_col'] = cp_col
            r_info['cp_col'] = cp_col

        if not cp_col:
            continue

        cell = ws2.cell(row, cp_col)
        if cp_found:
            cell.value = cp_found; cell.font = cp_font; cell.fill = cp_fill
            ecrites += 1
        else:
            cell.value = "?"; cell.font = miss_font; cell.fill = miss_fill
            manquees += 1

    return ecrites, manquees


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader("📂 Dépose ton fichier Excel", type=["xlsx"])

col1, col2 = st.columns(2)
with col1:
    creer_colonne    = st.checkbox("➕ Créer la colonne CP si absente", value=True)
with col2:
    enrichir_origine = st.checkbox("🔄 Enrichir aussi les départs sans CP", value=True)

if uploaded:
    raw_bytes = uploaded.read()
    wb        = openpyxl.load_workbook(io.BytesIO(raw_bytes))

    fmts = {sn: detect_format(wb[sn]) for sn in wb.sheetnames}
    has_grille   = any(v == 'grille'   for v in fmts.values())
    has_standard = any(v == 'standard' for v in fmts.values())

    fmt_parts = []
    if has_grille:   fmt_parts.append("**grille tarifaire** (from / to)")
    if has_standard: fmt_parts.append("**standard** (ville / CP / pays)")
    st.info(f"📋 Format détecté : {' + '.join(fmt_parts) if fmt_parts else 'non reconnu'}")

    lignes  = []
    if has_grille:   lignes += analyser_grille(wb, enrichir_origine)
    if has_standard: lignes += analyser_standard(wb, enrichir_origine)

    total   = len(lignes)
    uniques = len(set((r['ville'], r['iso']) for r in lignes))
    sheets  = len(set(r['sheet'] for r in lignes))

    st.markdown("### 📊 Analyse du fichier")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="stat-card"><div class="val">{total}</div><div class="lbl">Cellules sans CP</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-card"><div class="val">{uniques}</div><div class="lbl">Villes uniques à géocoder</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="stat-card"><div class="val">{sheets}</div><div class="lbl">Feuilles concernées</div></div>', unsafe_allow_html=True)

    if total == 0:
        st.success("✅ Toutes les villes ont déjà un code postal !")
    else:
        import pandas as pd
        with st.expander(f"👁️ Aperçu des {min(total, 150)} premières lignes sans CP", expanded=False):
            preview = [{"Feuille": r['sheet'], "Ligne": r['row'],
                        "Rôle": "Départ" if r['role'] == 'from' else "Destination",
                        "Ville": r['ville'], "Pays ISO": r['iso']}
                       for r in lignes[:150]]
            st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)
            if total > 150:
                st.caption(f"... et {total - 150} autres non affichées.")

        if not PTV_API_KEY:
            st.error("❌ Clé PTV_API_KEY non configurée. Enrichissement impossible.")
            st.info("Vérifie les variables d'environnement Streamlit Cloud.")
        else:
            st.markdown("---")
            if st.button("🚀 Lancer l'enrichissement", type="primary"):

                unique_pairs = list({(r['ville'], r['iso']) for r in lignes})
                cache_cp     = {}
                progress     = st.progress(0)
                status       = st.empty()
                log_area     = st.empty()
                logs         = []
                trouves = non_trouves = 0

                for i, (ville, iso) in enumerate(unique_pairs):
                    progress.progress((i + 1) / len(unique_pairs))
                    status.markdown(f"**🔍 {i+1}/{len(unique_pairs)}** — {ville} ({iso})")
                    key = (normalize(ville), iso)
                    cp  = fetch_cp_from_ptv(ville, iso)
                    cache_cp[key] = cp or ""
                    if cp:
                        trouves += 1
                        logs.append(f"✅ {ville} ({iso}) → **{cp}**")
                    else:
                        non_trouves += 1
                        logs.append(f"⚠️ {ville} ({iso}) → non trouvé")
                    log_area.markdown("\n".join(logs[-12:]))
                    time.sleep(0.15)

                status.markdown(f"**✅ Terminé** — {trouves} trouvés · {non_trouves} non trouvés")

                wb2 = openpyxl.load_workbook(io.BytesIO(raw_bytes))
                ecrites, manquees = ecrire_cp(wb2, lignes, cache_cp, creer_colonne)

                out_buf = io.BytesIO()
                wb2.save(out_buf)
                out_buf.seek(0)

                st.markdown("---")
                st.success(f"✅ **{ecrites}** codes postaux écrits — ⚠️ **{manquees}** non trouvés (cellules `?` en rouge)")

                fname = (uploaded.name or "fichier").replace(".xlsx", "_CP_enrichi.xlsx")
                st.download_button(
                    label="⬇️ Télécharger le fichier enrichi",
                    data=out_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                if manquees > 0:
                    with st.expander("⚠️ Villes non trouvées — à corriger manuellement"):
                        non_trouv = [r for r in lignes if not cache_cp.get((normalize(r['ville']), r['iso']), "")]
                        df_miss = pd.DataFrame([{"Ville": r['ville'], "Pays ISO": r['iso'],
                                                  "Rôle": "Départ" if r['role'] == 'from' else "Destination"}
                                                 for r in non_trouv]).drop_duplicates()
                        st.dataframe(df_miss, use_container_width=True, hide_index=True)
                        st.info("💡 Ces villes peuvent être ajoutées dans GPS_FIXES ou corrigées dans le fichier source.")
else:
    st.info("📂 Dépose un fichier Excel pour démarrer l'analyse.")
