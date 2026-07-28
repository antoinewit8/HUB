"""
4____Missions_CA_KM.py
──────────────────────────────────────────────────────────────────
Outil TX-FLEX : Analyse Missions + CA + Calcul KM PTV
──────────────────────────────────────────────────────────────────
Entrées :
  • Fichier Missions  (.xlsx) — colonnes : N°Dossier, Activité, Date, Heure,
                                Nom1, Adresse, Numéro, Code pays, CP, Localité,
                                Chauffeur, Immat. tracteur
  • Fichier CA        (.xlsx) — colonnes : N°Dossier, Prix transport, Total vente

Sorties :
  • Tableau consolidé par chauffeur/dossier avec stops + CA
  • Calcul PTV : km totaux (chaîne complète) + km à vide (DECHARGER→CHARGER)
  • Rentabilité par DÉPARTEMENT de déchargement (France)
  • Export Excel
──────────────────────────────────────────────────────────────────
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import os
import io
import re
import json
from typing import Optional, Tuple, Dict, List
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════════════════════════════
#  CONFIG PTV
# ══════════════════════════════════════════════════════════════════

PTV_API_KEY  = os.environ.get("PTV_API_KEY", "METS_TA_CLE_ICI")
PTV_BASE_URL = "https://api.myptv.com/routing/v1"
GEOCODE_URL  = "https://api.myptv.com/geocoding/v1"
HEADERS      = {"apiKey": PTV_API_KEY}
MAX_RETRIES  = 3
RETRY_DELAY  = 2
MAX_WORKERS  = 4
VEHICLE      = "EUR_TRAILER_TRUCK"

# ══════════════════════════════════════════════════════════════════
#  IMPORTS excel_handler_km + ptv_router_km
# ══════════════════════════════════════════════════════════════════

try:
    from excel_handler_km import (
        PAYS_MAP, CP_LENGTHS, ZONE_CORRECTIONS,
        CITY_CORRECTIONS, GPS_FIXES_ORIGIN,
        parse_origin_from_parts,
    )
    from ptv_router_km import geocode_address as _ptv_geocode_address
    _IMPORTS_OK = True
except ImportError:
    _IMPORTS_OK = False
    PAYS_MAP = {
        "F": "France", "B": "Belgium", "D": "Germany", "L": "Luxembourg",
        "NL": "Netherlands", "E": "Spain", "I": "Italy", "CH": "Switzerland",
        "GB": "United Kingdom", "A": "Austria", "P": "Portugal",
        "FR": "France", "BE": "Belgium", "DE": "Germany", "LU": "Luxembourg",
    }
    CP_LENGTHS        = {}
    ZONE_CORRECTIONS  = {}
    CITY_CORRECTIONS  = {}
    GPS_FIXES_ORIGIN  = {}

    def parse_origin_from_parts(city, cp, country):
        pays_full = PAYS_MAP.get(str(country).strip().upper(), country)
        return ", ".join(p for p in [city, cp, pays_full] if p and p != "nan")

    def _ptv_geocode_address(address):
        return None


@st.cache_data(show_spinner=False)
def geocode_address(address):
    address = str(address).strip()
    if not address or address.lower() in ("nan", ""):
        return None
    return _ptv_geocode_address(address)


PAYS_TO_ISO2 = {
    "F": "FR", "B": "BE", "D": "DE", "L": "LU", "I": "IT",
    "E": "ES", "A": "AT", "P": "PT", "CH": "CH", "GB": "GB",
    "NL": "NL", "FR": "FR", "BE": "BE", "DE": "DE", "LU": "LU",
    "IT": "IT", "ES": "ES", "AT": "AT", "PT": "PT",
}


# ══════════════════════════════════════════════════════════════════
#  DÉPARTEMENTS FRANÇAIS
# ══════════════════════════════════════════════════════════════════

DEPT_NOMS = {
    "01": "Ain", "02": "Aisne", "03": "Allier", "04": "Alpes-de-Haute-Provence",
    "05": "Hautes-Alpes", "06": "Alpes-Maritimes", "07": "Ardèche", "08": "Ardennes",
    "09": "Ariège", "10": "Aube", "11": "Aude", "12": "Aveyron",
    "13": "Bouches-du-Rhône", "14": "Calvados", "15": "Cantal", "16": "Charente",
    "17": "Charente-Maritime", "18": "Cher", "19": "Corrèze",
    "2A": "Corse-du-Sud", "2B": "Haute-Corse",
    "21": "Côte-d'Or", "22": "Côtes-d'Armor", "23": "Creuse", "24": "Dordogne",
    "25": "Doubs", "26": "Drôme", "27": "Eure", "28": "Eure-et-Loir",
    "29": "Finistère", "30": "Gard", "31": "Haute-Garonne", "32": "Gers",
    "33": "Gironde", "34": "Hérault", "35": "Ille-et-Vilaine", "36": "Indre",
    "37": "Indre-et-Loire", "38": "Isère", "39": "Jura", "40": "Landes",
    "41": "Loir-et-Cher", "42": "Loire", "43": "Haute-Loire", "44": "Loire-Atlantique",
    "45": "Loiret", "46": "Lot", "47": "Lot-et-Garonne", "48": "Lozère",
    "49": "Maine-et-Loire", "50": "Manche", "51": "Marne", "52": "Haute-Marne",
    "53": "Mayenne", "54": "Meurthe-et-Moselle", "55": "Meuse", "56": "Morbihan",
    "57": "Moselle", "58": "Nièvre", "59": "Nord", "60": "Oise",
    "61": "Orne", "62": "Pas-de-Calais", "63": "Puy-de-Dôme", "64": "Pyrénées-Atlantiques",
    "65": "Hautes-Pyrénées", "66": "Pyrénées-Orientales", "67": "Bas-Rhin", "68": "Haut-Rhin",
    "69": "Rhône", "70": "Haute-Saône", "71": "Saône-et-Loire", "72": "Sarthe",
    "73": "Savoie", "74": "Haute-Savoie", "75": "Paris", "76": "Seine-Maritime",
    "77": "Seine-et-Marne", "78": "Yvelines", "79": "Deux-Sèvres", "80": "Somme",
    "81": "Tarn", "82": "Tarn-et-Garonne", "83": "Var", "84": "Vaucluse",
    "85": "Vendée", "86": "Vienne", "87": "Haute-Vienne", "88": "Vosges",
    "89": "Yonne", "90": "Territoire de Belfort", "91": "Essonne", "92": "Hauts-de-Seine",
    "93": "Seine-Saint-Denis", "94": "Val-de-Marne", "95": "Val-d'Oise",
    "971": "Guadeloupe", "972": "Martinique", "973": "Guyane",
    "974": "La Réunion", "976": "Mayotte",
}

PAYS_FRANCE = {"F", "FR", "FRA", "FRANCE"}


def normalize_cp_fr(cp):
    """Nettoie un code postal français : '1000' -> '01000', '69 007' -> '69007'."""
    cp = re.sub(r"[^0-9A-Za-z]", "", str(cp or "").strip()).upper()
    if not cp or cp == "NAN":
        return ""
    if cp.isdigit() and len(cp) == 4:
        cp = "0" + cp
    return cp


def extract_departement(cp, pays):
    """
    Retourne le code département FR ('01'..'95', '2A', '2B', '971'...) ou None.
    Ne retourne un département que si le pays est la France (ou vide + CP FR à 5 chiffres).
    """
    pays_n = str(pays or "").strip().upper()
    cp_n   = normalize_cp_fr(cp)

    if not cp_n:
        return None

    if pays_n not in PAYS_FRANCE:
        # Pays vide : on accepte uniquement un CP français plausible (5 chiffres)
        if pays_n not in ("", "NAN", "NONE"):
            return None
        if not (cp_n.isdigit() and len(cp_n) == 5):
            return None

    # Corse
    if cp_n.startswith("2A"):
        return "2A"
    if cp_n.startswith("2B"):
        return "2B"
    if cp_n.startswith("20") and cp_n.isdigit() and len(cp_n) == 5:
        return "2A" if int(cp_n) < 20200 else "2B"

    # DOM
    if cp_n[:2] in ("97", "98") and len(cp_n) >= 3:
        return cp_n[:3]

    if len(cp_n) >= 2:
        return cp_n[:2]
    return None


def dept_label(code):
    if not code:
        return ""
    nom = DEPT_NOMS.get(code, "")
    return f"{code} — {nom}" if nom else str(code)


def _ptv_by_text(query):
    if not query:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-text",
                params={"searchText": query},
                headers=HEADERS,
                timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def _ptv_by_postal_code(cp, iso2):
    if not cp or not iso2:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-postal-code",
                params={"postalCode": cp, "countryCode": iso2},
                headers=HEADERS,
                timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code in (400, 404):
                return None
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def geocode_with_fallback(adresse_complete, ville, cp, pays):
    pays_full = PAYS_MAP.get(pays.upper(), pays) if pays else ""
    iso2      = PAYS_TO_ISO2.get(pays.upper(), pays.upper() if len(pays) == 2 else "")

    if ville and cp and pays_full:
        r = _ptv_by_text(f"{ville}, {cp}, {pays_full}")
        if r:
            return r
    if cp and iso2:
        r = _ptv_by_postal_code(cp, iso2)
        if r:
            return r
    if ville and pays_full:
        r = _ptv_by_text(f"{ville}, {pays_full}")
        if r:
            return r
    if adresse_complete:
        r = _ptv_by_text(adresse_complete)
        if r:
            return r
    return None


def build_address_string(row):
    def clean(v):
        v = str(v or "").strip()
        return "" if v.lower() in ("nan", "none") else v

    ville   = clean(row.get("localite",    ""))
    cp      = clean(row.get("code_postal", ""))
    pays    = clean(row.get("code_pays",   "")).upper()
    adresse = clean(row.get("adresse",     ""))
    numero  = clean(row.get("numero",      ""))
    nom     = clean(row.get("nom1",        ""))

    addr = parse_origin_from_parts(ville, cp, pays)
    rue  = " ".join(p for p in [numero, adresse] if p).strip()

    if rue and addr:
        addr = f"{rue}, {addr}"
    elif rue:
        addr = rue

    return addr if addr else nom


# ══════════════════════════════════════════════════════════════════
#  CALCUL ROUTE PTV
# ══════════════════════════════════════════════════════════════════

def calculate_route(coords_list):
    if len(coords_list) < 2:
        return None

    query_params = [
        ("profile", VEHICLE),
        ("results", "POLYLINE"),
    ]
    for i, (lat, lon) in enumerate(coords_list):
        if 0 < i < len(coords_list) - 1:
            query_params.append(("waypoints", f"{lat},{lon};radius=5000"))
        else:
            query_params.append(("waypoints", f"{lat},{lon}"))

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{PTV_BASE_URL}/routes",
                headers=HEADERS,
                params=query_params,
                timeout=30,
            )
            if resp.status_code != 200:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return None
            data = resp.json()
            return {
                "km":            round(data.get("distance", 0) / 1000, 1),
                "travel_time_h": round(data.get("travelTime", 0) / 3600, 2),
            }
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


# ══════════════════════════════════════════════════════════════════
#  PARSING FICHIER MISSIONS
# ══════════════════════════════════════════════════════════════════

ACTIVITE_KEYWORDS = {
    "charger":       "CHARGEMENT",
    "chargement":    "CHARGEMENT",
    "décharger":     "DECHARGEMENT",
    "dechargement":  "DECHARGEMENT",
    "déchargement":  "DECHARGEMENT",
    "decharger":     "DECHARGEMENT",
    "douane":        "DOUANE",
    "transit":       "DOUANE",
}


def normalize_activite(val):
    v = str(val).strip().lower()
    for kw, mapped in sorted(ACTIVITE_KEYWORDS.items(), key=lambda x: -len(x[0])):
        if kw in v:
            return mapped
    return str(val).strip().upper()


def _norm_col(s):
    s = str(s).strip().lower()
    for src, dst in [("é","e"),("è","e"),("ê","e"),("à","a"),("â","a"),
                     ("ô","o"),("û","u"),("î","i"),("ù","u"),("ç","c")]:
        s = s.replace(src, dst)
    return re.sub(r"[^a-z0-9]", "", s)


def detect_col(df, keywords):
    for col in df.columns:
        col_n = _norm_col(col)
        for kw in keywords:
            kw_n = _norm_col(kw)
            if kw_n == col_n or kw_n in col_n:
                return col
    return None


MISSIONS_COL_CANDIDATES = {
    "dossier":     ["N° Dossier", "N°Dossier", "N Dossier", "Dossier", "ndossier"],
    "activite":    ["Activité", "Activite", "Activité / Enregistrement"],
    "date":        ["Date"],
    "heure":       ["Heure"],
    "transport":   ["Type de transport", "Type transport"],
    "nom1":        ["Nom 1", "Nom1", "Nom"],
    "nom2":        ["Nom 2", "Nom2"],
    "adresse":     ["Adresse", "Address"],
    "numero":      ["Numéro", "Numero", "N°"],
    "code_pays":   ["Code pays", "Code Pays", "Pays", "Country"],
    "departement": ["Département", "Departement"],
    "code_postal": ["Code postal", "Code Postal", "Code Postal "],
    "localite":    ["Localité", "Localite", "Ville", "City"],
    "produit":     ["Produit"],
    "chauffeur":   ["Chauffeur", "Driver"],
    "tracteur":    ["Immat. tracteur", "Immat tracteur", "Immat.tracteur",
                    "Tracteur", "Immatriculation"],
    "remorque":    ["Remorque"],
}


def parse_missions(file):
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cols_lower = {_norm_col(c): c for c in df.columns}

    col_map = {}
    for role, candidates in MISSIONS_COL_CANDIDATES.items():
        found = None
        for cand in candidates:
            key = _norm_col(cand)
            if key in cols_lower:
                found = cols_lower[key]
                break
        if not found:
            found = detect_col(df, candidates)
        col_map[role] = found

    critiques = ["dossier", "activite", "date", "heure", "code_pays", "code_postal", "localite"]
    manquantes = [r for r in critiques if col_map.get(r) is None]
    if manquantes:
        st.warning(
            f"⚠️ Colonnes non détectées dans le fichier missions : {manquantes}\n"
            f"Colonnes disponibles : {list(df.columns)}"
        )

    rename = {v: k for k, v in col_map.items() if v}
    df = df.rename(columns=rename)

    for col in MISSIONS_COL_CANDIDATES.keys():
        if col not in df.columns:
            df[col] = ""

    df["dossier"] = df["dossier"].str.strip()
    df = df[df["dossier"].notna() & (df["dossier"] != "") & (df["dossier"] != "nan")]
    df = df[df["dossier"].str.match(r"^\d+", na=False)]

    df["activite_norm"] = df["activite"].apply(normalize_activite)

    import datetime as _dt

    def _combine(date_s, heure_s):
        date_s  = str(date_s  or "").strip()[:10]
        heure_s = str(heure_s or "").strip()
        if not date_s or date_s == "nan":
            return pd.NaT
        try:
            d = _dt.datetime.strptime(date_s, "%Y-%m-%d")
        except Exception:
            return pd.NaT
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                t = _dt.datetime.strptime(heure_s, fmt)
                return pd.Timestamp(d.replace(hour=t.hour, minute=t.minute, second=t.second))
            except Exception:
                pass
        return pd.Timestamp(d)

    df["datetime"]         = df.apply(lambda r: _combine(r.get("date", ""), r.get("heure", "")), axis=1)
    df["adresse_complete"] = df.apply(build_address_string, axis=1)

    return df


# ══════════════════════════════════════════════════════════════════
#  PARSING FICHIER CA
# ══════════════════════════════════════════════════════════════════

CA_COL_CANDIDATES = {
    "dossier":          ["N° Dossier", "N°Dossier", "Dossier"],
    "prix_transport":   ["Prix transport", "Prix Transport"],
    "total_vente":      ["Total des ventes", "Total ventes", "Total des vente"],
    "client":           ["Client facturation", "Client Facturation", "Client"],
    "etat_vente":       ["Etat vente", "État vente", "Etat"],
    "supplements":      ["Suppléments", "Supplements"],
    "sg":               ["S.G.", "SG"],
    "heures_attente":   ["Heures d'attente", "Heures attente"],
    "date_charg":       ["Date chargement", "Date Chargement"],
    "type_transport":   ["Type de transport", "Type transport"],
    "adr_charg":        ["Adresse chargement", "Adresse Chargement"],
    "localite_charg":   ["Localité chargement", "Localite chargement"],
    "cp_charg":         ["C.P. chargement", "CP chargement", "Code postal chargement"],
    "pays_charg":       ["Pays chargement", "Pays Chargement"],
    "adr_decharg":      ["Adresse déchargement", "Adresse dechargement"],
    "localite_decharg": ["Localité déchargement", "Localite dechargement"],
    "cp_decharg":       ["C.P. déchargement", "CP dechargement", "Code postal dechargement"],
    "pays_decharg":     ["Pays déchargement", "Pays dechargement"],
}


def parse_ca(file):
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cols_lower = {_norm_col(c): c for c in df.columns}

    col_map = {}
    for role, candidates in CA_COL_CANDIDATES.items():
        found = None
        for cand in candidates:
            key = _norm_col(cand)
            if key in cols_lower:
                found = cols_lower[key]
                break
        col_map[role] = found

    critiques_ca = ["dossier", "prix_transport", "total_vente"]
    manquantes_ca = [r for r in critiques_ca if col_map.get(r) is None]
    if manquantes_ca:
        st.warning(
            f"⚠️ Colonnes CA non détectées : {manquantes_ca} — "
            f"Colonnes disponibles : {list(df.columns)}"
        )

    rename = {v: k for k, v in col_map.items() if v}
    df = df.rename(columns=rename)

    for col in CA_COL_CANDIDATES.keys():
        if col not in df.columns:
            df[col] = ""

    df["dossier"] = df["dossier"].str.strip()
    df = df[df["dossier"].notna() & (df["dossier"] != "") & (df["dossier"] != "nan")]
    df = df[df["dossier"].str.match(r"^\d+", na=False)]

    def to_float(s):
        try:
            return float(
                str(s).replace(",", ".").replace("\xa0", "").replace(" ", "")
                .replace("€", "").strip()
            )
        except Exception:
            return 0.0

    df["prix_transport"] = df["prix_transport"].apply(to_float)
    df["total_vente"]    = df["total_vente"].apply(to_float)

    def _clean(v):
        v = str(v or "").strip()
        return "" if v.lower() in ("nan", "none") else v

    def _addr_from_ca(row, prefix):
        localite = _clean(row.get(f"localite_{prefix}", ""))
        cp       = _clean(row.get(f"cp_{prefix}", ""))
        pays     = _clean(row.get(f"pays_{prefix}", "")).upper()
        return parse_origin_from_parts(localite, cp, pays)

    df["adresse_charg_geo"]   = df.apply(lambda r: _addr_from_ca(r, "charg"),   axis=1)
    df["adresse_decharg_geo"] = df.apply(lambda r: _addr_from_ca(r, "decharg"), axis=1)

    df_agg = df.groupby("dossier", as_index=False).agg(
        prix_transport      = ("prix_transport",     "sum"),
        total_vente         = ("total_vente",        "sum"),
        client              = ("client",             "first"),
        etat_vente          = ("etat_vente",         "first"),
        adresse_charg_geo   = ("adresse_charg_geo",  "first"),
        adresse_decharg_geo = ("adresse_decharg_geo","first"),
        localite_charg      = ("localite_charg",     "first"),
        localite_decharg    = ("localite_decharg",   "first"),
        cp_charg            = ("cp_charg",           "first"),
        cp_decharg          = ("cp_decharg",         "first"),
        pays_charg          = ("pays_charg",         "first"),
        pays_decharg        = ("pays_decharg",       "first"),
    )

    return df_agg


# ══════════════════════════════════════════════════════════════════
#  CONSOLIDATION
# ══════════════════════════════════════════════════════════════════

def consolidate(df_missions, df_ca):
    rows = []

    def _c(v):
        v = str(v or "").strip()
        return "" if v.lower() in ("nan", "none") else v

    for dossier, grp in df_missions.groupby("dossier"):
        grp = grp.sort_values("datetime").reset_index(drop=True)

        chauffeur = next((v for v in grp["chauffeur"] if v and v not in ("nan", "")), "")
        tracteur  = next((v for v in grp["tracteur"]  if v and v not in ("nan", "")), "")
        remorque  = next(
            (str(v).strip() for v in grp["remorque"] if v and str(v).strip() not in ("nan", "")), ""
        ) if "remorque" in grp.columns else ""

        stops = []
        for _, r in grp.iterrows():
            stops.append({
                "activite":  r["activite_norm"],
                "datetime":  r["datetime"],
                "adresse":   r["adresse_complete"],
                "localite":  r.get("localite", ""),
                "nom":       r.get("nom1", ""),
                "ville_raw": str(r.get("localite",    "") or "").strip(),
                "cp_raw":    str(r.get("code_postal", "") or "").strip(),
                "pays_raw":  str(r.get("code_pays",   "") or "").strip().upper(),
            })

        # ── Stop de chargement / déchargement réels (fichier missions) ──
        stops_ch = [s for s in stops if s["activite"] == "CHARGEMENT"]
        stops_de = [s for s in stops if s["activite"] == "DECHARGEMENT"]
        s_ch = stops_ch[0]  if stops_ch else (stops[0]  if stops else None)
        s_de = stops_de[-1] if stops_de else (stops[-1] if stops else None)

        stop_labels = " → ".join(
            f"[{s['activite']}] {s['localite'] or s['nom'] or s['adresse']}"
            for s in stops
        )

        dates_valides = [s["datetime"] for s in stops if pd.notna(s["datetime"])]
        date_debut = min(dates_valides).strftime("%d/%m/%Y") if dates_valides else ""
        date_fin   = max(dates_valides).strftime("%d/%m/%Y") if dates_valides else ""

        rows.append({
            "dossier":            dossier,
            "chauffeur":          chauffeur,
            "tracteur":           tracteur,
            "remorque":           remorque,
            "date_debut":         date_debut,
            "date_fin":           date_fin,
            "nb_stops":           len(stops),
            "nb_dechargements":   len(stops_de),
            "stops_texte":        stop_labels,
            "stops_data":         stops,
            "loc_charg_stop":     _c(s_ch["ville_raw"]) if s_ch else "",
            "cp_charg_stop":      _c(s_ch["cp_raw"])    if s_ch else "",
            "pays_charg_stop":    _c(s_ch["pays_raw"])  if s_ch else "",
            "loc_decharg_stop":   _c(s_de["ville_raw"]) if s_de else "",
            "cp_decharg_stop":    _c(s_de["cp_raw"])    if s_de else "",
            "pays_decharg_stop":  _c(s_de["pays_raw"])  if s_de else "",
        })

    df_cons = pd.DataFrame(rows)

    ca_cols = [
        "dossier", "prix_transport", "total_vente", "client", "etat_vente",
        "adresse_charg_geo", "adresse_decharg_geo",
        "localite_charg", "localite_decharg",
        "cp_charg", "cp_decharg", "pays_charg", "pays_decharg",
    ]
    ca_cols_dispo = [c for c in ca_cols if c in df_ca.columns]
    df_cons = df_cons.merge(df_ca[ca_cols_dispo], on="dossier", how="left")
    df_cons["prix_transport"] = df_cons["prix_transport"].fillna(0.0)
    df_cons["total_vente"]    = df_cons["total_vente"].fillna(0.0)

    def _fill_from_stops(row):
        loc_ch  = _c(row.get("localite_charg",   ""))
        cp_ch   = _c(row.get("cp_charg",         ""))
        pays_ch = _c(row.get("pays_charg",       ""))
        loc_de  = _c(row.get("localite_decharg", ""))
        cp_de   = _c(row.get("cp_decharg",       ""))
        pays_de = _c(row.get("pays_decharg",     ""))

        if not (loc_ch or cp_ch):
            loc_ch  = _c(row.get("loc_charg_stop",  ""))
            cp_ch   = _c(row.get("cp_charg_stop",   ""))
            pays_ch = _c(row.get("pays_charg_stop", ""))
        if not (loc_de or cp_de):
            loc_de  = _c(row.get("loc_decharg_stop",  ""))
            cp_de   = _c(row.get("cp_decharg_stop",   ""))
            pays_de = _c(row.get("pays_decharg_stop", ""))

        return pd.Series({
            "localite_charg":   loc_ch,
            "cp_charg":         cp_ch,
            "pays_charg":       pays_ch,
            "localite_decharg": loc_de,
            "cp_decharg":       cp_de,
            "pays_decharg":     pays_de,
        })

    filled = df_cons.apply(_fill_from_stops, axis=1)
    df_cons["localite_charg"]   = filled["localite_charg"]
    df_cons["cp_charg"]         = filled["cp_charg"]
    df_cons["pays_charg"]       = filled["pays_charg"]
    df_cons["localite_decharg"] = filled["localite_decharg"]
    df_cons["cp_decharg"]       = filled["cp_decharg"]
    df_cons["pays_decharg"]     = filled["pays_decharg"]

    # ── Départements de déchargement (2 sources) ──
    df_cons["dept_ca"] = df_cons.apply(
        lambda r: extract_departement(r.get("cp_decharg", ""), r.get("pays_decharg", "")), axis=1
    )
    df_cons["dept_stop"] = df_cons.apply(
        lambda r: extract_departement(r.get("cp_decharg_stop", ""), r.get("pays_decharg_stop", "")), axis=1
    )

    return df_cons


# ══════════════════════════════════════════════════════════════════
#  STATS PAR DÉPARTEMENT
# ══════════════════════════════════════════════════════════════════

def build_dept_stats(df):
    """
    Agrège par département de déchargement.
    Fonctionne avec ou sans les colonnes km_total / km_vide.
    """
    if df is None or df.empty or "dept_decharg" not in df.columns:
        return pd.DataFrame()

    d = df[df["dept_decharg"].notna() & (df["dept_decharg"].astype(str) != "")].copy()
    if d.empty:
        return pd.DataFrame()

    for c in ("km_total", "km_vide"):
        if c not in d.columns:
            d[c] = np.nan
        d[c] = pd.to_numeric(d[c], errors="coerce")

    for c in ("prix_transport", "total_vente"):
        d[c] = pd.to_numeric(d.get(c), errors="coerce").fillna(0.0)

    g = d.groupby("dept_decharg", as_index=False).agg(
        nb_dossiers    = ("dossier",        "count"),
        km_charges     = ("km_total",       "sum"),
        km_vide        = ("km_vide",        "sum"),
        prix_transport = ("prix_transport", "sum"),
        total_vente    = ("total_vente",    "sum"),
    )

    g["km_charges"]    = g["km_charges"].round(0)
    g["km_vide"]       = g["km_vide"].round(0)
    g["km_complet"]    = g["km_charges"] + g["km_vide"]
    g["pct_vide"]      = (g["km_vide"] / g["km_complet"].replace(0, np.nan) * 100).round(1)
    g["renta_charge"]  = (g["total_vente"] / g["km_charges"].replace(0, np.nan)).round(2)
    g["renta_complet"] = (g["total_vente"] / g["km_complet"].replace(0, np.nan)).round(2)
    g["ca_moyen"]      = (g["total_vente"] / g["nb_dossiers"].replace(0, np.nan)).round(0)
    g["km_moyen"]      = (g["km_complet"] / g["nb_dossiers"].replace(0, np.nan)).round(0)
    g["nom_dept"]      = g["dept_decharg"].map(lambda c: DEPT_NOMS.get(c, "—"))

    return g.sort_values("total_vente", ascending=False).reset_index(drop=True)


DEPT_COL_RENAME = {
    "dept_decharg":   "Dépt",
    "nom_dept":       "Département",
    "nb_dossiers":    "Nb Dossiers",
    "km_charges":     "KM Chargés",
    "km_vide":        "KM À Vide",
    "km_complet":     "KM Complet",
    "pct_vide":       "% KM Vide",
    "km_moyen":       "KM moy/dossier",
    "prix_transport": "Prix Transport (€)",
    "total_vente":    "Total Vente (€)",
    "ca_moyen":       "CA moy/dossier (€)",
    "renta_charge":   "Renta €/km chargé",
    "renta_complet":  "Renta €/km complet",
}

DEPT_COL_ORDER = [
    "dept_decharg", "nom_dept", "nb_dossiers",
    "km_charges", "km_vide", "km_complet", "pct_vide", "km_moyen",
    "prix_transport", "total_vente", "ca_moyen",
    "renta_charge", "renta_complet",
]


def format_dept_table(df_dept, avec_km=True):
    if df_dept.empty:
        return df_dept
    cols = DEPT_COL_ORDER if avec_km else [
        c for c in DEPT_COL_ORDER
        if c not in ("km_charges", "km_vide", "km_complet", "pct_vide",
                     "km_moyen", "renta_charge", "renta_complet")
    ]
    cols = [c for c in cols if c in df_dept.columns]
    return df_dept[cols].rename(columns=DEPT_COL_RENAME)


# ══════════════════════════════════════════════════════════════════
#  CALCUL KM PAR CHAUFFEUR (PTV)
# ══════════════════════════════════════════════════════════════════

def compute_ptv_for_driver(df_cons, chauffeur, progress_cb=None):
    df_ch = df_cons[df_cons["chauffeur"] == chauffeur].copy()
    df_ch["_sort_date"] = pd.to_datetime(df_ch["date_debut"], format="%d/%m/%Y", errors="coerce")
    df_ch = df_ch.sort_values("_sort_date").reset_index(drop=True)

    def _c(v):
        v = str(v or "").strip()
        return "" if v.lower() in ("nan", "none") else v

    results          = []
    dossier_sequences = {}

    for _, row in df_ch.iterrows():
        dos = row["dossier"]

        addr_ch  = _c(row.get("adresse_charg_geo",  ""))
        addr_de  = _c(row.get("adresse_decharg_geo", ""))
        loc_ch   = _c(row.get("localite_charg",      ""))
        loc_de   = _c(row.get("localite_decharg",    ""))
        cp_ch    = _c(row.get("cp_charg",            ""))
        cp_de    = _c(row.get("cp_decharg",          ""))
        pays_ch  = _c(row.get("pays_charg",          ""))
        pays_de  = _c(row.get("pays_decharg",        ""))

        stops_mid = [
            s for s in row.get("stops_data", [])
            if s.get("activite", "") in ("DOUANE",)
        ]

        dossier_sequences[dos] = {
            "addr_ch":    addr_ch,  "loc_ch":  loc_ch,  "cp_ch":  cp_ch,  "pays_ch":  pays_ch,
            "addr_de":    addr_de,  "loc_de":  loc_de,  "cp_de":  cp_de,  "pays_de":  pays_de,
            "stops_mid":  stops_mid,
            "date_debut": row["date_debut"],
        }

    # ── Géocodage ──
    geo_cache        = {}
    points_to_geocode = {}

    for dos, seq in dossier_sequences.items():
        if seq["loc_ch"] or seq["cp_ch"]:
            points_to_geocode[(seq["loc_ch"], seq["cp_ch"], seq["pays_ch"])] = True
        if seq["loc_de"] or seq["cp_de"]:
            points_to_geocode[(seq["loc_de"], seq["cp_de"], seq["pays_de"])] = True
        for s in seq["stops_mid"]:
            k = (s.get("ville_raw", ""), s.get("cp_raw", ""), s.get("pays_raw", ""))
            if k[0] or k[1]:
                points_to_geocode[k] = True

    total_geo = len(points_to_geocode)
    for i, (ville_r, cp_r, pays_r) in enumerate(points_to_geocode.keys()):
        label = f"{ville_r} {cp_r}".strip()
        if progress_cb:
            progress_cb(f"🌍 Géocodage {i+1}/{total_geo} : {label}...")
        addr_display = parse_origin_from_parts(ville_r, cp_r, pays_r)
        coords = geocode_with_fallback(addr_display, ville_r, cp_r, pays_r)
        geo_cache[(ville_r, cp_r, pays_r)] = coords
        if coords is None:
            st.warning(f"⚠️ Géocodage échoué : {addr_display}")

    # ── KM chargés par dossier ──
    dossier_km = {}
    for dos, seq in dossier_sequences.items():
        coords_seq = []

        c_ch = geo_cache.get((seq["loc_ch"], seq["cp_ch"], seq["pays_ch"]))
        if c_ch:
            coords_seq.append(c_ch)

        for s in seq["stops_mid"]:
            c = geo_cache.get((s.get("ville_raw", ""), s.get("cp_raw", ""), s.get("pays_raw", "")))
            if c:
                coords_seq.append(c)

        c_de = geo_cache.get((seq["loc_de"], seq["cp_de"], seq["pays_de"]))
        if c_de:
            coords_seq.append(c_de)

        if len(coords_seq) >= 2:
            if progress_cb:
                progress_cb(f"📍 Calcul km dossier {dos} ({seq['loc_ch']} → {seq['loc_de']})...")
            res = calculate_route(coords_seq)
            dossier_km[dos] = res["km"] if res else None
        else:
            dossier_km[dos] = None

    # ── KM à vide inter-dossiers ──
    dossiers_ordonnes = sorted(
        dossier_sequences.keys(),
        key=lambda d: pd.to_datetime(
            dossier_sequences[d]["date_debut"], format="%d/%m/%Y", errors="coerce"
        ),
    )

    empty_legs = []
    for i in range(len(dossiers_ordonnes) - 1):
        dos_actuel  = dossiers_ordonnes[i]
        dos_suivant = dossiers_ordonnes[i + 1]
        seq_act     = dossier_sequences[dos_actuel]
        seq_suiv    = dossier_sequences[dos_suivant]

        coords_fin   = geo_cache.get((seq_act["loc_de"],  seq_act["cp_de"],  seq_act["pays_de"]))
        coords_debut = geo_cache.get((seq_suiv["loc_ch"], seq_suiv["cp_ch"], seq_suiv["pays_ch"]))

        if coords_fin and coords_debut:
            if progress_cb:
                progress_cb(f"⚡ Km à vide : {seq_act['loc_de']} → {seq_suiv['loc_ch']}...")
            res     = calculate_route([coords_fin, coords_debut])
            km_vide = res["km"] if res else None
        else:
            km_vide = None

        empty_legs.append({
            "dossier_depart":  dos_actuel,
            "dossier_arrivee": dos_suivant,
            "from_addr":       seq_act["addr_de"],
            "from_localite":   seq_act["loc_de"],
            "to_addr":         seq_suiv["addr_ch"],
            "to_localite":     seq_suiv["loc_ch"],
            "km_vide":         km_vide,
        })

    # ── Assemblage ──
    for _, row in df_ch.iterrows():
        dos = row["dossier"]

        km_vide_total = sum(
            leg["km_vide"] for leg in empty_legs
            if leg["dossier_depart"] == dos and leg["km_vide"] is not None
        )
        vide_details = [leg for leg in empty_legs if leg["dossier_depart"] == dos]

        results.append({
            "dossier":          dos,
            "chauffeur":        chauffeur,
            "tracteur":         row.get("tracteur", ""),
            "remorque":         row.get("remorque", ""),
            "date_debut":       row["date_debut"],
            "date_fin":         row["date_fin"],
            "client":           row.get("client",      ""),
            "etat_vente":       row.get("etat_vente",  ""),
            "stops_texte":      row["stops_texte"],
            "nb_stops":         row["nb_stops"],
            "localite_charg":   row.get("localite_charg",   ""),
            "localite_decharg": row.get("localite_decharg", ""),
            "cp_decharg":       row.get("cp_decharg",       ""),
            "pays_decharg":     row.get("pays_decharg",     ""),
            "dept_decharg":     row.get("dept_decharg",     None),
            "prix_transport":   row["prix_transport"],
            "total_vente":      row["total_vente"],
            "km_total":         dossier_km.get(dos),
            "km_vide":          km_vide_total if km_vide_total > 0 else None,
            "vide_details":     vide_details,
        })

    return results


# ══════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════

def export_excel(df_result, df_vide, df_dept=None):
    output = io.BytesIO()

    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            # ── Feuille principale ──
            col_rename = {
                "chauffeur":        "Chauffeur",
                "tracteur":         "Tracteur",
                "remorque":         "Remorque",
                "dossier":          "N° Dossier",
                "date_debut":       "Date début",
                "date_fin":         "Date fin",
                "client":           "Client",
                "etat_vente":       "État vente",
                "localite_decharg": "Ville déchargement",
                "dept_decharg":     "Dépt déch.",
                "nb_stops":         "Nb stops",
                "stops_texte":      "Séquence stops",
                "km_total":         "KM Chargés",
                "km_vide":          "KM À Vide",
                "prix_transport":   "Prix Transport (€)",
                "total_vente":      "Total Vente (€)",
            }
            cols_dispo = [c for c in col_rename if c in df_result.columns]
            df_export  = df_result[cols_dispo].copy()

            if "km_total" in df_result.columns and "km_vide" in df_result.columns:
                df_export["km_complet"]  = df_result["km_total"].fillna(0) + df_result["km_vide"].fillna(0)
                df_export["rentabilite"] = (
                    df_result["total_vente"] / df_export["km_complet"].replace(0, np.nan)
                ).round(2)
                col_rename["km_complet"]  = "KM Complet"
                col_rename["rentabilite"] = "Rentabilité €/km"
                cols_dispo = [c for c in col_rename if c in df_export.columns]
                df_export  = df_export[cols_dispo]

            df_export = df_export.rename(columns=col_rename).fillna("")
            df_export.to_excel(writer, sheet_name="Missions & CA", index=False)
            _style_sheet(writer.sheets["Missions & CA"], len(df_export))

            # ── Rentabilité par département ──
            if df_dept is not None and not df_dept.empty:
                df_dept_exp = format_dept_table(df_dept, avec_km=True).fillna("")
                df_dept_exp.to_excel(writer, sheet_name="Renta Départements", index=False)
                _style_sheet(writer.sheets["Renta Départements"], len(df_dept_exp))

            # ── Résumé chauffeurs ──
            df_resume = df_result.groupby("chauffeur", as_index=False).agg(
                Dossiers       = ("dossier",        "count"),
                KM_Charges     = ("km_total",        "sum"),
                KM_Vide        = ("km_vide",          "sum"),
                Prix_Transport = ("prix_transport",   "sum"),
                Total_Vente    = ("total_vente",      "sum"),
            ).round(1)
            df_resume["KM Complet"]       = df_resume["KM_Charges"] + df_resume["KM_Vide"]
            df_resume["% KM Vide"]        = (
                df_resume["KM_Vide"] / df_resume["KM Complet"].replace(0, np.nan) * 100
            ).round(1)
            df_resume["Rentabilité €/km"] = (
                df_resume["Total_Vente"] / df_resume["KM Complet"].replace(0, np.nan)
            ).round(2)
            df_resume = df_resume.rename(columns={
                "chauffeur":      "Chauffeur",
                "Dossiers":       "Nb Dossiers",
                "KM_Charges":     "KM Chargés",
                "KM_Vide":        "KM À Vide",
                "Prix_Transport": "Prix Transport (€)",
                "Total_Vente":    "Total Vente (€)",
            }).fillna("")
            df_resume.to_excel(writer, sheet_name="Résumé Chauffeurs", index=False)
            _style_sheet(writer.sheets["Résumé Chauffeurs"], len(df_resume))

            # ── KM à vide détail ──
            if not df_vide.empty:
                vide_rename = {
                    "chauffeur":       "Chauffeur",
                    "dossier_depart":  "Dossier départ",
                    "dossier_arrivee": "Dossier arrivée",
                    "from_localite":   "Ville départ",
                    "to_localite":     "Ville arrivée",
                    "km_vide":         "KM à vide",
                }
                df_vide_exp = df_vide[[c for c in vide_rename if c in df_vide.columns]].copy()
                df_vide_exp = df_vide_exp.rename(columns=vide_rename).fillna("")
                df_vide_exp.to_excel(writer, sheet_name="KM À Vide Détail", index=False)
                _style_sheet(writer.sheets["KM À Vide Détail"], len(df_vide_exp))

    except Exception as e:
        st.error(f"❌ Erreur génération Excel : {e}")
        return b""

    return output.getvalue()


def _style_sheet(ws, nb_rows):
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, nb_rows + 2):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = ALT_FILL

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


# ══════════════════════════════════════════════════════════════════
#  INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Missions & CA + KM PTV", page_icon="📦", layout="wide")

st.title("📦 Analyse Missions + CA + Calcul KM")
st.caption("Consolide les missions, le chiffre d'affaires et calcule les kilomètres via PTV.")

if not PTV_API_KEY or PTV_API_KEY == "METS_TA_CLE_ICI":
    st.error("⚠️ Clé PTV_API_KEY non configurée. Le calcul de distances ne fonctionnera pas.")

st.divider()

# ── Upload ──
col_up1, col_up2 = st.columns(2)
with col_up1:
    st.markdown("#### 📋 Fichier Missions")
    file_missions = st.file_uploader("Export missions (.xlsx)", type=["xlsx"], key="missions")
with col_up2:
    st.markdown("#### 💶 Fichier CA")
    file_ca = st.file_uploader("Export CA (.xlsx)", type=["xlsx"], key="ca")

st.divider()

# ── Parsing & Consolidation ──
if file_missions and file_ca:

    with st.spinner("📂 Lecture des fichiers..."):
        try:
            df_missions = parse_missions(file_missions)
            df_ca_raw   = parse_ca(file_ca)
        except Exception as e:
            st.error(f"❌ Erreur lecture fichiers : {e}")
            st.stop()

    df_cons = consolidate(df_missions, df_ca_raw)

    df_cons["_date_dt"] = pd.to_datetime(df_cons["date_debut"], format="%d/%m/%Y", errors="coerce")
    df_cons_f = df_cons.copy()

    MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
               "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
    dates_valides = df_cons["_date_dt"].dropna()
    if not dates_valides.empty:
        d_min = dates_valides.min()
        d_max = dates_valides.max()
        if d_min.month == d_max.month and d_min.year == d_max.year:
            periode_label = f"{MOIS_FR[d_min.month]} {d_min.year}"
        else:
            periode_label = f"{d_min.strftime('%d/%m/%Y')} → {d_max.strftime('%d/%m/%Y')}"
    else:
        periode_label = "Période inconnue"

    # ── KPIs globaux ──
    st.markdown(f"### 📊 Aperçu — {periode_label}")
    ca_total   = df_cons_f["total_vente"].sum()
    prix_total = df_cons_f["prix_transport"].sum()
    ca_moy     = ca_total / len(df_cons_f) if len(df_cons_f) > 0 else 0

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("📁 Dossiers",       len(df_cons_f))
    k2.metric("👨🏻‍✈️ Chauffeurs",     df_cons_f["chauffeur"].nunique())
    k3.metric("📍 Stops totaux",   int(df_cons_f["nb_stops"].sum()))
    k4.metric("💶 Prix Transport", f"{prix_total:,.0f} €")
    k5.metric("💶 Total Ventes",   f"{ca_total:,.0f} €")
    k6.metric("📈 CA moy/dossier", f"{ca_moy:,.0f} €")

    st.divider()

    # ══════════════════════════════════════════════════════════
    #  SOURCE DU DÉPARTEMENT DE DÉCHARGEMENT
    # ══════════════════════════════════════════════════════════
    st.markdown("### 🇫🇷 Département de déchargement")

    src_c1, src_c2 = st.columns([2, 3])
    with src_c1:
        source_dept = st.radio(
            "Source du lieu de déchargement :",
            options=["Fichier CA", "Dernier stop DÉCHARGEMENT (missions)"],
            index=0,
            horizontal=False,
            help=(
                "« Fichier CA » = colonne C.P. déchargement du fichier CA (avec repli sur "
                "les stops si vide). « Dernier stop » = dernière ligne DÉCHARGEMENT du "
                "fichier missions — plus fiable pour les tournées multi-livraisons."
            ),
        )

    if source_dept.startswith("Fichier CA"):
        df_cons_f["dept_decharg"] = df_cons_f["dept_ca"]
    else:
        df_cons_f["dept_decharg"] = df_cons_f["dept_stop"].fillna(df_cons_f["dept_ca"])

    df_cons_f["dept_label"] = df_cons_f["dept_decharg"].map(
        lambda c: dept_label(c) if c else "🌍 Hors France / inconnu"
    )

    nb_fr   = int(df_cons_f["dept_decharg"].notna().sum())
    nb_hors = len(df_cons_f) - nb_fr
    with src_c2:
        st.info(
            f"🇫🇷 **{nb_fr} dossiers** déchargés en France "
            f"(**{df_cons_f['dept_decharg'].nunique()} départements**) · "
            f"🌍 {nb_hors} hors France ou CP non exploitable."
        )
        if nb_hors:
            with st.expander(f"🔎 Voir les {nb_hors} dossiers sans département"):
                st.dataframe(
                    df_cons_f[df_cons_f["dept_decharg"].isna()][
                        ["dossier", "chauffeur", "localite_decharg", "cp_decharg",
                         "pays_decharg", "total_vente"]
                    ].rename(columns={
                        "dossier":          "N° Dossier",
                        "chauffeur":        "Chauffeur",
                        "localite_decharg": "Ville déch.",
                        "cp_decharg":       "CP",
                        "pays_decharg":     "Pays",
                        "total_vente":      "Total Vente (€)",
                    }),
                    use_container_width=True,
                    height=250,
                )

    # ── Aperçu CA par département (avant PTV) ──
    df_dept_ca = build_dept_stats(df_cons_f)
    if not df_dept_ca.empty:
        with st.expander("💶 CA par département (sans km — lance le PTV pour la rentabilité)"):
            st.dataframe(
                format_dept_table(df_dept_ca, avec_km=False),
                use_container_width=True,
                height=380,
            )

    st.divider()

    # ══════════════════════════════════════════════════════════
    #  TABLEAU CONSOLIDÉ
    # ══════════════════════════════════════════════════════════
    st.markdown("### 📋 Tableau consolidé")

    # ── Listes pour les filtres ──
    chauffeurs_dispo = sorted([
        c for c in df_cons_f["chauffeur"].dropna().unique()
        if c and c != "nan"
    ])
    remorques_dispo = sorted([
        str(r).strip() for r in df_cons_f["remorque"].dropna().unique()
        if str(r).strip() and str(r).strip() not in ("nan", "")
    ]) if "remorque" in df_cons_f.columns else []
    tracteurs_dispo = sorted([
        str(t).strip() for t in df_cons_f["tracteur"].dropna().unique()
        if str(t).strip() and str(t).strip() not in ("nan", "")
    ]) if "tracteur" in df_cons_f.columns else []
    depts_dispo = sorted([
        d for d in df_cons_f["dept_decharg"].dropna().unique() if d
    ])

    # ── Filtres ──
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        filtre_chauffeur = st.multiselect(
            "👨🏻‍✈️ Chauffeur :",
            options=chauffeurs_dispo,
            default=[],
            placeholder="Tous les chauffeurs",
        )
    with fc2:
        filtre_tracteur = st.multiselect(
            "🚜 Tracteur :",
            options=tracteurs_dispo,
            default=[],
            placeholder="Tous les tracteurs",
        )
    with fc3:
        filtre_remorque = st.multiselect(
            "🚛 Remorque :",
            options=remorques_dispo,
            default=[],
            placeholder="Toutes les remorques",
        )
    with fc4:
        filtre_dept = st.multiselect(
            "🇫🇷 Département déch. :",
            options=depts_dispo,
            default=[],
            format_func=dept_label,
            placeholder="Tous les départements",
        )

    # ── Application des filtres ──
    df_display = df_cons_f.copy()
    if filtre_chauffeur:
        df_display = df_display[df_display["chauffeur"].isin(filtre_chauffeur)]
    if filtre_remorque and "remorque" in df_display.columns:
        df_display = df_display[df_display["remorque"].isin(filtre_remorque)]
    if filtre_tracteur and "tracteur" in df_display.columns:
        df_display = df_display[df_display["tracteur"].isin(filtre_tracteur)]
    if filtre_dept:
        df_display = df_display[df_display["dept_decharg"].isin(filtre_dept)]

    # ── KPIs de la sélection ──
    if filtre_chauffeur or filtre_remorque or filtre_tracteur or filtre_dept:
        _sel_label = []
        if filtre_chauffeur:
            _sel_label.append(f"{len(filtre_chauffeur)} chauffeur(s)")
        if filtre_remorque:
            _sel_label.append(f"{len(filtre_remorque)} remorque(s)")
        if filtre_tracteur:
            _sel_label.append(f"{len(filtre_tracteur)} tracteur(s)")
        if filtre_dept:
            _sel_label.append(f"{len(filtre_dept)} département(s)")
        st.markdown(f"##### 📊 Aperçu — {', '.join(_sel_label)}")

        fk1, fk2, fk3, fk4, fk5, fk6 = st.columns(6)
        _tv = df_display["total_vente"].sum()
        _pt = df_display["prix_transport"].sum()
        _nd = len(df_display)
        fk1.metric("📁 Dossiers",       _nd)
        fk2.metric("📍 Stops",          int(df_display["nb_stops"].sum()))
        fk3.metric("💶 Prix Transport", f"{_pt:,.0f} €")
        fk4.metric("💶 Total Ventes",   f"{_tv:,.0f} €")
        fk5.metric("📈 CA moy/dossier", f"{(_tv / _nd if _nd else 0):,.0f} €")

        if "df_result" in st.session_state:
            _dr_f = st.session_state["df_result"].copy()
            _dr_f["dept_decharg"] = _dr_f["dossier"].map(
                df_cons_f.set_index("dossier")["dept_decharg"]
            )
            if filtre_chauffeur:
                _dr_f = _dr_f[_dr_f["chauffeur"].isin(filtre_chauffeur)]
            if filtre_remorque and "remorque" in _dr_f.columns:
                _dr_f = _dr_f[_dr_f["remorque"].isin(filtre_remorque)]
            if filtre_tracteur and "tracteur" in _dr_f.columns:
                _dr_f = _dr_f[_dr_f["tracteur"].isin(filtre_tracteur)]
            if filtre_dept:
                _dr_f = _dr_f[_dr_f["dept_decharg"].isin(filtre_dept)]
            _km   = pd.to_numeric(_dr_f["km_total"], errors="coerce").fillna(0).sum() + \
                    pd.to_numeric(_dr_f["km_vide"],  errors="coerce").fillna(0).sum()
            _rent = _dr_f["total_vente"].sum() / _km if _km > 0 else 0
            fk6.metric("⚡ Rentabilité", f"{_rent:.2f} €/km")
        else:
            fk6.metric("⚡ Rentabilité", "— (après PTV)")

    cols_show = [
        "dossier", "chauffeur", "tracteur", "remorque",
        "date_debut", "date_fin", "client", "etat_vente",
        "localite_decharg", "dept_label",
        "nb_stops", "stops_texte", "prix_transport", "total_vente",
    ]
    st.dataframe(
        df_display[[c for c in cols_show if c in df_display.columns]].rename(columns={
            "dossier":          "N° Dossier",
            "chauffeur":        "Chauffeur",
            "tracteur":         "Tracteur",
            "remorque":         "Remorque",
            "date_debut":       "Date début",
            "date_fin":         "Date fin",
            "client":           "Client",
            "etat_vente":       "État vente",
            "localite_decharg": "Ville déchargement",
            "dept_label":       "Département déch.",
            "nb_stops":         "Nb stops",
            "stops_texte":      "Séquence stops",
            "prix_transport":   "Prix Transport (€)",
            "total_vente":      "Total Vente (€)",
        }),
        use_container_width=True,
        height=400,
    )

    st.divider()

    # ══════════════════════════════════════════════════════════
    #  CALCUL KM via PTV
    # ══════════════════════════════════════════════════════════
    st.markdown("### 🗺️ Calcul KM via PTV")

    ptv_c1, ptv_c2, ptv_c3, ptv_c4 = st.columns(4)
    with ptv_c1:
        chauffeurs_ptv = st.multiselect(
            "🚛 Chauffeurs :",
            options=chauffeurs_dispo,
            default=[],
            placeholder="Sélectionner des chauffeurs...",
        )
    with ptv_c2:
        remorques_ptv = st.multiselect(
            "🔗 Remorques :",
            options=remorques_dispo,
            default=[],
            placeholder="Toutes les remorques",
        )
    with ptv_c3:
        tracteurs_ptv = st.multiselect(
            "🚜 Tracteurs :",
            options=tracteurs_dispo,
            default=[],
            placeholder="Tous les tracteurs",
        )
    with ptv_c4:
        depts_ptv = st.multiselect(
            "🇫🇷 Départements :",
            options=depts_dispo,
            default=[],
            format_func=dept_label,
            placeholder="Tous les départements",
        )

    # ── Construire la liste des chauffeurs à calculer ──
    chauffeurs_a_calculer = list(chauffeurs_ptv)

    if remorques_ptv and "remorque" in df_cons_f.columns:
        ch_remorque = (
            df_cons_f[df_cons_f["remorque"].isin(remorques_ptv)]["chauffeur"]
            .dropna().unique().tolist()
        )
        chauffeurs_a_calculer = list(set(chauffeurs_a_calculer + ch_remorque))

    if tracteurs_ptv and "tracteur" in df_cons_f.columns:
        ch_tracteur = (
            df_cons_f[df_cons_f["tracteur"].isin(tracteurs_ptv)]["chauffeur"]
            .dropna().unique().tolist()
        )
        chauffeurs_a_calculer = list(set(chauffeurs_a_calculer + ch_tracteur))

    if depts_ptv:
        ch_dept = (
            df_cons_f[df_cons_f["dept_decharg"].isin(depts_ptv)]["chauffeur"]
            .dropna().unique().tolist()
        )
        chauffeurs_a_calculer = list(set(chauffeurs_a_calculer + ch_dept))

    chauffeurs_a_calculer = [c for c in chauffeurs_a_calculer if c and c != "nan"]
    nb_dossiers_ptv = len(df_cons_f[df_cons_f["chauffeur"].isin(chauffeurs_a_calculer)])

    if chauffeurs_a_calculer:
        extras = []
        if remorques_ptv:
            extras.append(f"{len(remorques_ptv)} remorque(s)")
        if tracteurs_ptv:
            extras.append(f"{len(tracteurs_ptv)} tracteur(s)")
        if depts_ptv:
            extras.append(f"{len(depts_ptv)} département(s)")
        extras_str = f", {', '.join(extras)}" if extras else ""
        st.info(
            f"ℹ️ Calcul pour **{nb_dossiers_ptv} dossiers** "
            f"({len(chauffeurs_a_calculer)} chauffeur(s){extras_str})."
        )
        if depts_ptv:
            st.caption(
                "⚠️ Le calcul PTV porte sur **tous** les dossiers des chauffeurs concernés "
                "(nécessaire pour chaîner les km à vide). Le filtre département s'applique "
                "ensuite à l'analyse."
            )

    btn_ptv = st.button(
        "🚀 Lancer le calcul PTV",
        disabled=(not chauffeurs_a_calculer),
        type="primary",
    )

    if btn_ptv and chauffeurs_a_calculer:
        all_results = []
        all_vide    = []

        progress_bar = st.progress(0)
        status_text  = st.empty()
        total_ch     = len(chauffeurs_a_calculer)

        for ch_idx, chauffeur in enumerate(chauffeurs_a_calculer):
            status_text.text(f"⏳ Chauffeur {ch_idx+1}/{total_ch} : {chauffeur}")

            def _progress(msg):
                status_text.text(msg)

            try:
                res = compute_ptv_for_driver(df_cons_f, chauffeur, progress_cb=_progress)
                all_results.extend(res)

                for r in res:
                    for leg in r.get("vide_details", []):
                        all_vide.append({
                            "chauffeur":       chauffeur,
                            "dossier_depart":  leg["dossier_depart"],
                            "dossier_arrivee": leg["dossier_arrivee"],
                            "from_localite":   leg["from_localite"],
                            "to_localite":     leg["to_localite"],
                            "km_vide":         leg["km_vide"],
                        })
            except Exception as e:
                st.error(f"❌ Erreur chauffeur {chauffeur} : {e}")

            progress_bar.progress(int((ch_idx + 1) / total_ch * 100))

        status_text.success("✅ Calcul PTV terminé !")
        progress_bar.progress(100)

        st.session_state["df_result"] = pd.DataFrame(all_results)
        st.session_state["df_vide"]   = pd.DataFrame(all_vide)

    # ── Affichage résultats PTV ──
    if "df_result" in st.session_state and not st.session_state["df_result"].empty:
        df_result = st.session_state["df_result"].copy()
        df_vide   = st.session_state.get("df_vide", pd.DataFrame())

        # Re-mappe le département selon la source choisie (sans relancer le PTV)
        _map_dept = df_cons_f.set_index("dossier")["dept_decharg"]
        df_result["dept_decharg"] = df_result["dossier"].map(_map_dept)
        df_result["dept_label"]   = df_result["dept_decharg"].map(
            lambda c: dept_label(c) if c else "🌍 Hors France"
        )

        # Filtre département appliqué aux résultats
        if depts_ptv:
            df_result_f = df_result[df_result["dept_decharg"].isin(depts_ptv)].copy()
        else:
            df_result_f = df_result.copy()

        st.divider()
        st.markdown("### 📈 Résultats KM")
        if depts_ptv:
            st.caption(
                f"Filtré sur {len(depts_ptv)} département(s) de déchargement — "
                f"{len(df_result_f)} dossier(s) sur {len(df_result)}."
            )

        km_total_sum    = pd.to_numeric(df_result_f["km_total"], errors="coerce").fillna(0).sum()
        km_vide_sum     = pd.to_numeric(df_result_f["km_vide"],  errors="coerce").fillna(0).sum()
        _km_complet     = km_total_sum + km_vide_sum
        pct_vide        = (km_vide_sum / _km_complet * 100) if _km_complet > 0 else 0
        _ca_ptv         = df_result_f["total_vente"].sum()
        _rent_ptv       = _ca_ptv / _km_complet if _km_complet > 0 else 0

        kp1, kp2, kp3, kp4, kp5, kp6 = st.columns(6)
        kp1.metric("📏 KM Chargés",       f"{km_total_sum:,.0f} km")
        kp2.metric("⚡ KM À Vide",         f"{km_vide_sum:,.0f} km")
        kp3.metric("🔄 KM Total complet", f"{_km_complet:,.0f} km")
        kp4.metric("% À Vide",             f"{pct_vide:.1f}%")
        kp5.metric("💶 CA Total",          f"{_ca_ptv:,.0f} €")
        kp6.metric("📈 Rentabilité",       f"{_rent_ptv:.2f} €/km")

        tab1, tab2, tab3, tab4 = st.tabs([
            "🇫🇷 Rentabilité par département",
            "📋 Détail dossiers",
            "👤 Résumé par chauffeur",
            "⚡ Détail KM à vide",
        ])

        # ── TAB 1 : rentabilité par département ──
        with tab1:
            df_dept = build_dept_stats(df_result_f)

            if df_dept.empty:
                st.info(
                    "Aucun déchargement en France identifié dans les résultats. "
                    "Vérifie la colonne « C.P. déchargement » du fichier CA ou bascule "
                    "la source sur « Dernier stop DÉCHARGEMENT »."
                )
            else:
                dt_c1, dt_c2, dt_c3 = st.columns([1, 1, 2])
                with dt_c1:
                    min_dossiers = st.number_input(
                        "Nb dossiers min", min_value=1, max_value=50, value=1, step=1,
                        help="Masque les départements avec trop peu de dossiers (bruit).",
                    )
                with dt_c2:
                    tri_par = st.selectbox(
                        "Trier par",
                        options=[
                            "Renta €/km complet",
                            "Renta €/km chargé",
                            "Total Vente (€)",
                            "Nb Dossiers",
                            "KM Complet",
                            "% KM Vide",
                            "Dépt",
                        ],
                        index=0,
                    )
                with dt_c3:
                    ordre_asc = st.radio(
                        "Ordre",
                        options=["Décroissant", "Croissant"],
                        horizontal=True,
                        index=0,
                    ) == "Croissant"

                df_dept_v = df_dept[df_dept["nb_dossiers"] >= min_dossiers].copy()

                if df_dept_v.empty:
                    st.warning("Aucun département ne passe le seuil de dossiers minimum.")
                else:
                    _tri_map = {v: k for k, v in DEPT_COL_RENAME.items()}
                    col_tri  = _tri_map.get(tri_par, "renta_complet")
                    df_dept_v = df_dept_v.sort_values(
                        col_tri, ascending=ordre_asc, na_position="last"
                    ).reset_index(drop=True)

                    # ── Top / Flop ──
                    _valides = df_dept_v[df_dept_v["renta_complet"].notna()]
                    if not _valides.empty:
                        best  = _valides.sort_values("renta_complet", ascending=False).iloc[0]
                        worst = _valides.sort_values("renta_complet", ascending=True).iloc[0]
                        tf1, tf2, tf3, tf4 = st.columns(4)
                        tf1.metric(
                            "🥇 Meilleur département",
                            f"{best['renta_complet']:.2f} €/km",
                            f"{best['dept_decharg']} — {best['nom_dept']}",
                        )
                        tf2.metric(
                            "🥉 Moins rentable",
                            f"{worst['renta_complet']:.2f} €/km",
                            f"{worst['dept_decharg']} — {worst['nom_dept']}",
                        )
                        tf3.metric("🗺️ Départements", len(df_dept_v))
                        tf4.metric(
                            "📁 Dossiers analysés",
                            int(df_dept_v["nb_dossiers"].sum()),
                        )

                    st.dataframe(
                        format_dept_table(df_dept_v, avec_km=True),
                        use_container_width=True,
                        height=520,
                        column_config={
                            "Renta €/km complet": st.column_config.NumberColumn(format="%.2f €"),
                            "Renta €/km chargé":  st.column_config.NumberColumn(format="%.2f €"),
                            "% KM Vide":          st.column_config.NumberColumn(format="%.1f %%"),
                            "Total Vente (€)":    st.column_config.NumberColumn(format="%.0f €"),
                            "Prix Transport (€)": st.column_config.NumberColumn(format="%.0f €"),
                            "CA moy/dossier (€)": st.column_config.NumberColumn(format="%.0f €"),
                        },
                    )

                    # ── Graphique ──
                    _chart = df_dept_v[df_dept_v["renta_complet"].notna()].copy()
                    if not _chart.empty:
                        _chart["Dépt"] = _chart["dept_decharg"] + " " + _chart["nom_dept"]
                        st.bar_chart(
                            _chart.set_index("Dépt")["renta_complet"],
                            height=320,
                        )
                        st.caption("Rentabilité €/km complet (chargé + à vide) par département de déchargement.")

                    st.download_button(
                        "📥 Export CSV départements",
                        data=format_dept_table(df_dept_v, avec_km=True)
                             .to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig"),
                        file_name="Renta_par_departement.csv",
                        mime="text/csv",
                    )

                    # ── Détail d'un département ──
                    st.markdown("##### 🔍 Détail d'un département")
                    _dep_choix = st.selectbox(
                        "Département",
                        options=df_dept_v["dept_decharg"].tolist(),
                        format_func=dept_label,
                    )
                    _det = df_result_f[df_result_f["dept_decharg"] == _dep_choix]
                    st.dataframe(
                        _det[[c for c in [
                            "dossier", "chauffeur", "date_debut", "client",
                            "localite_charg", "localite_decharg",
                            "km_total", "km_vide", "prix_transport", "total_vente",
                        ] if c in _det.columns]].rename(columns={
                            "dossier":          "N° Dossier",
                            "chauffeur":        "Chauffeur",
                            "date_debut":       "Date",
                            "client":           "Client",
                            "localite_charg":   "Ville charg.",
                            "localite_decharg": "Ville déch.",
                            "km_total":         "KM Chargés",
                            "km_vide":          "KM À Vide",
                            "prix_transport":   "Prix Transport €",
                            "total_vente":      "Total Vente €",
                        }),
                        use_container_width=True,
                        height=300,
                    )

        # ── TAB 2 : détail dossiers ──
        with tab2:
            cols_res = [
                "dossier", "chauffeur", "tracteur", "date_debut", "client",
                "localite_decharg", "dept_label",
                "stops_texte", "km_total", "km_vide", "prix_transport", "total_vente",
            ]
            st.dataframe(
                df_result_f[[c for c in cols_res if c in df_result_f.columns]].rename(columns={
                    "dossier":          "N° Dossier",
                    "chauffeur":        "Chauffeur",
                    "tracteur":         "Tracteur",
                    "date_debut":       "Date",
                    "client":           "Client",
                    "localite_decharg": "Ville déchargement",
                    "dept_label":       "Département déch.",
                    "stops_texte":      "Séquence",
                    "km_total":         "KM Total",
                    "km_vide":          "KM À Vide",
                    "prix_transport":   "Prix Transport €",
                    "total_vente":      "Total Vente €",
                }),
                use_container_width=True,
                height=400,
            )

        # ── TAB 3 : résumé chauffeurs ──
        with tab3:
            df_resume_ptv = df_result_f.groupby("chauffeur", as_index=False).agg(
                Dossiers       = ("dossier",        "count"),
                KM_Total       = ("km_total",        "sum"),
                KM_Vide        = ("km_vide",          "sum"),
                Prix_Transport = ("prix_transport",   "sum"),
                Total_Vente    = ("total_vente",      "sum"),
            ).round(1)
            df_resume_ptv["KM Complet"]       = df_resume_ptv["KM_Total"] + df_resume_ptv["KM_Vide"]
            df_resume_ptv["% KM Vide"]        = (
                df_resume_ptv["KM_Vide"] / df_resume_ptv["KM Complet"].replace(0, np.nan) * 100
            ).round(1)
            df_resume_ptv["Rentabilité €/km"] = (
                df_resume_ptv["Total_Vente"] / df_resume_ptv["KM Complet"].replace(0, np.nan)
            ).round(2)
            df_resume_ptv.columns = [
                "Chauffeur", "Nb Dossiers", "KM Chargés", "KM À Vide",
                "Prix Transport €", "Total Vente €",
                "KM Complet", "% KM Vide", "Rentabilité €/km",
            ]
            st.dataframe(df_resume_ptv, use_container_width=True)

        # ── TAB 4 : km à vide ──
        with tab4:
            if not df_vide.empty:
                st.dataframe(
                    df_vide.rename(columns={
                        "chauffeur":       "Chauffeur",
                        "dossier_depart":  "Dossier départ",
                        "dossier_arrivee": "Dossier arrivée",
                        "from_localite":   "Ville départ",
                        "to_localite":     "Ville arrivée",
                        "km_vide":         "KM à vide",
                    }),
                    use_container_width=True,
                )
            else:
                st.info("Aucun trajet à vide détecté.")

        st.divider()
        excel_bytes = export_excel(
            df_result_f, df_vide, df_dept=build_dept_stats(df_result_f)
        )
        st.download_button(
            label="📥 Télécharger le rapport Excel",
            data=excel_bytes,
            file_name="Rapport_Missions_CA_KM.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

elif file_missions and not file_ca:
    st.info("📂 Fichier missions chargé. En attente du fichier CA...")

elif file_ca and not file_missions:
    st.info("📂 Fichier CA chargé. En attente du fichier missions...")

else:
    st.markdown("""
    #### Comment utiliser cet outil

    1. **Chargez le fichier Missions** (export avec N°Dossier, Activité, stops, chauffeur)
    2. **Chargez le fichier CA** (export avec N°Dossier, Prix transport, Total vente)
    3. Choisissez la **source du département de déchargement** (fichier CA ou dernier stop
       DÉCHARGEMENT du fichier missions)
    4. Consultez le **tableau consolidé** par dossier, filtrable par département
    5. Sélectionnez les chauffeurs et **lancez le calcul PTV** pour obtenir :
       - Les km totaux par dossier (toute la chaîne de stops)
       - Les km à vide entre chaque déchargement et le rechargement suivant
       - La **rentabilité €/km par département de déchargement**
    6. **Téléchargez le rapport Excel** (onglet « Renta Départements » inclus)

    > ⚙️ La clé PTV doit être configurée dans le fichier `.env` (`PTV_API_KEY`).
    """)
