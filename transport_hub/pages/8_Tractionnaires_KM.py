"""
6____Tractionnaires_KM.py
──────────────────────────────────────────────────────────────────
Outil TX-FLEX : Analyse Tractionnaires — KM PTV + CA + Rentabilité
──────────────────────────────────────────────────────────────────
Entrée : Export tractionnaires (.xlsx)
  Colonnes : Tractionnaire, Chauffeur, Véhicule, Remorque, Dossier,
             Référence, Type transport, CMR, Date chargement,
             CP chargement, Localité chargement, Pays chargement,
             Date déchargement, CP déchargement, Localité déchargement,
             Pays déchargement, Statut facturation, Ventes totales,
             Département vente, Client

──────────────────────────────────────────────────────────────────
CORRECTIFS v2 (écarts constatés vs relevés compteur CHESCA)
──────────────────────────────────────────────────────────────────
BUG 1 — Double comptage du KM à vide
  Avant : km_vide était mappé par n° de dossier sur TOUTES les lignes
  du dossier. Un dossier à 3 livraisons se voyait attribuer 3× le même
  tronçon à vide, et le résumé (somme de la colonne) le comptait 3 fois.
  → Avril 2026 : 6 983 km fantômes sur 48 487.
  Fix : les km (chargés ET à vide) sont désormais posés sur UNE SEULE
  ligne par dossier (la première). Les autres lignes sont à 0.

BUG 2 — Tronçons à vide auto-référents
  Avant : le chaînage bouclait sur les LIGNES du camion. Deux lignes
  d'un même dossier généraient un tronçon "à vide" entre le déchargement
  du dossier et... le chargement de ce même dossier.
  → 28 tronçons en avril, dont GL91CHE ISTRES → FLEURUS = 971,5 km.
  C'était aussi la cause des 75 tronçons rejetés sur 265 (gaps négatifs).
  Fix : consolidation en 1 événement par (véhicule, dossier) AVANT le
  chaînage. Départ du tronçon = DERNIÈRE livraison du dossier courant,
  arrivée = PREMIER chargement du dossier suivant.

BUG 3 — Tronçons à vide aberrants (trous de chaîne)
  Un tronçon à vide de 1 093 km (MONGRANDO → MONTAUBAN-DE-BRETAGNE)
  n'est pas du vide : c'est un dossier manquant dans l'export (filtré
  au bord du mois, autre département vente, trajet non facturé).
  Fix : seuil KM_VIDE_MAX. Au-delà, le tronçon est marqué
  "Trou de chaîne" et EXCLU du total, mais listé pour contrôle.

BUG 4 — Dossiers affectés à plusieurs véhicules (relais tracteur)
  Le même dossier apparaît sur 2 camions → km et CA comptés 2 fois.
  → Mai 2026 : 7 dossiers, 4 819 km en double.
  Fix : détection + répartition au prorata (option).

BUG 5 — Trajets inter-livraisons non comptés
  Avant : un dossier A → B1 → B2 ne comptait que A → B1.
  Fix : le km chargé est calculé sur l'itinéraire complet
  (tous les chargements puis toutes les livraisons, dans l'ordre).

BUG 6 — build_resume comptait les LIGNES, pas les dossiers uniques.
──────────────────────────────────────────────────────────────────
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import time
import os
import io
import re
from typing import Optional, Tuple, List
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ══════════════════════════════════════════════════════════════════
#  CONFIG PTV
# ══════════════════════════════════════════════════════════════════

PTV_API_KEY  = os.environ.get("PTV_API_KEY", "METS_TA_CLE_ICI")
PTV_BASE_URL = "https://api.myptv.com/routing/v1"
GEOCODE_URL  = "https://api.myptv.com/geocoding/v1"
HEADERS      = {"apiKey": PTV_API_KEY}
MAX_RETRIES  = 3
RETRY_DELAY  = 2
VEHICLE      = "EUR_TRAILER_TRUCK"

# Seuil par défaut au-delà duquel un tronçon à vide est jugé fictif
GAP_VIDE_DEFAUT   = 3    # jours
KM_VIDE_MAX_DEFAUT = 300  # km — au-delà : trou de chaîne, pas du vide

PAYS_MAP = {
    "F": "France", "B": "Belgium", "D": "Germany", "L": "Luxembourg",
    "NL": "Netherlands", "E": "Spain", "I": "Italy", "CH": "Switzerland",
    "GB": "United Kingdom", "A": "Austria", "P": "Portugal",
    "FR": "France", "BE": "Belgium", "DE": "Germany", "LU": "Luxembourg",
    "IT": "Italy", "ES": "Spain", "AT": "Austria", "PT": "Portugal",
}

PAYS_TO_ISO2 = {
    "F": "FR", "B": "BE", "D": "DE", "L": "LU", "I": "IT",
    "E": "ES", "A": "AT", "P": "PT", "CH": "CH", "GB": "GB",
    "NL": "NL", "FR": "FR", "BE": "BE", "DE": "DE", "LU": "LU",
    "IT": "IT", "ES": "ES", "AT": "AT", "PT": "PT",
}

MOIS_FR = ["", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

P_COMPLET = "Complet"
P_ENTRANT = "Entrant (chargé M-1)"
P_SORTANT = "Sortant (déchargé M+1)"
P_HORS    = "Hors mois"

# Version du schéma des résultats stockés en session.
# À incrémenter dès qu'on change les colonnes produites par compute_km :
# un résultat calculé par une ancienne version est alors invalidé au lieu
# de faire planter l'affichage (KeyError sur une colonne absente).
RESULT_SCHEMA = "v2"

# Statuts de tronçon à vide
S_OK          = "OK"
S_MEME_LIEU   = "Même lieu"
S_GEO         = "Géocodage manquant"
S_DATES       = "Dates incohérentes — ignoré"
S_ROUTE       = "Route PTV échouée"


# ══════════════════════════════════════════════════════════════════
#  UTILITAIRES
# ══════════════════════════════════════════════════════════════════

def _norm_col(s: str) -> str:
    s = str(s).strip().lower()
    for src, dst in [("é", "e"), ("è", "e"), ("ê", "e"), ("à", "a"), ("â", "a"),
                     ("ô", "o"), ("û", "u"), ("î", "i"), ("ù", "u"), ("ç", "c")]:
        s = s.replace(src, dst)
    return re.sub(r"[^a-z0-9]", "", s)


def _clean(v) -> str:
    v = str(v or "").strip()
    return "" if v.lower() in ("nan", "none", "nat") else v


def _to_float(s) -> float:
    try:
        return float(str(s).replace(",", ".").replace("\u00a0", "")
                     .replace(" ", "").replace("€", "").strip())
    except Exception:
        return 0.0


def _opts(series: pd.Series) -> List[str]:
    return sorted({v for v in series.dropna().astype(str)
                   if v.strip() and v.strip().lower() != "nan"})


def _libelle_mois(per) -> str:
    if per is None:
        return "Période inconnue"
    return f"{MOIS_FR[per.month]} {per.year}"


def _pt(row, sens: str) -> Tuple[str, str, str]:
    """Clé de géocodage d'un point (sens = 'charg' ou 'decharg')."""
    return (_clean(row[f"localite_{sens}"]),
            _clean(row[f"cp_{sens}"]),
            _clean(row[f"pays_{sens}"]))


def _pt_vide(p: Tuple[str, str, str]) -> bool:
    return not (p[0] or p[1])


def _ca_col(df: pd.DataFrame) -> str:
    """Colonne de CA à utiliser : 'ventes_retenues' si le prorata a été appliqué,
    sinon le CA brut. Évite toute dépendance dure à une colonne optionnelle."""
    return "ventes_retenues" if "ventes_retenues" in df.columns else "ventes_totales"


def _normaliser_resultat(df: pd.DataFrame) -> pd.DataFrame:
    """Filet de sécurité : garantit la présence des colonnes attendues par
    l'affichage, même sur un DataFrame issu d'une version antérieure."""
    df = df.copy()
    if "ventes_retenues" not in df.columns:
        df["ventes_retenues"] = df.get("ventes_totales", 0.0)
    if "_ligne_km" not in df.columns:
        df["_ligne_km"] = True
    if "_multi_veh" not in df.columns:
        df["_multi_veh"] = False
    for c in ("km_ptv", "km_vide"):
        if c not in df.columns:
            df[c] = 0.0
    if "km_total_complet" not in df.columns:
        df["km_total_complet"] = df["km_ptv"].fillna(0) + df["km_vide"].fillna(0)
    return df


# ══════════════════════════════════════════════════════════════════
#  GEOCODAGE
# ══════════════════════════════════════════════════════════════════

def _ptv_by_text(query: str) -> Optional[Tuple[float, float]]:
    if not query:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-text",
                params={"searchText": query},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def _ptv_by_postal_code(cp: str, iso2: str) -> Optional[Tuple[float, float]]:
    if not cp or not iso2:
        return None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{GEOCODE_URL}/locations/by-postal-code",
                params={"postalCode": cp, "countryCode": iso2},
                headers=HEADERS, timeout=15,
            )
            if resp.status_code == 429:
                time.sleep(RETRY_DELAY * attempt)
                continue
            if resp.status_code != 200:
                return None
            locs = resp.json().get("locations", [])
            if locs:
                pos = locs[0]["referencePosition"]
                return (pos["latitude"], pos["longitude"])
            return None
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def geocode_with_fallback(ville: str, cp: str, pays: str) -> Optional[Tuple[float, float]]:
    pays_full = PAYS_MAP.get(pays.upper(), pays) if pays else ""
    iso2      = PAYS_TO_ISO2.get(pays.upper(), pays.upper() if len(pays) == 2 else "")

    if ville and cp and pays_full:
        r = _ptv_by_text(f"{ville}, {cp}, {pays_full}")
        if r:
            return r
    if cp and iso2:
        r = _ptv_by_postal_code(cp, iso2)
        if r:
            return r
    if ville and pays_full:
        r = _ptv_by_text(f"{ville}, {pays_full}")
        if r:
            return r
    return None


# ══════════════════════════════════════════════════════════════════
#  CALCUL ROUTE PTV
# ══════════════════════════════════════════════════════════════════

def calculate_route(coords_list: list) -> Optional[dict]:
    if len(coords_list) < 2:
        return None
    query_params = [("profile", VEHICLE), ("results", "POLYLINE")]
    for lat, lon in coords_list:
        query_params.append(("waypoints", f"{lat},{lon}"))
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(
                f"{PTV_BASE_URL}/routes",
                headers=HEADERS, params=query_params, timeout=30,
            )
            if resp.status_code != 200:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return None
            data = resp.json()
            return {"km": round(data.get("distance", 0) / 1000, 1)}
        except Exception:
            time.sleep(RETRY_DELAY)
    return None


def _route_km_cached(cache: dict, a, b) -> Optional[float]:
    """Route A→B avec cache : une paire identique n'est demandée qu'une fois."""
    if a is None or b is None:
        return None
    key = (round(a[0], 4), round(a[1], 4), round(b[0], 4), round(b[1], 4))
    if key in cache:
        return cache[key]
    if key[0] == key[2] and key[1] == key[3]:
        cache[key] = 0.0
        return 0.0
    res = calculate_route([a, b])
    km = res["km"] if res else None
    cache[key] = km
    return km


# ══════════════════════════════════════════════════════════════════
#  PARSING
# ══════════════════════════════════════════════════════════════════

TRACT_COL_CANDIDATES = {
    "tractionnaire":   ["Tractionnaire"],
    "chauffeur":       ["Chauffeur"],
    "vehicule":        ["Véhicule", "Vehicule"],
    "remorque":        ["Remorque"],
    "dossier":         ["Dossier", "N° Dossier"],
    "reference":       ["Référence", "Reference"],
    "type_transport":  ["Type de transport", "Type transport"],
    "cmr":             ["CMR"],
    "date_charg":      ["Date chargement", "Date Chargement"],
    "cp_charg":        ["C.P. chargement", "CP chargement"],
    "localite_charg":  ["Localité chargement", "Localite chargement"],
    "pays_charg":      ["Pays chargement"],
    "date_decharg":    ["Date déchargement", "Date dechargement"],
    "cp_decharg":      ["C.P. déchargement", "CP dechargement"],
    "localite_decharg": ["Localité déchargement", "Localite dechargement"],
    "pays_decharg":    ["Pays déchargement", "Pays dechargement"],
    "statut":          ["Statut facturation", "Statut"],
    "ventes_totales":  ["Ventes totales", "Total ventes", "CA"],
    "dept_vente":      ["Département vente", "Dept vente"],
    "client":          ["Client"],
}


def parse_tractionnaires(file) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    cols_lower = {_norm_col(c): c for c in df.columns}
    col_map = {}
    for role, candidates in TRACT_COL_CANDIDATES.items():
        found = None
        for cand in candidates:
            key = _norm_col(cand)
            if key in cols_lower:
                found = cols_lower[key]
                break
        col_map[role] = found

    critiques = ["tractionnaire", "dossier", "localite_charg", "localite_decharg"]
    manquantes = [r for r in critiques if col_map.get(r) is None]
    if manquantes:
        st.warning(f"⚠️ Colonnes non détectées : {manquantes} — Colonnes dispo : {list(df.columns)}")

    rename = {v: k for k, v in col_map.items() if v}
    df = df.rename(columns=rename)
    for col in TRACT_COL_CANDIDATES:
        if col not in df.columns:
            df[col] = ""

    df["dossier"] = df["dossier"].astype(str).str.strip()
    df = df[df["dossier"].notna() & (df["dossier"] != "") & (df["dossier"] != "nan")]
    df = df[df["dossier"].str.match(r"^\d+", na=False)]

    df["ventes_totales"] = df["ventes_totales"].apply(_to_float)

    for col in ["tractionnaire", "chauffeur", "vehicule", "remorque",
                "localite_charg", "cp_charg", "pays_charg",
                "localite_decharg", "cp_decharg", "pays_decharg",
                "statut", "client"]:
        df[col] = df[col].apply(_clean)

    df["_date_charg_dt"] = pd.to_datetime(
        df["date_charg"].astype(str).str.strip().str[:10], format="%Y-%m-%d", errors="coerce"
    )
    df["_date_decharg_dt"] = pd.to_datetime(
        df["date_decharg"].astype(str).str.strip().str[:10], format="%Y-%m-%d", errors="coerce"
    )
    df["_date_dt"] = df["_date_charg_dt"]

    df["date_charg_fmt"] = df["_date_charg_dt"].dt.strftime("%d/%m/%Y")
    df["date_charg_fmt"] = df["date_charg_fmt"].fillna(df["date_charg"].apply(_clean))
    df["date_decharg_fmt"] = df["_date_decharg_dt"].dt.strftime("%d/%m/%Y")
    df["date_decharg_fmt"] = df["date_decharg_fmt"].fillna(df["date_decharg"].apply(_clean))

    # ── CORRECTIF BUG 1/2 : ordre d'origine des lignes, indispensable
    #    pour reconstituer l'itinéraire d'un dossier multi-lignes.
    df = df.reset_index(drop=True)
    df["_ligne"] = np.arange(len(df))

    return df


# ══════════════════════════════════════════════════════════════════
#  PERIODE
# ══════════════════════════════════════════════════════════════════

def detect_mois_analyse(df: pd.DataFrame):
    d = df["_date_decharg_dt"].dropna()
    if d.empty:
        d = df["_date_charg_dt"].dropna()
    if d.empty:
        return None
    modes = d.dt.to_period("M").mode()
    if modes.empty:
        return None
    return modes.iloc[0]


def classer_dossiers(df: pd.DataFrame, mois) -> pd.DataFrame:
    """
    Classement au niveau DOSSIER (et non ligne) : toutes les lignes d'un
    dossier reçoivent la même période, sinon un dossier multi-livraisons
    peut être coupé en deux (une partie gardée, une partie exclue).
    """
    df = df.copy()
    if mois is None:
        df["_periode"] = P_COMPLET
        return df

    m_prev, m_next = mois - 1, mois + 1

    def _cls(pc, pdch):
        if pd.isna(pc) and pd.isna(pdch):
            return P_HORS
        if pd.isna(pc):
            return P_COMPLET if pdch == mois else P_HORS
        if pd.isna(pdch):
            return P_COMPLET if pc == mois else P_HORS
        if pc == mois and pdch == mois:
            return P_COMPLET
        if pc == m_prev and pdch == mois:
            return P_ENTRANT
        if pc == mois and pdch == m_next:
            return P_SORTANT
        if pc == mois or pdch == mois:
            return P_ENTRANT if pdch == mois else P_SORTANT
        return P_HORS

    # Dates au niveau dossier : 1er chargement / dernier déchargement
    key = ["vehicule", "dossier"]
    agg = df.groupby(key).agg(
        _dc=("_date_charg_dt", "min"),
        _dd=("_date_decharg_dt", "max"),
    ).reset_index()
    agg["_periode"] = [
        _cls(pc.to_period("M") if pd.notna(pc) else pd.NaT,
             pd.NaT if pd.isna(pdch) else pdch.to_period("M"))
        for pc, pdch in zip(agg["_dc"], agg["_dd"])
    ]
    df = df.merge(agg[key + ["_periode"]], on=key, how="left")
    df["_periode"] = df["_periode"].fillna(P_HORS)
    return df


# ══════════════════════════════════════════════════════════════════
#  CONSOLIDATION PAR DOSSIER  (CORRECTIF BUG 1, 2 et 5)
# ══════════════════════════════════════════════════════════════════

def consolider_dossiers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Une ligne d'export ≠ un dossier.
    Un dossier peut avoir plusieurs chargements et plusieurs livraisons,
    éclatés sur plusieurs lignes (la localité de chargement n'est souvent
    reprise que sur la 1re ligne).

    Reconstitue, pour chaque (véhicule, dossier) :
      • pickups : points de chargement, dans l'ordre, dédoublonnés
      • drops   : points de livraison, dans l'ordre, dédoublonnés
      • ligne_km : la ligne du détail qui portera les km (la première)
    """
    rows = []
    for (veh, dos), grp in df.groupby(["vehicule", "dossier"], sort=False):
        grp = grp.sort_values("_ligne")
        pickups, drops = [], []
        for _, r in grp.iterrows():
            p = _pt(r, "charg")
            d = _pt(r, "decharg")
            if not _pt_vide(p) and p not in pickups:
                pickups.append(p)
            if not _pt_vide(d) and d not in drops:
                drops.append(d)
        rows.append({
            "vehicule":      veh,
            "dossier":       dos,
            "tractionnaire": _clean(grp["tractionnaire"].iloc[0]),
            "pickups":       pickups,
            "drops":         drops,
            "ligne_km":      int(grp["_ligne"].iloc[0]),
            "nb_lignes":     len(grp),
            "date_charg":    grp["_date_charg_dt"].min(),
            "date_decharg":  grp["_date_decharg_dt"].max(),
            "ville_debut":   pickups[0][0] if pickups else "",
            "ville_fin":     drops[-1][0] if drops else (pickups[-1][0] if pickups else ""),
            "ca":            grp["ventes_totales"].sum(),
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════
#  CALCUL KM PTV
# ══════════════════════════════════════════════════════════════════

def compute_km(df: pd.DataFrame,
               progress_cb=None,
               calc_vide: bool = True,
               gap_max_jours: int = GAP_VIDE_DEFAUT,
               km_vide_max: int = KM_VIDE_MAX_DEFAUT,
               prorata_multi_veh: bool = True):
    """
    Retourne (df_enrichi, df_vide, df_dos, stats).

    Règles :
      • KM chargé = itinéraire COMPLET du dossier
        (chargement(s) → livraison(s), tous les points, dans l'ordre).
      • KM à vide  = dernière livraison du dossier N → premier chargement
        du dossier N+1 du MÊME camion, dossiers triés chronologiquement.
      • Les km sont posés sur UNE SEULE ligne par dossier → aucun risque
        de double comptage dans les résumés.
      • Un tronçon à vide > km_vide_max n'est pas du vide : c'est un trou
        dans la chaîne (dossier absent de l'export). Il est exclu du total
        et listé pour contrôle.
    """
    df = df.copy()
    stats = {"appels_route": 0, "routes_cache": 0,
             "vide_ignores": 0, "vide_meme_lieu": 0, "vide_trous": 0,
             "km_trous": 0.0, "dossiers_multi_veh": 0, "km_inter_livraisons": 0.0}

    # ── 1. Consolidation par dossier ─────────────────────────
    df_dos = consolider_dossiers(df)

    # ── 2. Géocodage de tous les points uniques ──────────────
    points = set()
    for _, d in df_dos.iterrows():
        points.update(d["pickups"])
        points.update(d["drops"])

    geo_cache = {}
    total = len(points)
    echecs_geo = []
    for i, p in enumerate(sorted(points)):
        if progress_cb:
            progress_cb(f"🌍 Géocodage {i+1}/{total} : {p[0]} {p[1]}...",
                        (i + 1) / max(total, 1) * 0.3)
        coords = geocode_with_fallback(*p)
        geo_cache[p] = coords
        if coords is None:
            echecs_geo.append(p)
    stats["echecs_geo"] = echecs_geo

    route_cache = {}

    def _route(a, b):
        before = len(route_cache)
        km = _route_km_cached(route_cache, a, b)
        if len(route_cache) > before:
            stats["appels_route"] += 1
        else:
            stats["routes_cache"] += 1
        return km

    def _route_itineraire(pts) -> Optional[float]:
        """Somme des tronçons consécutifs (chaque tronçon est mis en cache)."""
        coords = [geo_cache.get(p) for p in pts]
        coords = [c for c in coords if c]
        if len(coords) < 2:
            return None
        total_km, ok = 0.0, False
        for a, b in zip(coords[:-1], coords[1:]):
            km = _route(a, b)
            if km is None:
                return None
            total_km += km
            ok = True
        return round(total_km, 1) if ok else None

    # ── 3. KM chargés : itinéraire complet du dossier ─────────
    km_charges, km_inter = [], []
    n = len(df_dos)
    for i, (_, d) in enumerate(df_dos.iterrows()):
        if progress_cb:
            progress_cb(f"📍 KM chargé dossier {d['dossier']} ({i+1}/{n})...",
                        0.3 + (i + 1) / max(n, 1) * (0.4 if calc_vide else 0.7))
        itineraire = list(d["pickups"]) + list(d["drops"])
        km_tot = _route_itineraire(itineraire)
        km_charges.append(km_tot)

        # Part des trajets inter-livraisons (ignorés dans l'ancienne version)
        km_direct = None
        if d["pickups"] and d["drops"]:
            km_direct = _route(geo_cache.get(d["pickups"][0]), geo_cache.get(d["drops"][0]))
        if km_tot is not None and km_direct is not None:
            km_inter.append(round(km_tot - km_direct, 1))
        else:
            km_inter.append(0.0)

    df_dos["km_charge"] = km_charges
    df_dos["km_inter_livraisons"] = km_inter
    stats["km_inter_livraisons"] = round(float(np.nansum(df_dos["km_inter_livraisons"])), 1)

    # ── 4. KM à vide : chaînage AU NIVEAU DOSSIER ────────────
    vide_legs = []
    if calc_vide:
        vehicules = [v for v in df_dos["vehicule"].unique() if v and v != "nan"]
        for v_idx, vehicule in enumerate(vehicules):
            g = (df_dos[df_dos["vehicule"] == vehicule]
                 .sort_values(["date_charg", "date_decharg", "ligne_km"],
                              na_position="last")
                 .reset_index(drop=True))

            for i in range(len(g) - 1):
                cur, nxt = g.iloc[i], g.iloc[i + 1]

                # Départ = DERNIÈRE livraison du dossier courant
                p_dep = cur["drops"][-1] if cur["drops"] else (
                    cur["pickups"][-1] if cur["pickups"] else None)
                # Arrivée = PREMIER chargement du dossier suivant
                p_arr = nxt["pickups"][0] if nxt["pickups"] else (
                    nxt["drops"][0] if nxt["drops"] else None)

                c_dep = geo_cache.get(p_dep) if p_dep else None
                c_arr = geo_cache.get(p_arr) if p_arr else None

                d_fin, d_debut = cur["date_decharg"], nxt["date_charg"]
                gap = (d_debut - d_fin).days if (pd.notna(d_fin) and pd.notna(d_debut)) else None

                km_vide, km_brut, statut = None, None, S_OK

                if c_dep is None or c_arr is None:
                    statut = S_GEO
                elif gap is not None and gap < 0:
                    # Chevauchement réel entre 2 dossiers du même camion :
                    # on ne sait pas reconstituer l'ordre → tronçon ignoré.
                    statut = S_DATES
                    stats["vide_ignores"] += 1
                elif gap is not None and gap > gap_max_jours:
                    statut = f"Gap {gap} j > {gap_max_jours} j — ignoré"
                    stats["vide_ignores"] += 1
                else:
                    if progress_cb:
                        progress_cb(
                            f"⚡ KM vide {vehicule} : {p_dep[0]} → {p_arr[0]}...",
                            0.7 + (v_idx + 1) / max(len(vehicules), 1) * 0.3,
                        )
                    km_brut = _route(c_dep, c_arr)
                    if km_brut is None:
                        statut = S_ROUTE
                    elif km_brut == 0.0:
                        statut = S_MEME_LIEU
                        km_vide = 0.0
                        stats["vide_meme_lieu"] += 1
                    elif km_brut > km_vide_max:
                        # ── CORRECTIF BUG 3 : ce n'est pas du vide, c'est un
                        #    dossier manquant dans l'export. Listé, mais exclu.
                        statut = f"Trou de chaîne ({km_brut:,.0f} km > {km_vide_max}) — exclu"
                        stats["vide_trous"] += 1
                        stats["km_trous"] += km_brut
                        km_vide = None
                    else:
                        km_vide = km_brut

                vide_legs.append({
                    "vehicule":        vehicule,
                    "tractionnaire":   cur["tractionnaire"],
                    "dossier_depart":  cur["dossier"],
                    "dossier_arrivee": nxt["dossier"],
                    "ville_depart":    p_dep[0] if p_dep else "",
                    "ville_arrivee":   p_arr[0] if p_arr else "",
                    "date_depart":     d_fin.strftime("%d/%m/%Y") if pd.notna(d_fin) else "",
                    "gap_jours":       gap if gap is not None else "",
                    "statut_leg":      statut,
                    "km_vide":         km_vide,
                    "km_brut":         km_brut,
                })

    df_vide = pd.DataFrame(vide_legs)

    # ── 5. Rattachement du vide au dossier de DÉPART ─────────
    if not df_vide.empty:
        v = (df_vide.groupby(["vehicule", "dossier_depart"])["km_vide"]
             .sum(min_count=1).rename("km_vide").reset_index()
             .rename(columns={"dossier_depart": "dossier"}))
        df_dos = df_dos.merge(v, on=["vehicule", "dossier"], how="left")
    else:
        df_dos["km_vide"] = np.nan
    df_dos["km_vide"] = df_dos["km_vide"].fillna(0.0)

    # ── 6. Dossiers sur plusieurs véhicules (CORRECTIF BUG 4) ─
    df_dos["_nb_veh"] = df_dos.groupby("dossier")["vehicule"].transform("nunique")
    df_dos["multi_vehicule"] = df_dos["_nb_veh"] > 1
    stats["dossiers_multi_veh"] = int(
        df_dos.loc[df_dos["multi_vehicule"], "dossier"].nunique()
    )
    if prorata_multi_veh:
        df_dos["km_charge"] = df_dos["km_charge"] / df_dos["_nb_veh"]
        df_dos["coef_ca"]   = 1.0 / df_dos["_nb_veh"]
    else:
        df_dos["coef_ca"] = 1.0

    # ── 7. Report sur le détail : km sur UNE ligne par dossier ─
    #    (CORRECTIF BUG 1 : plus de valeur répétée sur les lignes filles)
    df["km_ptv"]          = 0.0
    df["km_vide"]         = 0.0
    df["_ligne_km"]       = False
    df["_multi_veh"]      = False
    df["_coef_ca"]        = 1.0

    idx = df.set_index("_ligne")
    for _, d in df_dos.iterrows():
        L = d["ligne_km"]
        idx.at[L, "km_ptv"]     = d["km_charge"] if pd.notna(d["km_charge"]) else np.nan
        idx.at[L, "km_vide"]    = d["km_vide"]
        idx.at[L, "_ligne_km"]  = True
    # coef CA et flag multi-véhicule : sur toutes les lignes du dossier
    coef = df_dos.set_index(["vehicule", "dossier"])["coef_ca"]
    mult = df_dos.set_index(["vehicule", "dossier"])["multi_vehicule"]
    keys = list(zip(idx["vehicule"], idx["dossier"]))
    idx["_coef_ca"]   = [coef.get(k, 1.0) for k in keys]
    idx["_multi_veh"] = [bool(mult.get(k, False)) for k in keys]

    df = idx.reset_index()
    df["ventes_retenues"] = df["ventes_totales"] * df["_coef_ca"]
    df["km_total_complet"] = df["km_ptv"].fillna(0) + df["km_vide"].fillna(0)

    return df, df_vide, df_dos, stats


# ══════════════════════════════════════════════════════════════════
#  RESUMES
# ══════════════════════════════════════════════════════════════════

COLS_APERCU = {
    "dossier":          "N° Dossier",
    "_periode":         "Période",
    "tractionnaire":    "Tractionnaire",
    "vehicule":         "Véhicule",
    "date_charg_fmt":   "Date chargement",
    "localite_charg":   "Chargement",
    "date_decharg_fmt": "Date déchargement",
    "localite_decharg": "Déchargement",
    "client":           "Client",
    "statut":           "Statut",
    "ventes_totales":   "CA (€)",
}


def tableau_apercu(dfx: pd.DataFrame) -> pd.DataFrame:
    if dfx.empty:
        return pd.DataFrame(columns=list(COLS_APERCU.values()))
    d = dfx.sort_values(["_date_charg_dt", "dossier"], na_position="last")
    return d[[c for c in COLS_APERCU if c in d.columns]].rename(columns=COLS_APERCU)


def build_resume(df: pd.DataFrame, cle: str, label: str) -> pd.DataFrame:
    """
    CORRECTIF BUG 6 : 'Dossiers' compte les dossiers UNIQUES, pas les lignes.
    Les km sont portés par une seule ligne par dossier → la somme est juste.
    """
    ca_col = _ca_col(df)
    res = df.groupby(cle, as_index=False).agg(
        Dossiers   = ("dossier",  "nunique"),
        Lignes     = ("dossier",  "count"),
        KM_Charges = ("km_ptv",   "sum"),
        KM_Vide    = ("km_vide",  "sum"),
        CA_Total   = (ca_col,     "sum"),
    ).round(1)
    res["KM Complet"]       = (res["KM_Charges"] + res["KM_Vide"]).round(1)
    res["% KM Vide"]        = (res["KM_Vide"] / res["KM Complet"].replace(0, np.nan) * 100).round(1)
    res["CA moy/dossier"]   = (res["CA_Total"] / res["Dossiers"].replace(0, np.nan)).round(0)
    res["Rentabilité €/km"] = (res["CA_Total"] / res["KM Complet"].replace(0, np.nan)).round(2)
    res = res.rename(columns={
        cle: label,
        "KM_Charges": "KM Chargés",
        "KM_Vide": "KM À Vide",
        "CA_Total": "CA Total (€)",
    }).sort_values("CA Total (€)", ascending=False)
    return res


# ══════════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ══════════════════════════════════════════════════════════════════

def export_excel(df_detail, df_resume_tract, df_resume_veh=None,
                 df_vide=None, df_dos=None) -> bytes:
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            col_rename_detail = {
                "tractionnaire":    "Tractionnaire",
                "chauffeur":        "Chauffeur",
                "vehicule":         "Véhicule",
                "remorque":         "Remorque",
                "dossier":          "N° Dossier",
                "client":           "Client",
                "statut":           "Statut facturation",
                "_periode":         "Période",
                "date_charg_fmt":   "Date chargement",
                "localite_charg":   "Localité chargement",
                "date_decharg_fmt": "Date déchargement",
                "localite_decharg": "Localité déchargement",
                "ventes_totales":   "Ventes totales (€)",
                "ventes_retenues":  "CA retenu (€)",
                "km_ptv":           "KM Chargé",
                "km_vide":          "KM À Vide",
                "km_total_complet": "KM Complet",
                "rentabilite":      "Rentabilité €/km",
                "_ligne_km":        "Ligne porteuse des KM",
                "_multi_veh":       "Dossier multi-véhicules",
            }
            df_d = _normaliser_resultat(df_detail)
            df_d["rentabilite"] = (
                df_d[_ca_col(df_d)] / df_d["km_total_complet"].replace(0, np.nan)
            ).round(2)
            cols = [c for c in col_rename_detail if c in df_d.columns]
            df_d = df_d[cols].rename(columns=col_rename_detail).fillna("")
            df_d.to_excel(writer, sheet_name="Détail dossiers", index=False)
            _style_sheet(writer.sheets["Détail dossiers"], len(df_d))

            df_resume_tract.to_excel(writer, sheet_name="Résumé tractionnaires", index=False)
            _style_sheet(writer.sheets["Résumé tractionnaires"], len(df_resume_tract))

            if df_resume_veh is not None and not df_resume_veh.empty:
                df_resume_veh.to_excel(writer, sheet_name="Résumé véhicules", index=False)
                _style_sheet(writer.sheets["Résumé véhicules"], len(df_resume_veh))

            # Une ligne par dossier : la vraie maille de calcul
            if df_dos is not None and not df_dos.empty:
                dd = df_dos.copy()
                dd["Chargements"] = dd["pickups"].apply(lambda L: " → ".join(p[0] for p in L))
                dd["Livraisons"]  = dd["drops"].apply(lambda L: " → ".join(p[0] for p in L))
                dd = dd[["vehicule", "dossier", "Chargements", "Livraisons", "nb_lignes",
                         "km_charge", "km_inter_livraisons", "km_vide",
                         "multi_vehicule", "ca"]].rename(columns={
                    "vehicule": "Véhicule", "dossier": "N° Dossier",
                    "nb_lignes": "Nb lignes export", "km_charge": "KM Chargé",
                    "km_inter_livraisons": "dont inter-livraisons",
                    "km_vide": "KM À Vide (après ce dossier)",
                    "multi_vehicule": "Multi-véhicules", "ca": "CA (€)",
                })
                dd.to_excel(writer, sheet_name="Dossiers consolidés", index=False)
                _style_sheet(writer.sheets["Dossiers consolidés"], len(dd))

            if df_vide is not None and not df_vide.empty:
                vide_rename = {
                    "vehicule": "Véhicule", "tractionnaire": "Tractionnaire",
                    "dossier_depart": "Dossier départ", "dossier_arrivee": "Dossier arrivée",
                    "ville_depart": "Ville départ", "ville_arrivee": "Ville arrivée",
                    "date_depart": "Date", "gap_jours": "Gap (j)",
                    "statut_leg": "Statut tronçon", "km_vide": "KM à vide retenu",
                    "km_brut": "KM brut PTV",
                }
                df_v = df_vide[[c for c in vide_rename if c in df_vide.columns]].rename(
                    columns=vide_rename).fillna("")
                df_v.to_excel(writer, sheet_name="KM À Vide Détail", index=False)
                _style_sheet(writer.sheets["KM À Vide Détail"], len(df_v))

    except Exception as e:
        st.error(f"❌ Erreur génération Excel : {e}")
        return b""

    return output.getvalue()


def _style_sheet(ws, nb_rows: int):
    HEADER_FILL = PatternFill("solid", fgColor="003087")   # navy CB Groupe
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row_idx in range(2, nb_rows + 2):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


# ══════════════════════════════════════════════════════════════════
#  INTERFACE STREAMLIT
# ══════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Tractionnaires KM + CA", page_icon="🚛", layout="wide")

st.title("🚛 Analyse Tractionnaires — KM PTV + CA")
st.caption("Calcule les km estimés PTV et analyse le CA par tractionnaire / véhicule.")

if not PTV_API_KEY or PTV_API_KEY == "METS_TA_CLE_ICI":
    st.error("⚠️ Clé PTV_API_KEY non configurée.")

st.divider()

file_tract = st.file_uploader("📋 Export tractionnaires (.xlsx)", type=["xlsx"])

if file_tract:
    with st.spinner("📂 Lecture du fichier..."):
        try:
            df_raw = parse_tractionnaires(file_tract)
        except Exception as e:
            st.error(f"❌ Erreur lecture : {e}")
            st.stop()

    if df_raw.empty:
        st.error("❌ Aucun dossier valide détecté dans le fichier.")
        st.stop()

    # ══════════════════════════════════════════════════════════
    #  PERIODE D'ANALYSE
    # ══════════════════════════════════════════════════════════
    mois_auto = detect_mois_analyse(df_raw)

    st.markdown("### 🗓️ Période d'analyse")

    mois_dispo = sorted({
        p for p in pd.concat([
            df_raw["_date_decharg_dt"], df_raw["_date_charg_dt"]
        ]).dropna().dt.to_period("M").unique()
    })
    labels_mois = {_libelle_mois(p): p for p in mois_dispo}
    default_label = _libelle_mois(mois_auto) if mois_auto is not None else None
    liste_labels = list(labels_mois.keys())
    idx_default = liste_labels.index(default_label) if default_label in liste_labels else 0

    pc1, pc2 = st.columns([1, 2])
    with pc1:
        mois_label = st.selectbox(
            "Mois analysé (déduit des déchargements) :",
            options=liste_labels, index=idx_default, key="mois_sel",
        )
        mois = labels_mois.get(mois_label)

    df_all = classer_dossiers(df_raw, mois)

    cnt = df_all.drop_duplicates(["vehicule", "dossier"])["_periode"].value_counts().to_dict()
    n_complet = cnt.get(P_COMPLET, 0)
    n_entrant = cnt.get(P_ENTRANT, 0)
    n_sortant = cnt.get(P_SORTANT, 0)
    n_hors    = cnt.get(P_HORS, 0)

    with pc2:
        regle = st.radio(
            "Règle d'attribution des dossiers à cheval :",
            options=[
                "Tout inclure (comme l'export)",
                "Mois strict (chargement ET déchargement dans le mois)",
                "Personnalisé",
            ],
            index=0, horizontal=False, key="regle_periode",
        )

    if regle == "Tout inclure (comme l'export)":
        cats_gardees = [P_COMPLET, P_ENTRANT, P_SORTANT]
    elif regle == "Mois strict (chargement ET déchargement dans le mois)":
        cats_gardees = [P_COMPLET]
    else:
        cats_gardees = st.multiselect(
            "Catégories à conserver :",
            options=[P_COMPLET, P_ENTRANT, P_SORTANT, P_HORS],
            default=[P_COMPLET, P_ENTRANT, P_SORTANT],
            key="cats_periode",
        )

    p1, p2, p3, p4 = st.columns(4)
    p1.metric(f"✅ {P_COMPLET}", n_complet)
    p2.metric("↩️ Entrants (chargés M-1)", n_entrant,
              help="Ex. chargé le 30/04, déchargé le 02/05 → présent dans l'export de mai.")
    p3.metric("↪️ Sortants (déchargés M+1)", n_sortant,
              help="Chargé dans le mois, déchargé le mois suivant.")
    p4.metric("🚫 Hors mois", n_hors)

    st.info(
        "ℹ️ **Attention aux bords de mois.** Les dossiers exclus créent des trous dans la "
        "chronologie du camion : le tronçon à vide qui les enjambe devient anormalement long. "
        "Le calcul les repère et les écarte (voir « Trou de chaîne » dans le détail à vide), "
        "mais les km correspondants ne seront **jamais** retrouvés. "
        "Pour se rapprocher du compteur, garder « Tout inclure »."
    )

    for _cat, _icone in [(P_ENTRANT, "↩️"), (P_SORTANT, "↪️"), (P_HORS, "🚫")]:
        _sub = df_all[df_all["_periode"] == _cat]
        if _sub.empty:
            continue
        _ca = _sub["ventes_totales"].sum()
        _statut = "conservés" if _cat in cats_gardees else "exclus"
        with st.expander(
            f"{_icone} {_cat} — {_sub['dossier'].nunique()} dossier(s) · {_ca:,.0f} € · **{_statut}**"
        ):
            st.dataframe(tableau_apercu(_sub), use_container_width=True, hide_index=True)

    df = df_all[df_all["_periode"].isin(cats_gardees)].copy()
    df_exclus = df_all[~df_all["_periode"].isin(cats_gardees)].copy()
    n_exclus = df_exclus["dossier"].nunique()

    if n_exclus:
        ca_exclu = df_exclus["ventes_totales"].sum()
        st.warning(
            f"🚫 **{n_exclus} dossier(s) exclus** du périmètre "
            f"({ca_exclu:,.0f} € de CA) — aucun km ni appel PTV ne sera calculé dessus."
        )
        with st.expander(f"📄 Voir les {n_exclus} dossiers exclus (n° + dates)", expanded=False):
            df_exc_tab = tableau_apercu(df_exclus)
            st.dataframe(df_exc_tab, use_container_width=True, hide_index=True, height=280)
            _rep = df_exclus.drop_duplicates(["vehicule", "dossier"])["_periode"].value_counts()
            st.caption("Motifs : " + " · ".join(f"{k} = {v}" for k, v in _rep.items()))
            st.download_button(
                "📥 Exporter les dossiers exclus (CSV)",
                data=df_exc_tab.to_csv(index=False, sep=";").encode("utf-8-sig"),
                file_name="Dossiers_exclus.csv",
                mime="text/csv", key="dl_exclus",
            )

    if df.empty:
        st.error("❌ Le périmètre retenu est vide. Élargis la règle d'attribution.")
        st.stop()

    periode_label = _libelle_mois(mois)
    st.divider()

    # ── KPIs globaux ──────────────────────────────────────────
    st.markdown(f"### 📊 Aperçu — {periode_label}")
    ca_total = df["ventes_totales"].sum()
    nb_dossiers = df["dossier"].nunique()
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("📁 Dossiers",       nb_dossiers, help=f"{len(df)} lignes d'export")
    k2.metric("🏢 Tractionnaires", df["tractionnaire"].nunique())
    k3.metric("🚛 Véhicules",      len(_opts(df["vehicule"])))
    k4.metric("💶 CA Total",       f"{ca_total:,.0f} €")
    k5.metric("📈 CA moy/dossier", f"{ca_total / nb_dossiers:,.0f} €" if nb_dossiers else "—")

    st.divider()

    # ══════════════════════════════════════════════════════════
    #  FILTRES TABLEAU DÉTAIL
    # ══════════════════════════════════════════════════════════
    st.markdown("### 📋 Tableau détail")

    tract_dispo = _opts(df["tractionnaire"])

    f1, f2 = st.columns(2)
    with f1:
        filtre_tract = st.multiselect(
            "🏢 Tractionnaire :", options=tract_dispo, default=[],
            placeholder="Tous les tractionnaires", key="flt_tract",
        )
    with f2:
        _base_veh = df[df["tractionnaire"].isin(filtre_tract)] if filtre_tract else df
        veh_dispo = _opts(_base_veh["vehicule"])
        filtre_veh = st.multiselect(
            "🚛 Véhicule :", options=veh_dispo, default=[],
            placeholder="Tous les camions", key="flt_veh",
        )

    f3, f4 = st.columns(2)
    with f3:
        _base_chf = _base_veh[_base_veh["vehicule"].isin(filtre_veh)] if filtre_veh else _base_veh
        chauff_dispo = _opts(_base_chf["chauffeur"])
        filtre_chauff = st.multiselect(
            "👤 Chauffeur :", options=chauff_dispo, default=[],
            placeholder="Tous les chauffeurs", key="flt_chauff",
        )
    with f4:
        statuts_dispo = _opts(df["statut"])
        filtre_statut = st.multiselect(
            "📄 Statut facturation :", options=statuts_dispo, default=[],
            placeholder="Tous", key="flt_statut",
        )

    df_display = df.copy()
    if filtre_tract:
        df_display = df_display[df_display["tractionnaire"].isin(filtre_tract)]
    if filtre_veh:
        df_display = df_display[df_display["vehicule"].isin(filtre_veh)]
    if filtre_chauff:
        df_display = df_display[df_display["chauffeur"].isin(filtre_chauff)]
    if filtre_statut:
        df_display = df_display[df_display["statut"].isin(filtre_statut)]

    cols_show = {
        "dossier": "N° Dossier", "tractionnaire": "Tractionnaire",
        "chauffeur": "Chauffeur", "vehicule": "Véhicule", "remorque": "Remorque",
        "client": "Client", "statut": "Statut", "_periode": "Période",
        "date_charg_fmt": "Date chargement", "localite_charg": "Chargement",
        "date_decharg_fmt": "Date déchargement", "localite_decharg": "Déchargement",
        "ventes_totales": "CA (€)",
    }
    df_table = df_display.sort_values("_date_charg_dt", na_position="last")
    df_table = df_table[[c for c in cols_show if c in df_table.columns]].rename(columns=cols_show)
    st.dataframe(df_table, use_container_width=True, height=380)

    st.divider()

    # ══════════════════════════════════════════════════════════
    #  CALCUL PTV
    # ══════════════════════════════════════════════════════════
    st.markdown("### 🗺️ Calcul KM via PTV")
    st.caption(
        "⚠️ Le périmètre PTV suit la règle de période ci-dessus, mais **pas** les filtres "
        "du tableau (statut/chauffeur) : les km à vide se calculent sur la chronologie "
        "complète de chaque camion — un filtre statut fausserait l'enchaînement."
    )

    pv1, pv2 = st.columns(2)
    with pv1:
        tract_ptv = st.multiselect(
            "🏢 Tractionnaires à calculer :", options=tract_dispo, default=[],
            placeholder="Tous les tractionnaires", key="ptv_tract",
        )
    with pv2:
        _base_ptv_veh = df[df["tractionnaire"].isin(tract_ptv)] if tract_ptv else df
        veh_ptv_dispo = _opts(_base_ptv_veh["vehicule"])
        veh_ptv = st.multiselect(
            "🚛 Véhicules à calculer :", options=veh_ptv_dispo, default=[],
            placeholder="Tous les camions du périmètre", key="ptv_veh",
        )

    o1, o2, o3, o4 = st.columns(4)
    with o1:
        calc_vide = st.toggle("⚡ Calculer les KM à vide", value=True, key="ptv_calc_vide")
    with o2:
        gap_max = st.number_input(
            "Gap max à vide (jours)", min_value=0, max_value=31,
            value=GAP_VIDE_DEFAUT, step=1, key="ptv_gap",
            help="Au-delà de ce délai entre un déchargement et le chargement suivant "
                 "du même camion, le tronçon à vide est considéré comme fictif.",
            disabled=not calc_vide,
        )
    with o3:
        km_vide_max = st.number_input(
            "KM max d'un tronçon à vide", min_value=50, max_value=2000,
            value=KM_VIDE_MAX_DEFAUT, step=50, key="ptv_kmmax",
            help="Un tronçon à vide plus long est un TROU DE CHAÎNE (dossier absent de "
                 "l'export), pas du repositionnement. Il est listé mais exclu du total. "
                 "Repère : sur avril 2026, 43 tronçons > 300 km pesaient 54 % du 'vide' calculé.",
            disabled=not calc_vide,
        )
    with o4:
        prorata = st.toggle(
            "🔀 Prorata dossiers multi-camions", value=True, key="ptv_prorata",
            help="Un dossier relayé par 2 tracteurs apparaît 2 fois dans l'export. "
                 "Activé : km et CA répartis à parts égales. Désactivé : comptés deux fois.",
        )

    df_ptv_scope = df.copy()
    if tract_ptv:
        df_ptv_scope = df_ptv_scope[df_ptv_scope["tractionnaire"].isin(tract_ptv)]
    if veh_ptv:
        df_ptv_scope = df_ptv_scope[df_ptv_scope["vehicule"].isin(veh_ptv)]

    selection_faite = bool(tract_ptv or veh_ptv)
    if not selection_faite:
        df_ptv_scope = pd.DataFrame(columns=df.columns)

    if selection_faite and not df_ptv_scope.empty:
        nb_dos = df_ptv_scope["dossier"].nunique()
        nb_veh = len(_opts(df_ptv_scope["vehicule"]))
        nb_pts = len(set(
            list(zip(df_ptv_scope["localite_charg"], df_ptv_scope["cp_charg"], df_ptv_scope["pays_charg"])) +
            list(zip(df_ptv_scope["localite_decharg"], df_ptv_scope["cp_decharg"], df_ptv_scope["pays_decharg"]))
        ))
        st.info(
            f"ℹ️ **{nb_dos} dossiers** ({len(df_ptv_scope)} lignes) · **{nb_veh} camion(s)** · "
            f"~{nb_pts} géocodages"
            + (" · km à vide activés" if calc_vide else " · km à vide désactivés")
        )

    btn_ptv = st.button("🚀 Lancer le calcul PTV",
                        disabled=(not selection_faite or df_ptv_scope.empty), type="primary")

    if btn_ptv and selection_faite and not df_ptv_scope.empty:
        progress_bar = st.progress(0.0)
        status_text  = st.empty()

        def _progress(msg, pct=None):
            status_text.text(msg)
            if pct is not None:
                progress_bar.progress(min(max(pct, 0.0), 1.0))

        try:
            df_ptv_result, df_vide_result, df_dos_result, stats = compute_km(
                df_ptv_scope, progress_cb=_progress,
                calc_vide=calc_vide, gap_max_jours=int(gap_max),
                km_vide_max=int(km_vide_max), prorata_multi_veh=bool(prorata),
            )
            st.session_state["df_ptv_result"]  = df_ptv_result
            st.session_state["df_vide_result"] = df_vide_result
            st.session_state["df_dos_result"]  = df_dos_result
            st.session_state["ptv_stats"]      = stats
            st.session_state["ptv_schema"]     = RESULT_SCHEMA
            progress_bar.progress(1.0)
            status_text.success("✅ Calcul PTV terminé !")
        except Exception as e:
            # Le calcul a échoué : on PURGE les anciens résultats, sinon le bloc
            # d'affichage ci-dessous ré-affiche le DataFrame du run précédent
            # (éventuellement produit par une version antérieure du code) et
            # plante sur une colonne absente.
            for k in ("df_ptv_result", "df_vide_result", "df_dos_result",
                      "ptv_stats", "ptv_schema"):
                st.session_state.pop(k, None)
            st.error(f"❌ Erreur calcul PTV : {type(e).__name__} — {e}")
            with st.expander("🐞 Détail technique"):
                import traceback
                st.code(traceback.format_exc())

    # ══════════════════════════════════════════════════════════
    #  RESULTATS PTV
    # ══════════════════════════════════════════════════════════
    # Résultat en session mais produit par une version antérieure du code :
    # on l'invalide au lieu de le rendre (c'était la cause du KeyError
    # 'ventes_retenues').
    if ("df_ptv_result" in st.session_state
            and st.session_state.get("ptv_schema") != RESULT_SCHEMA):
        for k in ("df_ptv_result", "df_vide_result", "df_dos_result",
                  "ptv_stats", "ptv_schema"):
            st.session_state.pop(k, None)
        st.warning(
            "♻️ Les résultats en mémoire venaient d'une version précédente de "
            "l'outil et ont été effacés. Relance le calcul PTV."
        )

    if "df_ptv_result" in st.session_state:
        df_r      = _normaliser_resultat(st.session_state["df_ptv_result"])
        df_vide_r = st.session_state.get("df_vide_result", pd.DataFrame())
        df_dos_r  = st.session_state.get("df_dos_result", pd.DataFrame())
        stats     = st.session_state.get("ptv_stats", {})

        st.divider()
        st.markdown("### 📈 Résultats KM")

        if stats:
            st.caption(
                f"🔧 {stats.get('appels_route', 0)} routes appelées · "
                f"{stats.get('routes_cache', 0)} servies par le cache · "
                f"{stats.get('vide_ignores', 0)} tronçons ignorés (gap/incohérence) · "
                f"{stats.get('vide_meme_lieu', 0)} à 0 km (même lieu)"
            )

        # ── Contrôles qualité ─────────────────────────────────
        c_alertes = []
        if stats.get("vide_trous", 0):
            c_alertes.append(
                f"🕳️ **{stats['vide_trous']} trou(s) de chaîne** détecté(s) "
                f"({stats['km_trous']:,.0f} km écartés). Ce ne sont pas des km à vide : "
                f"il manque des dossiers entre les deux (bord de mois, autre département "
                f"vente, trajet non facturé). Voir l'onglet « Détail KM à vide »."
            )
        if stats.get("dossiers_multi_veh", 0):
            mode = "répartis au prorata" if prorata else "**comptés en double**"
            c_alertes.append(
                f"🔀 **{stats['dossiers_multi_veh']} dossier(s) sur plusieurs camions** "
                f"(relais tracteur) — km et CA {mode}."
            )
        if stats.get("km_inter_livraisons", 0):
            c_alertes.append(
                f"📦 **{stats['km_inter_livraisons']:,.0f} km inter-livraisons** intégrés "
                f"(dossiers à plusieurs déchargements) — ignorés dans l'ancienne version."
            )
        if stats.get("echecs_geo"):
            c_alertes.append(
                f"🌍 **{len(stats['echecs_geo'])} point(s) non géocodés** : "
                + ", ".join(f"{v} {cp}" for v, cp, _ in stats["echecs_geo"][:8])
                + (" …" if len(stats["echecs_geo"]) > 8 else "")
            )
        if c_alertes:
            with st.expander("🔎 Contrôles qualité du calcul", expanded=True):
                for a in c_alertes:
                    st.markdown("- " + a)

        rf1, rf2, rf3 = st.columns([2, 2, 1])
        with rf1:
            res_tract_dispo = _opts(df_r["tractionnaire"])
            res_f_tract = st.multiselect(
                "🏢 Filtrer résultats — Tractionnaire :", options=res_tract_dispo,
                default=[], placeholder="Tous", key="res_flt_tract",
            )
        with rf2:
            _base_res = df_r[df_r["tractionnaire"].isin(res_f_tract)] if res_f_tract else df_r
            res_veh_dispo = _opts(_base_res["vehicule"])
            res_f_veh = st.multiselect(
                "🚛 Filtrer résultats — Véhicule :", options=res_veh_dispo,
                default=[], placeholder="Tous", key="res_flt_veh",
            )
        with rf3:
            st.write("")
            if st.button("🗑️ Effacer résultats"):
                for k in ("df_ptv_result", "df_vide_result", "df_dos_result", "ptv_stats"):
                    st.session_state.pop(k, None)
                st.rerun()

        df_rf = df_r.copy()
        if res_f_tract:
            df_rf = df_rf[df_rf["tractionnaire"].isin(res_f_tract)]
        if res_f_veh:
            df_rf = df_rf[df_rf["vehicule"].isin(res_f_veh)]

        df_vide_rf = df_vide_r.copy()
        if not df_vide_rf.empty:
            if res_f_tract:
                df_vide_rf = df_vide_rf[df_vide_rf["tractionnaire"].isin(res_f_tract)]
            if res_f_veh:
                df_vide_rf = df_vide_rf[df_vide_rf["vehicule"].isin(res_f_veh)]

        df_dos_rf = df_dos_r.copy()
        if not df_dos_rf.empty and res_f_veh:
            df_dos_rf = df_dos_rf[df_dos_rf["vehicule"].isin(res_f_veh)]

        km_charges  = df_rf["km_ptv"].sum()
        km_vide_sum = df_rf["km_vide"].sum()
        km_complet  = km_charges + km_vide_sum
        ca_sum      = df_rf["ventes_retenues"].sum()
        rent        = ca_sum / km_complet if km_complet > 0 else 0
        pct_vide    = km_vide_sum / km_complet * 100 if km_complet > 0 else 0
        dos_ok      = int(df_rf.loc[df_rf["_ligne_km"], "km_ptv"].notna().sum())
        dos_tot     = df_rf["dossier"].nunique()

        rk1, rk2, rk3, rk4, rk5, rk6 = st.columns(6)
        rk1.metric("📁 Dossiers calculés", f"{dos_ok} / {dos_tot}")
        rk2.metric("📏 KM Chargés",         f"{km_charges:,.0f} km")
        rk3.metric("⚡ KM À Vide",           f"{km_vide_sum:,.0f} km")
        rk4.metric("🔄 KM Complet",          f"{km_complet:,.0f} km")
        rk5.metric("% À Vide",               f"{pct_vide:.1f}%",
                   help="Repère : un taux de vide réaliste tourne entre 12 % et 20 %. "
                        "Au-delà de 25 %, vérifier les trous de chaîne.")
        rk6.metric("📈 Rentabilité",         f"{rent:.2f} €/km")

        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📋 Détail dossiers", "🏢 Résumé tractionnaires",
            "🚛 Résumé véhicules", "📦 Dossiers consolidés", "⚡ Détail KM à vide",
        ])

        with tab1:
            st.caption(
                "Les km d'un dossier sont portés par sa **première ligne** ; les lignes "
                "suivantes (livraisons complémentaires) sont à 0 pour éviter tout double "
                "comptage. Le km chargé inclut les trajets entre livraisons."
            )
            df_detail_show = df_rf.copy()
            df_detail_show["rentabilite"] = (
                df_detail_show["ventes_retenues"] / df_detail_show["km_total_complet"].replace(0, np.nan)
            ).round(2)
            df_detail_show = df_detail_show.sort_values(
                ["vehicule", "_date_charg_dt", "_ligne"], na_position="last"
            )
            cols_det = {
                "dossier": "N° Dossier", "tractionnaire": "Tractionnaire",
                "chauffeur": "Chauffeur", "vehicule": "Véhicule", "client": "Client",
                "_periode": "Période",
                "date_charg_fmt": "Date chargement", "localite_charg": "Chargement",
                "date_decharg_fmt": "Date déchargement", "localite_decharg": "Déchargement",
                "ventes_retenues": "CA retenu (€)", "km_ptv": "KM Chargé",
                "km_vide": "KM À Vide", "km_total_complet": "KM Complet",
                "rentabilite": "€/km", "_multi_veh": "Multi-camions",
            }
            df_det_tab = df_detail_show[[c for c in cols_det if c in df_detail_show.columns]].rename(columns=cols_det)
            st.dataframe(df_det_tab, use_container_width=True, height=400)

        with tab2:
            df_res_ptv_t = build_resume(df_rf, "tractionnaire", "Tractionnaire")
            st.dataframe(df_res_ptv_t, use_container_width=True)

        with tab3:
            df_veh_scope = df_rf[df_rf["vehicule"] != ""]
            if not df_veh_scope.empty:
                df_res_ptv_v = build_resume(df_veh_scope, "vehicule", "Véhicule")
                st.dataframe(df_res_ptv_v, use_container_width=True)
                st.caption(
                    "Comparaison au relevé compteur : coller ici les KM CB / KM CAN. "
                    "Un écart durable > 10 % par camion pointe soit des trous de chaîne, "
                    "soit des dossiers hors périmètre."
                )
            else:
                df_res_ptv_v = pd.DataFrame()
                st.info("Aucun véhicule renseigné sur les dossiers calculés.")

        with tab4:
            if not df_dos_rf.empty:
                dd = df_dos_rf.copy()
                dd["Chargements"] = dd["pickups"].apply(lambda L: " → ".join(p[0] for p in L))
                dd["Livraisons"]  = dd["drops"].apply(lambda L: " → ".join(p[0] for p in L))
                st.dataframe(
                    dd[["vehicule", "dossier", "Chargements", "Livraisons", "nb_lignes",
                        "km_charge", "km_inter_livraisons", "km_vide", "multi_vehicule"]]
                    .rename(columns={
                        "vehicule": "Véhicule", "dossier": "N° Dossier",
                        "nb_lignes": "Lignes export", "km_charge": "KM Chargé",
                        "km_inter_livraisons": "dont inter-livraisons",
                        "km_vide": "KM À Vide après", "multi_vehicule": "Multi-camions",
                    }),
                    use_container_width=True, height=400,
                )
            else:
                st.info("Aucun dossier consolidé.")

        with tab5:
            if not df_vide_rf.empty:
                trous = df_vide_rf["statut_leg"].astype(str).str.startswith("Trou").sum()
                nb_ignores = int(df_vide_rf["km_vide"].isna().sum())
                if trous:
                    st.error(
                        f"🕳️ {trous} tronçon(s) « Trou de chaîne » : trajet à vide "
                        f"invraisemblable → dossier manquant dans l'export. Exclus du total."
                    )
                if nb_ignores:
                    st.caption(f"⚠️ {nb_ignores} tronçon(s) non retenus — voir « Statut tronçon ».")
                st.dataframe(
                    df_vide_rf.rename(columns={
                        "vehicule": "Véhicule", "tractionnaire": "Tractionnaire",
                        "dossier_depart": "Dossier départ", "dossier_arrivee": "Dossier arrivée",
                        "ville_depart": "Ville départ", "ville_arrivee": "Ville arrivée",
                        "date_depart": "Date", "gap_jours": "Gap (j)",
                        "statut_leg": "Statut tronçon", "km_vide": "KM à vide retenu",
                        "km_brut": "KM brut PTV",
                    }),
                    use_container_width=True, height=400,
                )
            else:
                st.info("Aucun km à vide calculé.")

        st.divider()
        excel_bytes = export_excel(df_rf, df_res_ptv_t, df_res_ptv_v, df_vide_rf, df_dos_rf)
        if excel_bytes:
            st.download_button(
                label="📥 Télécharger le rapport Excel",
                data=excel_bytes,
                file_name=f"Rapport_Tractionnaires_KM_{periode_label.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

else:
    st.markdown("""
    #### Comment utiliser cet outil

    1. **Chargez l'export tractionnaires** (.xlsx)
    2. Vérifiez la **période d'analyse** : le mois est déduit des déchargements.
    3. Filtrez le tableau par **tractionnaire, véhicule, chauffeur ou statut**
    4. Sélectionnez les **tractionnaires et/ou camions** à calculer, puis lancez le PTV
    5. Contrôlez les alertes qualité, puis **téléchargez le rapport Excel**

    ##### Ce que corrige la v2
    - **Un dossier = un itinéraire**, pas une ligne. Les dossiers à plusieurs
      chargements / livraisons sont consolidés avant tout calcul.
    - Les km sont posés sur **une seule ligne par dossier** → plus de double
      comptage dans les résumés.
    - Le tronçon à vide part de la **dernière livraison** du dossier et arrive au
      **premier chargement** du suivant (plus de tronçon d'un dossier vers lui-même).
    - Les trajets **entre deux livraisons** d'un même dossier sont enfin comptés.
    - Un tronçon à vide trop long est un **trou de chaîne**, pas du vide : il est
      signalé et exclu.
    - Un dossier relayé par **deux tracteurs** n'est plus compté deux fois.

    ##### Repères de contrôle
    - Taux de vide réaliste : **12 % à 20 %**. Au-delà de 25 %, chercher les trous.
    - Comparer au **KM CB**, pas au KM CAN : le CAN inclut lavages, approches,
      décrochages et km privés, absents des dossiers.

    > ⚙️ La clé PTV doit être configurée dans `.env` (`PTV_API_KEY`).
    """)
