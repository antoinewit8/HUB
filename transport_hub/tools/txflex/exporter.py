import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
 
def export_to_excel(empty_km_list, friday_data, output_path="outputs/rapport_camion.xlsx"):
 
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
 
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
 
        # ========== FEUILLE 1 : KM À VIDE ==========
        if empty_km_list:
            df_vide = pd.DataFrame(empty_km_list)
            # On garde uniquement les colonnes utiles (sans 'alerte' et 'raison')
            df_vide = df_vide[["date_depart", "date_arrivee", "ville_depart", "ville_arrivee", "km_vide"]]
            df_vide.columns = ["Date départ", "Date arrivée", "Ville départ", "Ville arrivée", "KM à vide"]
        else:
            df_vide = pd.DataFrame(columns=["Date départ", "Date arrivée", "Ville départ", "Ville arrivée", "KM à vide"])
 
        df_vide.to_excel(writer, sheet_name="KM à vide", index=False)
        ws1 = writer.sheets["KM à vide"]
 
        # Header rouge
        header_fill = PatternFill("solid", fgColor="C00000")
        header_font = Font(bold=True, color="FFFFFF")
 
        for col in range(1, 6):
            cell = ws1.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
 
        # Colorier les lignes selon les km à vide
        orange_fill = PatternFill("solid", fgColor="FFD966")
        red_fill    = PatternFill("solid", fgColor="FF7070")
 
        for row_idx, entry in enumerate(empty_km_list, start=2):
            km = entry["km_vide"]
            if km > 300:
                fill = red_fill
            elif km > 200:
                fill = orange_fill
            else:
                fill = None
 
            if fill:
                for col in range(1, 6):
                    ws1.cell(row=row_idx, column=col).fill = fill
 
        # Largeur colonnes
        for col in range(1, 6):
            ws1.column_dimensions[get_column_letter(col)].width = 30
 
        # Ligne total
        total    = sum(e["km_vide"] for e in empty_km_list)
        last_row = len(empty_km_list) + 2
        total_cell_label = ws1.cell(row=last_row, column=4)
        total_cell_value = ws1.cell(row=last_row, column=5)
        total_cell_label.value = "TOTAL KM À VIDE"
        total_cell_label.font  = Font(bold=True)
        total_cell_value.value = total
        total_cell_value.font  = Font(bold=True)
 
        # ========== FEUILLE 2 : VENDREDIS ==========
        rows_friday = []  # BUG CORRIGÉ : indentation incorrecte (13 espaces → 8)
        for f in friday_data:
            rows_friday.append({
                "Date":           f["date"],
                "KM parcourus":   f["km_parcourus"],
                "KM à vide":      f["km_vide"],
                "Nb activités":   f["nb_activites"],
                "Heure de fin":   f["heure_fin"],
                "Villes visitées": " → ".join(f["villes_visitees"]),
                "Statut":         "🔴 ALERTE" if f["alerte"] else "✅ Normal",
                "Raisons":        " | ".join(f["raisons"]) if f["raisons"] else ""
            })
 
        df_friday = pd.DataFrame(rows_friday)
        df_friday.to_excel(writer, sheet_name="Vendredis", index=False)
 
        ws2 = writer.sheets["Vendredis"]
        nb_cols = len(df_friday.columns)
 
        # Header bleu marine — couvre TOUTES les colonnes (était limité à 4 avant)
        for col in range(1, nb_cols + 1):
            cell = ws2.cell(row=1, column=col)
            cell.fill      = PatternFill("solid", fgColor="1F3864")
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
 
        # Colorier alertes — couvre TOUTES les colonnes
        for row_idx, f in enumerate(friday_data, start=2):
            if f["alerte"]:
                for col in range(1, nb_cols + 1):
                    ws2.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="FF7070")
 
        # Largeur colonnes
        col_widths = {
            "A": 15,   # Date
            "B": 15,   # KM parcourus
            "C": 15,   # KM à vide
            "D": 14,   # Nb activités
            "E": 12,   # Heure de fin
            "F": 50,   # Villes visitées
            "G": 14,   # Statut
            "H": 60,   # Raisons
        }
        for col_letter, width in col_widths.items():
            ws2.column_dimensions[col_letter].width = width
 
    print(f"\n✅ Rapport exporté : {output_path}")
