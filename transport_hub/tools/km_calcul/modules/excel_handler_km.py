import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ============================================================
# CONSTANTES
# ============================================================

PAYS_MAP = {
    "FR": "France", "BE": "Belgium", "DE": "Germany", "NL": "Netherlands",
    "LU": "Luxembourg", "IT": "Italy", "ES": "Spain", "PT": "Portugal",
    "GB": "United Kingdom", "CH": "Switzerland", "AT": "Austria",
    "PL": "Poland", "CZ": "Czech Republic", "HU": "Hungary",
    "RO": "Romania", "BG": "Bulgaria", "SK": "Slovakia", "SI": "Slovenia",
    "HR": "Croatia", "DK": "Denmark", "SE": "Sweden", "NO": "Norway",
    "FI": "Finland", "IE": "Ireland", "GR": "Greece",
    "F": "France", "B": "Belgium", "D": "Germany", "I": "Italy",
    "E": "Spain", "L": "Luxembourg", "A": "Austria", "P": "Portugal",
}

CP_LENGTHS = {
    "FR": 5, "F": 5, "BE": 4, "B": 4, "DE": 5, "D": 5,
    "IT": 5, "I": 5, "ES": 5, "E": 5, "PL": 5, "CZ": 5,
    "HR": 5, "SK": 5, "GR": 5, "SE": 5, "FI": 5, "NL": 4,
    "AT": 4, "A": 4, "CH": 4, "HU": 4, "DK": 4, "SI": 4,
    "NO": 4, "LU": 4, "L": 4, "PT": 7, "P": 7, "RO": 6,
}

CITY_CORRECTIONS = {
    "Basse Indre":    ("Indre", "44610", "FR"),
    "Saint Herblain": ("Saint-Herblain", "44800", "FR"),
    "Montataire":     ("Montataire", "60160", "FR"),
}

ZONE_CORRECTIONS = {
    "02000, Italy": "Rieti, 02100, Italy",
    "06000, Italy": "Perugia, 06100, Italy",
    "10000, Italy": "Torino, 10100, Italy",
    "16000, Italy": "Genova, 16100, Italy",
    "20000, Italy": "Milano, 20100, Italy",
    "21000, Italy": "Varese, 21100, Italy",
    "23000, Italy": "Sondrio, 23100, Italy",
    "24000, Italy": "Bergamo, 24100, Italy",
    "25000, Italy": "Brescia, 25100, Italy",
    "30000, Italy": "Mestre, 30170, Italy",
    "31000, Italy": "Treviso, 31100, Italy",
    "33000, Italy": "Udine, 33100, Italy",
    "36000, Italy": "Vicenza, 36100, Italy",
    "00000, Italy": "Roma, 00133, Italy",
    "50000, Italy": "Firenze, 50127, Italy",
    "80000, Italy": "Napoli, 80146, Italy",
    "6000, Netherlands": "Weert, 6000, Netherlands",
    "6200, Netherlands": "Maastricht, 6221, Netherlands",
    "6600, Netherlands": "Wijchen, 6602, Netherlands",
    "1000, Netherlands": "Amsterdam, 1043, Netherlands",
    "3000, Netherlands": "Rotterdam, 3089, Netherlands",
    "8800, Belgium": "Roeselare, 8800, Belgium",
    "1000, Belgium": "Bruxelles, 1120, Belgium",
    "2000, Belgium": "Antwerpen, 2030, Belgium",
    "4000, Belgium": "Liège, 4020, Belgium",
    "50000, Germany": "Köln, 50769, Germany",
    "58000, Germany": "Hagen, 58099, Germany",
    "59000, Germany": "Hamm, 59067, Germany",
    "10000, Germany": "Berlin, 13405, Germany",
    "20000, Germany": "Hamburg, 21129, Germany",
    "80000, Germany": "München, 80939, Germany",
    "60000, Germany": "Frankfurt, 60549, Germany",
    "75000, France": "Paris, 75012, France",
    "69000, France": "Vénissieux, 69200, France",
    "13000, France": "Marseille, 13015, France",
    "33000, France": "Bordeaux, 33300, France",
    "59000, France": "Lille, 59160, France",
    "67000, France": "Strasbourg, 67100, France",
    "28000, Spain": "Madrid, 28052, Spain",
    "08000, Spain": "Barcelona, 08040, Spain",
    "46000, Spain": "Valencia, 46024, Spain",
    "01000, Spain": "Vitoria-Gasteiz, 01015, Spain",
    "03000, Spain": "Alicante, 03008, Spain",
    "06000, Spain": "Badajoz, 06006, Spain",
    "09000, Spain": "Burgos, 09007, Spain",
    "30000, Spain": "Murcia, 30169, Spain",
    "36000, Spain": "Pontevedra, 36158, Spain",
    "1000, Austria": "Wien, 1110, Austria",
    "3400, Luxembourg": "Dudelange, 3400, Luxembourg",
    "1000, Luxembourg": "Luxembourg, 1000, Luxembourg",
    "15000, Bulgaria": "Sofia, 1528, Bulgaria",
}

GPS_FIXES_ORIGIN = {
    "rumbek":  "Rumbeke, 8800, Belgium",
    "rumbeke": "Rumbeke, 8800, Belgium",
    "indre":   "Indre, 44610, France",
}

# ============================================================
# SYNONYMES POUR DÉTECTION INTELLIGENTE
# ============================================================

# Chaque rôle a une liste de synonymes (en minuscule)
ORIG_PAYS_KEYWORDS = [
    "pays", "country", "country of origin", "orig cntry",
    "pays départ", "pays depart", "pays origine",
]
ORIG_CP_KEYWORDS = [
    "code postal", "cp", "postal code", "postal code origin",
    "orig reg", "cp départ", "cp depart", "code postal origine",
]
ORIG_VILLE_KEYWORDS = [
    "ville", "city", "city of origin", "orig zone txt",
    "localité", "localite", "origin", "depart", "départ",
    "origine", "ville départ", "ville depart",
]
DEST_PAYS_KEYWORDS = [
    "pays", "country", "country of destination", "dest cntry",
    "pays destination", "dest pays",
]
DEST_CP_KEYWORDS = [
    "département", "departement", "code postal", "cp",
    "postal code", "postal code destination", "dest reg",
    "cp destination", "cp dech", "cp déchargement",
]
DEST_VILLE_KEYWORDS = [
    "ville2", "ville", "city", "city of destination",
    "dest zone txt", "destination", "ville destination",
    "ville dest",
]

# Mots-clés des lignes de groupe (ligne au-dessus des headers)
ORIGIN_GROUP_KEYWORDS = ["depart", "départ", "origin", "origine", "chargement", "loading"]
DEST_GROUP_KEYWORDS   = ["destination", "dest", "arrivée", "arrivee", "livraison", "delivery", "unloading"]

# Mots-clés d'en-tête parasite
HEADER_KEYWORDS = {
    "depart", "départ", "origin", "pays", "country", "cp", "code postal",
    "postal code", "destination", "dest", "city", "ville", "zone",
    "country of destination", "dest cntry", "dest zone txt", "dest reg",
    "postal code destination", "city of destination",
}


# ============================================================
# UTILITAIRES
# ============================================================

def normalize(val):
    if val is None:
        return ""
    return re.sub(r'\s+', ' ', str(val).strip().lower())


def pad_postal_code(cp, country_prefix):
    if not cp:
        return ""
    target_len = CP_LENGTHS.get(country_prefix.upper(), 5)
    if target_len == 0:
        return cp
    if len(cp) < target_len:
        cp = cp + "0" * (target_len - len(cp))
    return cp


def is_header_row_data(values):
    """Retourne True si la ligne ressemble à une ligne d'en-tête parasite."""
    return any(str(v).strip().lower() in HEADER_KEYWORDS for v in values if v)


def find_first_empty_column(ws, header_row=1, data_start=2):
    col = 1
    while col < 100:
        if ws.cell(row=header_row, column=col).value is not None:
            col += 1
            continue
        has_data = False
        for row in range(data_start, min(data_start + 10, ws.max_row + 1)):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if not has_data:
            return col
        col += 1
    return col


# ============================================================
# DÉTECTION INTELLIGENTE DES EN-TÊTES
# ============================================================

def detect_header_row(ws, max_scan=15):
    """
    Parcourt les premières lignes pour trouver celle qui contient
    le plus de headers reconnus parmi tous les synonymes.
    """
    all_keywords = set()
    for kw_list in [ORIG_PAYS_KEYWORDS, ORIG_CP_KEYWORDS, ORIG_VILLE_KEYWORDS,
                    DEST_PAYS_KEYWORDS, DEST_CP_KEYWORDS, DEST_VILLE_KEYWORDS]:
        for kw in kw_list:
            all_keywords.add(kw)

    best_row   = None
    best_score = 0

    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        score = 0
        for col in range(1, ws.max_column + 1):
            val = normalize(ws.cell(row=row_idx, column=col).value)
            if val in all_keywords:
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row if best_score >= 2 else None


def detect_groups(ws, header_row):
    """
    Regarde la ligne juste au-dessus des headers pour détecter
    les groupes DEPART / Destination via cellules fusionnées ou non.
    Retourne un dict {col_index: "origin" | "dest"}.
    """
    if header_row <= 1:
        return {}

    group_row = header_row - 1
    groups = {}
    current_group = None

    # Gérer les cellules fusionnées : mapper chaque colonne à sa valeur de merge
    merge_map = {}
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row <= group_row <= merge_range.max_row:
            val = ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[c] = val

    for col in range(1, ws.max_column + 1):
        val = normalize(merge_map.get(col, ws.cell(row=group_row, column=col).value))

        if any(k in val for k in ORIGIN_GROUP_KEYWORDS):
            current_group = "origin"
        elif any(k in val for k in DEST_GROUP_KEYWORDS):
            current_group = "dest"
        elif val and val not in ("", " "):
            # Un texte non reconnu = on ne sait plus dans quel groupe
            current_group = None

        if current_group:
            groups[col] = current_group

    return groups


def map_columns(ws):
    """
    Détecte la ligne d'en-têtes, les groupes, et mappe chaque rôle
    à un numéro de colonne. Gère les doublons (2x "Pays") via les groupes.
    """
    header_row = detect_header_row(ws)
    if not header_row:
        return None, None, None

    groups = detect_groups(ws, header_row)

    # Lire tous les headers de la ligne détectée
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = normalize(ws.cell(row=header_row, column=col).value)
        if val:
            headers[col] = val

    # Définition des rôles à mapper
    ROLES = {
        "orig_pays":  (ORIG_PAYS_KEYWORDS,  "origin"),
        "orig_cp":    (ORIG_CP_KEYWORDS,     "origin"),
        "orig_ville": (ORIG_VILLE_KEYWORDS,  "origin"),
        "dest_pays":  (DEST_PAYS_KEYWORDS,   "dest"),
        "dest_cp":    (DEST_CP_KEYWORDS,     "dest"),
        "dest_ville": (DEST_VILLE_KEYWORDS,  "dest"),
    }

    mapping = {}
    used_cols = set()

    for role, (keywords, expected_side) in ROLES.items():
        # Passe 1 : avec groupe (résout les doublons "pays", "ville")
        for col, header_val in headers.items():
            if col in used_cols:
                continue
            col_side = groups.get(col)
            if col_side == expected_side and header_val in keywords:
                mapping[role] = col
                used_cols.add(col)
                break

        # Passe 2 : sans groupe (fallback pour fichiers sans ligne de groupe)
        if role not in mapping:
            for col, header_val in headers.items():
                if col in used_cols:
                    continue
                if header_val in keywords:
                    mapping[role] = col
                    used_cols.add(col)
                    break

    data_start_row = header_row + 1
    return mapping, header_row, data_start_row


# ============================================================
# PARSEURS
# ============================================================

def parse_origin(origin_str):
    if not origin_str:
        return ""
    origin_str = str(origin_str).strip()

    last_word = origin_str.split()[-1].lower() if origin_str.split() else ""
    if last_word in GPS_FIXES_ORIGIN:
        return GPS_FIXES_ORIGIN[last_word]

    if "(" in origin_str and ")" in origin_str:
        ville = origin_str[:origin_str.index("(")].strip()
        code  = origin_str[origin_str.index("(") + 1:origin_str.index(")")].strip()

        pays_prefix = ""
        cp = ""
        for i, c in enumerate(code):
            if c.isdigit():
                pays_prefix = code[:i]
                cp = code[i:]
                break

        cp   = pad_postal_code(cp, pays_prefix)
        pays = PAYS_MAP.get(pays_prefix.upper(), pays_prefix)

        if ville in CITY_CORRECTIONS:
            ville_corr, cp_corr, pays_corr = CITY_CORRECTIONS[ville]
            pays_corr_full = PAYS_MAP.get(pays_corr.upper(), pays_corr)
            return f"{ville_corr}, {cp_corr}, {pays_corr_full}"

        return f"{ville}, {cp}, {pays}"

    return origin_str


def parse_origin_from_parts(city, postal_code, country):
    country     = str(country     or "").strip()
    city        = str(city        or "").strip()
    postal_code = str(postal_code or "").strip()

    pays_full = PAYS_MAP.get(country.upper(), country)
    cp_num    = pad_postal_code(postal_code, country)

    if city:
        return f"{city}, {cp_num}, {pays_full}"
    elif cp_num:
        return f"{cp_num}, {pays_full}"
    else:
        return pays_full


def parse_destination(city, postal_code, country):
    country     = str(country     or "").strip()
    city        = str(city        or "").strip()
    postal_code = str(postal_code or "").strip()

    cp_pays = ""
    cp_num  = ""
    if "-" in postal_code:
        parts   = postal_code.split("-", 1)
        cp_pays = parts[0].strip()
        cp_num  = parts[1].strip()
    else:
        cp_num = postal_code

    pays_full = PAYS_MAP.get(country.upper(), country)
    if not pays_full or pays_full == country:
        pays_full = PAYS_MAP.get(cp_pays.upper(), pays_full or cp_pays)

    pays_prefix = cp_pays if cp_pays else country
    cp_num = pad_postal_code(cp_num, pays_prefix)

    if city.lower() in ("all cities", "all", ""):
        base_zone = f"{cp_num}, {pays_full}"
        return ZONE_CORRECTIONS.get(base_zone, base_zone)
    else:
        return f"{city}, {cp_num}, {pays_full}"


# ============================================================
# LECTURE EXCEL — DÉTECTION INTELLIGENTE
# ============================================================

def read_all_sheets(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── Détection intelligente ────────────────────────────────────────
        result = map_columns(ws)
        if result[0] is None:
            print(f"  ⚠️  Feuille '{sheet_name}' ignorée [headers non détectés]")
            continue

        mapping, header_row, data_start_row = result

        # Vérification minimale
        has_orig = any(k.startswith("orig_") for k in mapping)
        has_dest = any(k.startswith("dest_") for k in mapping)

        if not has_orig or not has_dest:
            print(f"  ⚠️  Feuille '{sheet_name}' ignorée [orig={has_orig}, dest={has_dest}]")
            print(f"       Mapping: {mapping}")
            continue

        print(f"  ✅ Feuille '{sheet_name}' — headers ligne {header_row}")
        print(f"       Mapping: {mapping}")

        # ── Lecture des routes ────────────────────────────────────────────
        routes       = []
        lignes_vides = 0

        for row in range(data_start_row, ws.max_row + 6):
            # Lecture origine
            orig_pays_raw = ws.cell(row=row, column=mapping["orig_pays"]).value   if "orig_pays"  in mapping else None
            orig_cp_raw   = ws.cell(row=row, column=mapping["orig_cp"]).value     if "orig_cp"    in mapping else None
            orig_city_raw = ws.cell(row=row, column=mapping["orig_ville"]).value  if "orig_ville" in mapping else None

            # Lecture destination
            dest_pays_raw = ws.cell(row=row, column=mapping["dest_pays"]).value   if "dest_pays"  in mapping else None
            dest_cp_raw   = ws.cell(row=row, column=mapping["dest_cp"]).value     if "dest_cp"    in mapping else None
            dest_city_raw = ws.cell(row=row, column=mapping["dest_ville"]).value  if "dest_ville" in mapping else None

            # Ligne vide ?
            if not orig_pays_raw and not dest_pays_raw and not orig_city_raw and not dest_city_raw:
                lignes_vides += 1
                if lignes_vides >= 3:
                    break
                continue

            lignes_vides = 0

            # En-tête parasite ?
            all_vals = [orig_pays_raw, orig_cp_raw, orig_city_raw,
                        dest_pays_raw, dest_cp_raw, dest_city_raw]
            if is_header_row_data(all_vals):
                print(f"  ⏭️  Ligne {row} ignorée (en-tête parasite)")
                continue

            # ── Parse origine ─────────────────────────────────────────────
            origin = parse_origin_from_parts(
                str(orig_city_raw or "").strip(),
                str(orig_cp_raw   or "").strip(),
                str(orig_pays_raw or "").strip(),
            )

            # ── Parse destination ─────────────────────────────────────────
            dest = parse_destination(
                dest_city_raw,
                dest_cp_raw,
                dest_pays_raw,
            )

            if not origin or not dest:
                continue

            label = (f"{orig_city_raw or orig_pays_raw} ({orig_cp_raw or ''}) "
                     f"→ {dest_city_raw or 'Zone'} ({dest_cp_raw or ''})")

            routes.append({
                "row":    row,
                "origin": origin,
                "dest":   dest,
                "label":  label,
            })

        if routes:
            sheets_data[sheet_name] = (ws, routes)
            print(f"     → {len(routes)} route(s) extraite(s)")
        else:
            print(f"  ⚠️  Feuille '{sheet_name}' : aucune route valide")

    return wb, sheets_data


# ============================================================
# ÉCRITURE RÉSULTATS
# ============================================================

def unmerge_and_write(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    for merge_range in list(ws.merged_cells.ranges):
        if (merge_range.min_row <= row <= merge_range.max_row and
                merge_range.min_col <= col <= merge_range.max_col):
            ws.unmerge_cells(str(merge_range))
            break
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    return cell


def find_safe_col(ws, header_row, start_col):
    col = start_col
    while True:
        is_merged = False
        for merge_range in ws.merged_cells.ranges:
            if (merge_range.min_row <= header_row <= merge_range.max_row and
                    merge_range.min_col <= col <= merge_range.max_col):
                col = merge_range.max_col + 1
                is_merged = True
                break
        if not is_merged:
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                return col
            col += 1


def write_km_results(ws, results, calculer_peage=False):
    # Détecte le header_row dynamiquement
    detected = detect_header_row(ws)
    header_row = detected if detected else 1

    raw_start  = find_first_empty_column(ws, header_row=header_row)
    safe_start = find_safe_col(ws, header_row, raw_start)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    headers = ["KM PTV", "Carte PTV"]
    if calculer_peage:
        headers.append("Péage (€)")

    col_map = {}
    for i, h in enumerate(headers):
        target_col = find_safe_col(ws, header_row, safe_start + i)
        unmerge_and_write(ws, header_row, target_col, h,
                          font=header_font, fill=header_fill,
                          alignment=header_align, border=border)
        col_map[h] = target_col

    link_font = Font(color="0563C1", underline="single")

    for r in results:
        if r is None:
            continue
        row  = r["row"]
        data = r.get("data")
        if not data:
            continue

        unmerge_and_write(ws, row, col_map["KM PTV"], data["km"], border=border)

        carte_url = data.get("carte_url", "")
        if carte_url:
            cell = unmerge_and_write(
                ws, row, col_map["Carte PTV"],
                "🗺️ Voir carte",
                font=link_font, border=border,
            )
            cell.hyperlink = carte_url
            cell.style = "Hyperlink"
        else:
            unmerge_and_write(ws, row, col_map["Carte PTV"], "", border=border)

        if calculer_peage:
            unmerge_and_write(
                ws, row, col_map["Péage (€)"],
                data.get("prix_peage", 0.0),
                border=border,
            )
